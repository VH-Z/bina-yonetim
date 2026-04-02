"""
╔══════════════════════════════════════════════════════════════╗
║   BİNA YÖNETİM SİSTEMİ  v3.0  —  Python / Tkinter          ║
║   Koyu Tema  |  Sürüklenebilir Çerçeveler                   ║
║   Aylık Özel Aidat  |  Ödeme Düzenleme  |  Ay Seçimi        ║
╚══════════════════════════════════════════════════════════════╝
Gereksinim: Python 3.8+  (tkinter dahil)
Veri: ~/bina_yonetim_v3.json
"""

import tkinter as tk
from tkinter import ttk, messagebox
import json, os, datetime, webbrowser, tempfile, copy
from pathlib import Path

# ─── RENKLER ─────────────────────────────────────────────────────────────────
T = {
    "bg"        : "#0f1117",
    "bg2"       : "#161b27",
    "bg3"       : "#1e2535",
    "bg4"       : "#252d3d",
    "border"    : "#2a3550",
    "gold"      : "#c9a84c",
    "gold2"     : "#e8c97a",
    "text"      : "#e8eaf0",
    "text2"     : "#8a9ab5",
    "text3"     : "#4a5a7a",
    "green"     : "#2ecc71",
    "green_bg"  : "#1a3a2a",
    "red"       : "#e74c3c",
    "red_bg"    : "#3a1a1a",
    "orange"    : "#f39c12",
    "orange_bg" : "#3a2a0a",
    "blue"      : "#3498db",
    "blue_bg"   : "#1a2a3a",
    "purple"    : "#9b59b6",
    "entry_bg"  : "#1a2235",
    "sel"       : "#2a3a55",
}

DATA_DIR  = Path.home() / "bina_yonetim_data"   # Klasör
FILE_AYAR = DATA_DIR / "bina_ayarlar.json"        # Ayarlar
FILE_VER  = DATA_DIR / "bina_veriler.json"         # Gelir/Gider/Alacaklılar
# Daireler: DATA_DIR / "daire_{no}.json"  (dinamik)

# Geriye dönük uyumluluk için eski tek-dosya yolu
DATA_FILE_LEGACY = Path.home() / "bina_yonetim_v3.json"

def _data_dir():
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    return DATA_DIR

def _daire_dosya(no):
    return DATA_DIR / f"daire_{no:03d}.json"

# ─── VERİ YARDIMCILARI ───────────────────────────────────────────────────────
def load_data():
    _data_dir()

    # Eski tek-dosya varsa migrate et
    if DATA_FILE_LEGACY.exists() and not FILE_AYAR.exists():
        _migrate_legacy()

    if not FILE_AYAR.exists():
        return None

    # 1. Ayarlar dosyası
    with open(FILE_AYAR, "r", encoding="utf-8") as f:
        ayar = json.load(f)

    # 2. Daire dosyaları
    daireler = []
    for dosya in sorted(DATA_DIR.glob("daire_*.json")):
        with open(dosya, "r", encoding="utf-8") as f:
            daireler.append(json.load(f))
    daireler.sort(key=lambda d: d["no"])

    # Daire dosyaları yoksa ayar dosyasındaki daire sayısına göre boş oluştur
    if not daireler and ayar.get("bina", {}).get("daire", 0) > 0:
        for i in range(1, ayar["bina"]["daire"] + 1):
            daireler.append({"no": i, "isim": "", "tel": "", "email": "",
                             "borc": 0.0, "faiz": 0.0, "son_odeme": "",
                             "odemeler": []})

    # 3. Genel veriler dosyası
    if FILE_VER.exists():
        with open(FILE_VER, "r", encoding="utf-8") as f:
            veriler = json.load(f)
    else:
        veriler = {"gelirler": [], "giderler": [], "alacaklilar": []}

    # Birleştir
    data = {
        "bina"          : ayar.get("bina", {}),
        "daireler"      : daireler,
        "odemeler"      : [],            # daire dosyalarından topla
        "gelirler"      : veriler.get("gelirler", []),
        "giderler"      : veriler.get("giderler", []),
        "alacaklilar"   : veriler.get("alacaklilar", []),
        "ortak_giderler": veriler.get("ortak_giderler", []),
    }

    # Her daireden ödemeleri ana listeye topla
    for d in daireler:
        for o in d.get("odemeler", []):
            data["odemeler"].append(o)

    # Migrate kontrolleri
    if "alacaklilar" not in data:
        data["alacaklilar"] = []
    return data

def _migrate_legacy():
    """Eski tek JSON dosyasını yeni çok-dosya yapısına taşı."""
    _data_dir()
    with open(DATA_FILE_LEGACY, "r", encoding="utf-8") as f:
        eski = json.load(f)

    # Ayarlar
    ayar = {"bina": eski.get("bina", {})}
    with open(FILE_AYAR, "w", encoding="utf-8") as f:
        json.dump(ayar, f, ensure_ascii=False, indent=2)

    # Daireler (ödemeler dahil)
    odemeler = eski.get("odemeler", [])
    for d in eski.get("daireler", []):
        d_odemeler = [o for o in odemeler if o["daireNo"] == d["no"]]
        d_data = dict(d)
        d_data["odemeler"] = d_odemeler
        with open(_daire_dosya(d["no"]), "w", encoding="utf-8") as f:
            json.dump(d_data, f, ensure_ascii=False, indent=2)

    # Genel veriler
    veriler = {
        "gelirler"   : eski.get("gelirler", []),
        "giderler"   : eski.get("giderler", []),
        "alacaklilar": eski.get("alacaklilar", []),
    }
    with open(FILE_VER, "w", encoding="utf-8") as f:
        json.dump(veriler, f, ensure_ascii=False, indent=2)

    # Eski dosyayı yeniden adlandır (silmiyoruz)
    DATA_FILE_LEGACY.rename(DATA_FILE_LEGACY.with_suffix(".json.bak"))

def _yuvarla_tutarlar(data):
    """Kaydedilmeden önce tüm para alanlarını 2 ondalığa yuvarla."""
    b = data["bina"]
    for k in ("aidat", "faiz"):
        if k in b:
            b[k] = round(float(b[k]), 2)
    for ozel in b.get("ozel_aidatlar", {}).items():
        b["ozel_aidatlar"][ozel[0]] = round(float(ozel[1]), 2)
    for d in data.get("daireler", []):
        d["borc"]  = round(float(d.get("borc",  0)), 2)
        d["faiz"]  = round(float(d.get("faiz",  0)), 2)
    for o in data.get("odemeler", []):
        o["tutar"] = round(float(o.get("tutar", 0)), 2)
    for g in data.get("gelirler", []):
        g["tutar"] = round(float(g.get("tutar", 0)), 2)
    for g in data.get("giderler", []):
        g["tutar"] = round(float(g.get("tutar", 0)), 2)
    for a in data.get("alacaklilar", []):
        a["tutar"] = round(float(a.get("tutar", 0)), 2)

def save_data(data):
    """Veriyi 3 ayrı alana kaydet."""
    _yuvarla_tutarlar(data)
    _data_dir()

    # 1. Ayarlar dosyası
    ayar = {"bina": data["bina"]}
    with open(FILE_AYAR, "w", encoding="utf-8") as f:
        json.dump(ayar, f, ensure_ascii=False, indent=2)

    # 2. Her daire ayrı dosyaya
    odemeler_by_daire = {}
    for o in data.get("odemeler", []):
        odemeler_by_daire.setdefault(o["daireNo"], []).append(o)

    for d in data.get("daireler", []):
        d_data = dict(d)
        d_data["odemeler"] = odemeler_by_daire.get(d["no"], [])
        # ozel_aidatlar (daireye özgü eksik ödeme devri) zaten d içinde yer alıyor
        with open(_daire_dosya(d["no"]), "w", encoding="utf-8") as f:
            json.dump(d_data, f, ensure_ascii=False, indent=2)

    # 3. Genel veriler
    veriler = {
        "gelirler"      : data.get("gelirler", []),
        "giderler"      : data.get("giderler", []),
        "alacaklilar"   : data.get("alacaklilar", []),
        "ortak_giderler": data.get("ortak_giderler", []),
    }
    with open(FILE_VER, "w", encoding="utf-8") as f:
        json.dump(veriler, f, ensure_ascii=False, indent=2)

def fmt(v):
    """Para birimi formatı: ₺1.234,56  (binlik=nokta, ondalık=virgül, 2 hane)"""
    try:
        v = float(v)
        neg = v < 0
        abs_v = abs(v)
        # iki ondalıklı string
        s = f"{abs_v:,.2f}"          # örn: "1,234.56"
        # Amerikan → Türk format: önce virgülü geçici karaktere al
        s = s.replace(",", "X").replace(".", ",").replace("X", ".")
        return ("−₺" if neg else "₺") + s
    except:
        return "₺0,00"

def buay():
    n = datetime.date.today()
    return f"{n.year}-{n.month:02d}"

def bugun_str():
    return datetime.date.today().strftime("%d.%m.%Y")

def _guncelle_gelecek_aylar(data, yeni_aidat, baslangic_ay=None):
    """
    baslangic_ay'dan itibaren (dahil) özel aidat kaydı OLMAYAN
    gelecek tüm aylara yeni_aidat değerini yazar.
    Daha önce elle girilmiş özel tutarlar korunur.
    Döndürür: değiştirilen ay listesi
    """
    if baslangic_ay is None:
        baslangic_ay = buay()

    # 5 yıl ileriye kadar
    today = datetime.date.today()
    bitis = f"{today.year + 5}-{today.month:02d}"
    ileri_aylar = ay_listesi(baslangic_ay, bitis)

    ozel = data["bina"].setdefault("ozel_aidatlar", {})
    degisen = []
    for ay in ileri_aylar:
        # Sadece özel girilmemiş ya da varsayılanla aynı olanları değiştir
        mevcut = ozel.get(ay)
        if mevcut is None:
            # Kayıt yok → yeni varsayılanı özel olarak yaz
            ozel[ay] = yeni_aidat
            degisen.append(ay)
    return degisen

def simdi():
    return datetime.datetime.now().isoformat(timespec="seconds")

def tarih_str(iso):
    try:
        return datetime.datetime.fromisoformat(iso).strftime("%d.%m.%Y")
    except: return "—"

AYLAR_TR = ["","Ocak","Şubat","Mart","Nisan","Mayıs","Haziran",
             "Temmuz","Ağustos","Eylül","Ekim","Kasım","Aralık"]

def ay_label(ay):
    try:
        y, m = int(ay[:4]), int(ay[5:7])
        return f"{AYLAR_TR[m]} {y}"
    except: return ay

def ay_listesi(baslangic, bitis=None):
    """baslangic'tan bitis'e (dahil) tüm YYYY-MM listesi."""
    if not baslangic: baslangic = buay()
    if not bitis:     bitis = buay()
    result = []
    sy, sm = int(baslangic[:4]), int(baslangic[5:7])
    ey, em = int(bitis[:4]),     int(bitis[5:7])
    while (sy, sm) <= (ey, em):
        result.append(f"{sy}-{sm:02d}")
        sm += 1
        if sm > 12: sm = 1; sy += 1
    return result

def ay_aidat(data, ay, daire_no=None):
    """Belirli ay için geçerli aidat tutarını döndür.
    Önce daireye özgü aidatı, sonra genel özel aidatı, yoksa bina aidatını kullan."""
    # Daireye özgü özel aidat (eksik ödemelerden oluşan borç devri)
    if daire_no is not None:
        d = next((x for x in data["daireler"] if x["no"] == daire_no), None)
        if d:
            daire_ozel = d.get("ozel_aidatlar", {})
            if ay in daire_ozel:
                return daire_ozel[ay]
    # Genel özel aidat tablosu
    ozel = data["bina"].get("ozel_aidatlar", {})
    return ozel.get(ay, data["bina"]["aidat"])

def faiz_hesapla(aidat, faiz_oran, ay_str, son_gun):
    """Geçmiş bir ay için gecikme faizi."""
    today = datetime.date.today()
    ay_y, ay_m = int(ay_str[:4]), int(ay_str[5:7])
    try:
        son = datetime.date(ay_y, ay_m, min(son_gun, 28))
    except:
        son = datetime.date(ay_y, ay_m, 10)
    if today <= son:
        return 0.0
    fark_ay = (today.year - ay_y) * 12 + (today.month - ay_m)
    return round(aidat * (faiz_oran / 100.0) * max(1, fark_ay), 2)

def yeni_makbuz_no(data):
    return f"MKB-{datetime.date.today().year}-{len(data['odemeler'])+1:04d}"

_id_sayac = 0
def yeni_id():
    global _id_sayac
    _id_sayac += 1
    return int(datetime.datetime.now().timestamp() * 1000) * 1000 + _id_sayac

# ─── STİL ────────────────────────────────────────────────────────────────────
def apply_style():
    s = ttk.Style()
    s.theme_use("clam")
    s.configure("Dark.Treeview",
        background=T["bg3"], foreground=T["text"],
        fieldbackground=T["bg3"], rowheight=28, borderwidth=0,
        font=("Segoe UI", 10))
    s.configure("Dark.Treeview.Heading",
        background=T["bg4"], foreground=T["text2"],
        font=("Segoe UI", 9, "bold"), relief="flat")
    s.map("Dark.Treeview",
        background=[("selected", T["sel"])],
        foreground=[("selected", T["gold2"])])
    s.configure("Dark.Vertical.TScrollbar",
        background=T["bg3"], troughcolor=T["bg"],
        borderwidth=0, arrowcolor=T["text3"])
    s.configure("Dark.Horizontal.TScrollbar",
        background=T["bg3"], troughcolor=T["bg"],
        borderwidth=0, arrowcolor=T["text3"])
    s.configure("Dark.TPanedwindow", background=T["gold"])
    s.configure("Dark.TNotebook", background=T["bg2"], borderwidth=0)
    s.configure("Dark.TNotebook.Tab",
        background=T["bg2"], foreground=T["text2"],
        padding=[14, 8], font=("Segoe UI", 10))
    s.map("Dark.TNotebook.Tab",
        background=[("selected", T["bg3"])],
        foreground=[("selected", T["gold2"])])
    s.configure("Dark.TCombobox",
        fieldbackground=T["entry_bg"], background=T["bg3"],
        foreground=T["text"], selectbackground=T["sel"],
        selectforeground=T["text"], borderwidth=0)
    s.map("Dark.TCombobox",
        fieldbackground=[("readonly", T["entry_bg"])],
        foreground=[("readonly", T["text"])])
    s.configure("Gold.TSeparator", background=T["gold"])

# ─── UI YARDIMCILARI ─────────────────────────────────────────────────────────
def frm(parent, bg=None, **kw):
    return tk.Frame(parent, bg=bg or T["bg"], **kw)

def lbl(parent, text, fg=None, bg=None, font=None, **kw):
    return tk.Label(parent, text=text,
        fg=fg or T["text"], bg=bg or T["bg"],
        font=font or ("Segoe UI", 10), **kw)

COLORS = {
    "gold"  : (T["gold"],   "#1a1200"),
    "green" : (T["green"],  "#001a0a"),
    "red"   : (T["red"],    T["text"]),
    "blue"  : (T["blue"],   T["text"]),
    "gray"  : (T["bg4"],    T["text2"]),
    "orange": (T["orange"], "#1a1000"),
    "purple": (T["purple"], T["text"]),
}

def btn(parent, text, cmd, color="gold", ipadx=10, ipady=5, **kw):
    bg, fg = COLORS.get(color, (T["bg4"], T["text"]))
    return tk.Button(parent, text=text, command=cmd,
        bg=bg, fg=fg, activebackground=T["bg4"],
        activeforeground=T["text"], relief="flat",
        font=("Segoe UI", 9, "bold"), cursor="hand2",
        padx=ipadx, pady=ipady, bd=0, **kw)

def ent(parent, textvariable=None, width=20, **kw):
    return tk.Entry(parent,
        bg=T["entry_bg"], fg=T["text"],
        insertbackground=T["gold"],
        relief="flat", bd=0,
        highlightthickness=1,
        highlightcolor=T["gold"],
        highlightbackground=T["border"],
        font=("Segoe UI", 10),
        textvariable=textvariable,
        width=width, **kw)

def cmb(parent, var, values, width=18, **kw):
    cb = ttk.Combobox(parent, textvariable=var, values=values,
        state="readonly", style="Dark.TCombobox",
        font=("Segoe UI", 10), width=width, **kw)
    return cb

def make_tree(parent, cols_widths, height=12, selectmode="extended"):
    cols = [c for c, _ in cols_widths]
    tree = ttk.Treeview(parent, columns=cols, show="headings",
        style="Dark.Treeview", height=height, selectmode=selectmode)
    for col, (h, w) in zip(cols, cols_widths):
        tree.heading(col, text=h, anchor="center")
        tree.column(col, width=w, anchor="center", stretch=True)
    tree.tag_configure("odendi",   background=T["green_bg"], foreground=T["green"])
    tree.tag_configure("borclu",   background=T["red_bg"],   foreground=T["red"])
    tree.tag_configure("bekliyor", background=T["orange_bg"],foreground=T["orange"])
    tree.tag_configure("ileri",    background=T["bg3"],      foreground=T["text2"])
    tree.tag_configure("normal",   background=T["bg3"],      foreground=T["text"])
    tree.tag_configure("gelir",    background=T["bg3"],      foreground=T["green"])
    tree.tag_configure("gider",    background=T["bg3"],      foreground=T["red"])
    tree.tag_configure("ozel",     background=T["blue_bg"],  foreground=T["blue"])
    return tree

def scrolled(parent, cols_widths, height=12, selectmode="extended"):
    f = frm(parent, bg=T["bg3"])
    tree = make_tree(f, cols_widths, height, selectmode)
    vsb = ttk.Scrollbar(f, orient="vertical", command=tree.yview,
                        style="Dark.Vertical.TScrollbar")
    hsb = ttk.Scrollbar(f, orient="horizontal", command=tree.xview,
                        style="Dark.Horizontal.TScrollbar")
    tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
    tree.grid(row=0, column=0, sticky="nsew")
    vsb.grid(row=0, column=1, sticky="ns")
    hsb.grid(row=1, column=0, sticky="ew")
    f.rowconfigure(0, weight=1); f.columnconfigure(0, weight=1)
    return f, tree

def pane(parent, **kw):
    pw = ttk.PanedWindow(parent, style="Dark.TPanedwindow", **kw)
    return pw

def card_panel(parent, title, subtitle=""):
    outer = frm(parent, bg=T["bg2"],
        highlightthickness=1, highlightbackground=T["border"])
    tk.Frame(outer, bg=T["gold"], height=3).pack(fill="x")
    h = frm(outer, bg=T["bg2"])
    h.pack(fill="x", padx=12, pady=(8,6))
    lbl(h, title, font=("Segoe UI",11,"bold"), bg=T["bg2"]).pack(anchor="w")
    if subtitle:
        lbl(h, subtitle, fg=T["text3"], font=("Segoe UI",8),
            bg=T["bg2"]).pack(anchor="w")
    tk.Frame(outer, bg=T["border"], height=1).pack(fill="x")
    return outer

def stat_card(parent, title, value, sub, color="gold"):
    acc = {"gold":T["gold"],"green":T["green"],
           "red":T["red"],"blue":T["blue"],"orange":T["orange"]}.get(color, T["gold"])
    f = frm(parent, bg=T["bg2"],
        highlightthickness=1, highlightbackground=acc)
    tk.Frame(f, bg=acc, height=3).pack(fill="x")
    inner = frm(f, bg=T["bg2"])
    inner.pack(fill="both", expand=True, padx=14, pady=10)
    lbl(inner, title, fg=T["text2"], font=("Segoe UI",9), bg=T["bg2"]).pack(anchor="w")
    vl = lbl(inner, value, fg=acc, font=("Segoe UI",20,"bold"), bg=T["bg2"])
    vl.pack(anchor="w")
    sl = lbl(inner, sub, fg=T["text3"], font=("Segoe UI",8), bg=T["bg2"])
    sl.pack(anchor="w", pady=(2,0))
    return f, vl, sl

def sep(parent, bg=None):
    tk.Frame(parent, bg=bg or T["border"], height=1).pack(fill="x", pady=4)

# ─── MAKBUZ ──────────────────────────────────────────────────────────────────
def makbuz_yazdir(data, odeme):
    b    = data["bina"]
    d    = next((x for x in data["daireler"] if x["no"] == odeme["daireNo"]), {})
    isim = d.get("isim") or "—"
    not_ = odeme.get("not_","")
    html = f"""<!DOCTYPE html><html lang="tr"><head><meta charset="UTF-8">
<title>Makbuz {odeme['makbuzNo']}</title>
<style>
  body{{font-family:Arial;background:#f0f0f0;display:flex;justify-content:center;padding:30px;}}
  .card{{background:#fff;max-width:420px;width:100%;padding:32px;border-radius:8px;
         box-shadow:0 4px 20px rgba(0,0,0,.15);}}
  .top{{text-align:center;border-bottom:2px solid #000;padding-bottom:16px;margin-bottom:16px;}}
  .badge{{text-align:center;background:#f5f5f5;padding:12px;border-radius:6px;margin-bottom:20px;}}
  table{{width:100%;border-collapse:collapse;font-size:14px;}}
  td{{padding:9px 0;border-bottom:1px solid #eee;}}
  td:first-child{{color:#666;width:42%;}}
  .tutar{{text-align:center;background:#111;padding:16px;border-radius:6px;margin-top:20px;}}
  .tutar p{{color:#aaa;font-size:12px;margin:0;}}
  .tutar h3{{color:#fff;font-size:26px;font-weight:bold;margin:4px 0 0;}}
  .footer{{text-align:center;font-size:11px;color:#999;margin-top:16px;}}
  .np{{text-align:center;margin-bottom:20px;}}
  button{{padding:10px 28px;background:#111;color:#fff;border:none;
          border-radius:5px;cursor:pointer;font-size:14px;}}
  @media print{{.np{{display:none;}}body{{background:#fff;padding:0;}}}}
</style></head><body>
<div><div class="np"><button onclick="window.print()">Yazdir Makbuzu Yazdır</button></div>
<div class="card">
  <div class="top">
    <div style="font-size:32px">Bina</div>
    <h1 style="font-size:18px;margin:8px 0 4px">{b['adi']}</h1>
    <p style="color:#666;font-size:12px">Bina Yönetim Sistemi</p>
  </div>
  <div class="badge">
    <h2 style="font-size:20px;letter-spacing:2px;margin:0">AİDAT MAKBUZU</h2>
    <p style="color:#666;font-size:13px;margin:4px 0 0">No: {odeme['makbuzNo']}</p>
  </div>
  <table>
    <tr><td>Daire No</td><td><strong>Daire {odeme['daireNo']}</strong></td></tr>
    <tr><td>Sakin</td><td>{isim}</td></tr>
    <tr><td>Dönem</td><td>{ay_label(odeme['ay'])}</td></tr>
    <tr><td>Ödeme Tarihi</td><td>{tarih_str(odeme['tarih'])}</td></tr>
    <tr><td>Ödeme Yöntemi</td><td>{odeme['yontem']}</td></tr>
    {"<tr><td>Not</td><td>" + not_ + "</td></tr>" if not_ else ""}
  </table>
  <div class="tutar"><p>ÖDENEN TUTAR</p><h3>{fmt(odeme['tutar'])}</h3></div>
  <p class="footer">Bu makbuz {b['adi']} tarafından düzenlenmiştir.</p>
</div></div></body></html>"""
    tmp = tempfile.NamedTemporaryFile(suffix=".html", delete=False,
                                      mode="w", encoding="utf-8")
    tmp.write(html); tmp.close()
    webbrowser.open(f"file:///{tmp.name}")

def og_makbuz_yazdir(data, alacakli, odeme_kaydi):
    """Ortak gider ödemesi için ayrı makbuz oluşturur."""
    b    = data["bina"]
    dno  = alacakli.get("daire_no")
    d    = next((x for x in data["daireler"] if x["no"] == dno), {})
    isim = d.get("isim") or "—"
    gider_adi = alacakli.get("kisi") or alacakli.get("aciklama","") or "Ortak Gider"
    mkb_no = f"OG-{datetime.date.today().year}-{odeme_kaydi['id'] % 10000:04d}"
    not_  = odeme_kaydi.get("not_", "")
    toplam_tutar = alacakli.get("tutar", 0)
    odenen_toplam = sum(p["tutar"] for p in alacakli.get("odemeler", []))
    kalan = round(toplam_tutar - odenen_toplam, 2)
    html = f"""<!DOCTYPE html><html lang="tr"><head><meta charset="UTF-8">
<title>Ortak Gider Makbuzu {mkb_no}</title>
<style>
  body{{font-family:Arial;background:#f0f0f0;display:flex;justify-content:center;padding:30px;}}
  .card{{background:#fff;max-width:440px;width:100%;padding:32px;border-radius:8px;
         box-shadow:0 4px 20px rgba(0,0,0,.15);}}
  .top{{text-align:center;border-bottom:2px solid #000;padding-bottom:16px;margin-bottom:16px;}}
  .badge{{text-align:center;background:#f5f5f5;padding:12px;border-radius:6px;margin-bottom:20px;}}
  .badge2{{text-align:center;background:#fff3e0;border:1px solid #f39c12;padding:8px;border-radius:6px;margin-bottom:16px;}}
  table{{width:100%;border-collapse:collapse;font-size:14px;}}
  td{{padding:9px 0;border-bottom:1px solid #eee;}}
  td:first-child{{color:#666;width:42%;}}
  .tutar{{text-align:center;background:#1a1a2e;padding:16px;border-radius:6px;margin-top:20px;}}
  .tutar p{{color:#aaa;font-size:12px;margin:0;}}
  .tutar h3{{color:#f39c12;font-size:26px;font-weight:bold;margin:4px 0 0;}}
  .kalan{{text-align:center;background:#fff3e0;padding:10px;border-radius:6px;margin-top:10px;}}
  .kalan p{{color:#888;font-size:11px;margin:0;}}
  .kalan h4{{color:#e67e22;font-size:18px;margin:4px 0 0;}}
  .footer{{text-align:center;font-size:11px;color:#999;margin-top:16px;}}
  .np{{text-align:center;margin-bottom:20px;}}
  button{{padding:10px 28px;background:#1a1a2e;color:#f39c12;border:1px solid #f39c12;
          border-radius:5px;cursor:pointer;font-size:14px;font-weight:bold;}}
  @media print{{.np{{display:none;}}body{{background:#fff;padding:0;}}}}
</style></head><body>
<div><div class="np"><button onclick="window.print()">Yazdir Makbuzu Yazdır</button></div>
<div class="card">
  <div class="top">
    <div style="font-size:32px">Arac</div>
    <h1 style="font-size:18px;margin:8px 0 4px">{b['adi']}</h1>
    <p style="color:#666;font-size:12px">Bina Yönetim Sistemi</p>
  </div>
  <div class="badge">
    <h2 style="font-size:18px;letter-spacing:2px;margin:0;color:#e67e22">ORTAK GİDER MAKBUZU</h2>
    <p style="color:#666;font-size:13px;margin:4px 0 0">No: {mkb_no}</p>
  </div>
  <div class="badge2">
    <p style="font-size:13px;font-weight:bold;color:#e67e22;margin:0">Arac {gider_adi}</p>
    <p style="font-size:11px;color:#888;margin:4px 0 0">Ortak Gider / Tadilat</p>
  </div>
  <table>
    <tr><td>Daire No</td><td><strong>Daire {dno}</strong></td></tr>
    <tr><td>Sakin</td><td>{isim}</td></tr>
    <tr><td>Ödeme Tarihi</td><td>{tarih_str(odeme_kaydi['tarih'])}</td></tr>
    <tr><td>Gider Toplam</td><td>{fmt(toplam_tutar)}</td></tr>
    {"<tr><td>Not</td><td>" + not_ + "</td></tr>" if not_ else ""}
  </table>
  <div class="tutar"><p>BU ÖDEMEDE ÖDENEN</p><h3>{fmt(odeme_kaydi['tutar'])}</h3></div>
  {"<div class='kalan'><p>Kalan Borç</p><h4>" + fmt(kalan) + "</h4></div>" if kalan > 0.01 else "<div class='kalan' style='background:#f0faf4;border-color:#27ae60'><p style='color:#27ae60'>OK Borç Tamamen Kapatıldı</p><h4 style='color:#27ae60'>v Sıfırlandı</h4></div>"}
  <p class="footer">Bu makbuz {b['adi']} tarafından düzenlenmiştir.</p>
</div></div></body></html>"""
    tmp = tempfile.NamedTemporaryFile(suffix=".html", delete=False,
                                      mode="w", encoding="utf-8")
    tmp.write(html); tmp.close()
    webbrowser.open(f"file:///{tmp.name}")

# ─── TOPLU MAKBUZ YARDIMCILARI ────────────────────────────────────────────────
def _toplu_makbuz_html(b, isim, daire_no, baslik, rows_html):
    """Birden fazla makbuzu tek HTML sayfasında göster."""
    return f"""<!DOCTYPE html><html lang="tr"><head><meta charset="UTF-8">
<title>{baslik} — Daire {daire_no}</title>
<style>
  body{{font-family:Arial;background:#f0f0f0;padding:24px;}}
  h1{{text-align:center;font-size:18px;color:#333;margin-bottom:4px;}}
  .sub{{text-align:center;color:#888;font-size:12px;margin-bottom:24px;}}
  .np{{text-align:center;margin-bottom:20px;}}
  .np button{{padding:9px 22px;background:#111;color:#fff;border:none;
              border-radius:4px;cursor:pointer;font-size:14px;margin:0 6px;}}
  .np button.og-btn{{background:#1a1a2e;color:#f39c12;border:1px solid #f39c12;}}
  .makbuz{{background:#fff;max-width:420px;margin:0 auto 28px;
            padding:28px;border-radius:8px;
            box-shadow:0 3px 16px rgba(0,0,0,.13);
            page-break-inside:avoid;}}
  .makbuz.og{{border-top:4px solid #f39c12;}}
  .mbadge{{text-align:center;background:#f5f5f5;padding:10px;
            border-radius:6px;margin-bottom:16px;font-size:14px;
            font-weight:bold;letter-spacing:1px;color:#333;}}
  .og-badge{{background:#fff3e0;color:#e67e22;}}
  .gider-adi{{text-align:center;font-size:13px;font-weight:bold;
              color:#e67e22;background:#fff8f0;border:1px solid #f5d6a0;
              border-radius:4px;padding:6px;margin-bottom:12px;}}
  table{{width:100%;border-collapse:collapse;font-size:13px;}}
  td{{padding:8px 0;border-bottom:1px solid #eee;}}
  td:first-child{{color:#666;width:40%;}}
  .tutar{{text-align:center;background:#111;padding:14px;
          border-radius:6px;margin-top:16px;}}
  .tutar.og-tutar{{background:#1a1a2e;}}
  .tutar p{{color:#aaa;font-size:11px;margin:0;}}
  .tutar h3{{color:#fff;font-size:22px;font-weight:bold;margin:4px 0 0;}}
  .tutar.og-tutar h3{{color:#f39c12;}}
  .kalan-kart{{text-align:center;padding:8px;border-radius:5px;margin-top:8px;}}
  .kalan-var{{background:#fff3e0;border:1px solid #f39c12;}}
  .kalan-var p{{color:#888;font-size:11px;margin:0;}}
  .kalan-var h4{{color:#e67e22;font-size:16px;margin:3px 0 0;}}
  .kalan-tamam{{background:#f0faf4;border:1px solid #27ae60;}}
  .kalan-tamam p{{color:#27ae60;font-size:12px;margin:0;}}
  .kalan-tamam h4{{color:#27ae60;font-size:14px;margin:3px 0 0;}}
  .footer{{text-align:center;font-size:10px;color:#bbb;margin-top:12px;}}
  .sayfa-baslik{{background:#fff;max-width:420px;margin:0 auto 10px;
                  padding:14px 24px;border-radius:8px;
                  box-shadow:0 2px 10px rgba(0,0,0,.1);
                  display:flex;align-items:center;gap:12px;}}
  .sayfa-baslik .emoji{{font-size:28px;}}
  .sayfa-baslik .info h2{{margin:0;font-size:15px;}}
  .sayfa-baslik .info p{{margin:0;color:#888;font-size:12px;}}
  @media print{{.np{{display:none;}}body{{background:#fff;padding:0;}}
                .makbuz{{box-shadow:none;border:1px solid #eee;}}}}
</style></head><body>
<div class="np">
  <button onclick="window.print()">Yazdir Tümünü Yazdır</button>
</div>
<h1>Bina {b['adi']}</h1>
<p class="sub">{baslik}  •  Daire {daire_no}  —  {isim}</p>
{rows_html}
</body></html>"""

def _makbuz_tarayici_ac(html):
    """HTML içeriğini geçici dosyaya yazıp tarayıcıda aç."""
    tmp = tempfile.NamedTemporaryFile(suffix=".html", delete=False,
                                      mode="w", encoding="utf-8")
    tmp.write(html); tmp.close()
    webbrowser.open(f"file:///{tmp.name}")


# ══════════════════════════════════════════════════════════════════════════════
#  AY/YIL SEÇİCİ WİDGET
# ══════════════════════════════════════════════════════════════════════════════
class AySecici(tk.Frame):
    """Kompakt ay-yıl seçici: < Ocak 2024 >"""
    def __init__(self, parent, initial=None, callback=None, fg=None, bg=None, **kw):
        super().__init__(parent, bg=bg or T["bg"], **kw)
        self.callback = callback
        today = datetime.date.today()
        if initial:
            y, m = int(initial[:4]), int(initial[5:7])
        else:
            y, m = today.year, today.month
        self._y, self._m = y, m
        self._build(fg, bg)

    def _build(self, fg, bg):
        bg = bg or T["bg"]
        tk.Button(self, text="◀", command=self._prev,
            bg=bg, fg=T["gold"], font=("Segoe UI",11,"bold"),
            relief="flat", cursor="hand2", bd=0, padx=4
        ).pack(side="left")
        self._lbl = lbl(self, self._text(),
            fg=fg or T["gold2"], bg=bg,
            font=("Segoe UI",10,"bold"))
        self._lbl.pack(side="left", padx=6)
        tk.Button(self, text="▶", command=self._next,
            bg=bg, fg=T["gold"], font=("Segoe UI",11,"bold"),
            relief="flat", cursor="hand2", bd=0, padx=4
        ).pack(side="left")

    def _text(self):
        return f"{AYLAR_TR[self._m]} {self._y}"

    def _prev(self):
        self._m -= 1
        if self._m < 1: self._m = 12; self._y -= 1
        self._lbl.config(text=self._text())
        if self.callback: self.callback(self.get())

    def _next(self):
        self._m += 1
        if self._m > 12: self._m = 1; self._y += 1
        self._lbl.config(text=self._text())
        if self.callback: self.callback(self.get())

    def get(self):
        return f"{self._y}-{self._m:02d}"

    def set(self, ay):
        self._y, self._m = int(ay[:4]), int(ay[5:7])
        self._lbl.config(text=self._text())

# ══════════════════════════════════════════════════════════════════════════════
#  KURULUM EKRANI
# ══════════════════════════════════════════════════════════════════════════════

# ══════════════════════════════════════════════════════════════════════════════
#  VERİ BULUNAMADI — BAŞLANGIÇ SEÇİM EKRANI
# ══════════════════════════════════════════════════════════════════════════════
class VeriYokEkrani(tk.Toplevel):
    """
    Program açılışında veri dosyaları bulunamadığında gösterilir.
    Kullanıcıya 3 seçenek sunar:
      1) Yedekten geri yükle (ZIP veya JSON)
      2) Yeni kurulum (SetupEkrani'ya yönlendir)
      3) Çıkış
    """
    def __init__(self, parent, callback_kurulum, callback_yuklendi):
        super().__init__(parent)
        self.parent            = parent
        self.callback_kurulum  = callback_kurulum   # yeni kurulum tamamlanınca
        self.callback_yuklendi = callback_yuklendi  # yedek yüklenince

        self.title("Bina Yönetim Sistemi — Veri Bulunamadı")
        self.configure(bg=T["bg"])
        self.resizable(False, False)
        self.protocol("WM_DELETE_WINDOW", self._cikis)
        self.grab_set()
        self.update_idletasks()
        sw, sh = self.winfo_screenwidth(), self.winfo_screenheight()
        self.geometry(f"520x560+{(sw-520)//2}+{(sh-560)//2}")
        self._build()
        self._yedek_dosyalari_bul()

    # ─── ARAYÜZ ──────────────────────────────────────────────────────────────
    def _build(self):
        tk.Frame(self, bg=T["gold"], height=4).pack(fill="x")

        # Başlık
        hdr = frm(self, bg=T["bg2"])
        hdr.pack(fill="x")
        tk.Frame(hdr, bg=T["bg2"], height=18).pack()
        lbl(hdr, "Uyari", font=("Segoe UI Emoji", 36),
            bg=T["bg2"], fg=T["orange"]).pack()
        lbl(hdr, "Veri Dosyaları Bulunamadı",
            font=("Segoe UI", 16, "bold"), bg=T["bg2"]).pack(pady=(6, 2))
        lbl(hdr, f"Klasör: {DATA_DIR}",
            fg=T["text3"], font=("Segoe UI", 8), bg=T["bg2"]).pack(pady=(0, 14))
        tk.Frame(hdr, bg=T["border"], height=1).pack(fill="x")

        body = frm(self, bg=T["bg"])
        body.pack(fill="both", expand=True, padx=32, pady=20)

        # ── SEÇENEK 1: Yedekten Geri Yükle ──────────────────────────────────
        s1 = frm(body, bg=T["bg3"],
                 highlightthickness=1, highlightbackground=T["orange"])
        s1.pack(fill="x", pady=(0, 12))
        tk.Frame(s1, bg=T["orange"], height=3).pack(fill="x")
        i1 = frm(s1, bg=T["bg3"]); i1.pack(fill="x", padx=16, pady=12)
        lbl(i1, "♻️  Yedekten Geri Yükle",
            font=("Segoe UI", 11, "bold"), fg=T["orange"], bg=T["bg3"]).pack(anchor="w")
        lbl(i1, "ZIP veya JSON yedek dosyası seçerek tüm verilerinizi geri getirin.",
            fg=T["text2"], font=("Segoe UI", 9), bg=T["bg3"]).pack(anchor="w", pady=(2, 8))

        # Otomatik bulunan yedekler
        self._yedek_frame = frm(i1, bg=T["bg3"])
        self._yedek_frame.pack(fill="x", pady=(0, 6))

        btn(i1, "Klasor  Yedek Dosyası Seç…", self._yedek_sec, "orange").pack(
            anchor="w", ipadx=10, ipady=5)

        # ── SEÇENEK 2: Yeni Kurulum ──────────────────────────────────────────
        s2 = frm(body, bg=T["bg3"],
                 highlightthickness=1, highlightbackground=T["gold"])
        s2.pack(fill="x", pady=(0, 12))
        tk.Frame(s2, bg=T["gold"], height=3).pack(fill="x")
        i2 = frm(s2, bg=T["bg3"]); i2.pack(fill="x", padx=16, pady=12)
        lbl(i2, "  Yeni Kurulum Başlat",
            font=("Segoe UI", 11, "bold"), fg=T["gold2"], bg=T["bg3"]).pack(anchor="w")
        lbl(i2, "Bina bilgilerini girerek sistemi sıfırdan kurun.",
            fg=T["text2"], font=("Segoe UI", 9), bg=T["bg3"]).pack(anchor="w", pady=(2, 8))
        btn(i2, "+  Yeni Kurulum", self._yeni_kurulum, "gold").pack(
            anchor="w", ipadx=10, ipady=5)

        # ── SEÇENEK 3: Çıkış ─────────────────────────────────────────────────
        btn(body, "X  Programdan Çık", self._cikis, "gray").pack(
            anchor="e", ipadx=10, ipady=5, pady=(4, 0))

    # ─── YARDIMCI: Olası yedekleri bul ───────────────────────────────────────
    def _yedek_dosyalari_bul(self):
        """Masaüstü, İndirilenler ve ev klasöründe bina_yedek*.zip / *.json ara."""
        import glob
        ev = Path.home()
        arama_klasorleri = [
            ev / "Desktop",
            ev / "Masaüstü",
            ev / "Downloads",
            ev / "İndirilenler",
            ev,
        ]
        bulunanlar = []
        for klasor in arama_klasorleri:
            if klasor.exists():
                for p in sorted(klasor.glob("bina_yedek*.zip"), reverse=True)[:3]:
                    if p not in bulunanlar:
                        bulunanlar.append(p)
                for p in sorted(klasor.glob("bina_yedek*.json"), reverse=True)[:2]:
                    if p not in bulunanlar:
                        bulunanlar.append(p)

        for w in self._yedek_frame.winfo_children():
            w.destroy()

        if bulunanlar:
            lbl(self._yedek_frame, "Ara  Bulunan yedek dosyaları:",
                fg=T["text3"], font=("Segoe UI", 8, "bold"),
                bg=T["bg3"]).pack(anchor="w", pady=(0, 4))
            for p in bulunanlar[:4]:
                pf = frm(self._yedek_frame, bg=T["bg3"])
                pf.pack(fill="x", pady=1)
                lbl(pf, f"   {p.name}",
                    fg=T["text2"], font=("Segoe UI", 8),
                    bg=T["bg3"]).pack(side="left")
                p2 = p
                btn(pf, "Yükle", lambda f=p2: self._yukle_dosya(str(f)),
                    "orange").pack(side="right", ipadx=6, ipady=2)

    # ─── EYLEMLER ────────────────────────────────────────────────────────────
    def _yedek_sec(self):
        from tkinter import filedialog
        kaynak = filedialog.askopenfilename(
            title="Yedek Dosyası Seç",
            filetypes=[
                ("ZIP Yedeği", "*.zip"),
                ("JSON Yedek", "*.json"),
                ("Tüm Dosyalar", "*.*"),
            ],
            parent=self,
        )
        if kaynak:
            self._yukle_dosya(kaynak)

    def _yukle_dosya(self, kaynak):
        import zipfile, shutil
        ext = Path(kaynak).suffix.lower()

        if ext == ".zip":
            try:
                with zipfile.ZipFile(kaynak, "r") as zf:
                    json_dosyalar = [d for d in zf.namelist() if d.endswith(".json")]
            except Exception as e:
                messagebox.showerror("Hata", f"ZIP okunamadı:\n{e}", parent=self)
                return
            if not json_dosyalar:
                messagebox.showerror("Hata",
                    "ZIP içinde JSON dosyası bulunamadı.", parent=self)
                return
            if not messagebox.askyesno("Geri Yükle",
                    f"{len(json_dosyalar)} dosya geri yüklenecek:\n"
                    + "\n".join(f"  • {d}" for d in json_dosyalar)
                    + "\n\nEmin misiniz?", parent=self):
                return
            try:
                DATA_DIR.mkdir(parents=True, exist_ok=True)
                with zipfile.ZipFile(kaynak, "r") as zf:
                    for dosya in json_dosyalar:
                        hedef = DATA_DIR / Path(dosya).name
                        with zf.open(dosya) as src, open(hedef, "wb") as dst:
                            dst.write(src.read())
            except Exception as e:
                messagebox.showerror("Hata", f"Geri yükleme başarısız:\n{e}", parent=self)
                return

        elif ext == ".json":
            try:
                with open(kaynak, "r", encoding="utf-8") as f:
                    test = json.load(f)
                if "bina" not in test or "daireler" not in test:
                    raise ValueError("Geçersiz bina yönetim yedeği.")
            except Exception as e:
                messagebox.showerror("Hata", f"Dosya okunamadı:\n{e}", parent=self)
                return
            if not messagebox.askyesno("Geri Yükle",
                    "JSON yedekten geri yüklenecek. Emin misiniz?", parent=self):
                return
            try:
                DATA_DIR.mkdir(parents=True, exist_ok=True)
                shutil.copy2(kaynak, DATA_FILE_LEGACY)
            except Exception as e:
                messagebox.showerror("Hata", f"Kopyalama başarısız:\n{e}", parent=self)
                return
        else:
            messagebox.showerror("Hata",
                "Desteklenmeyen dosya türü. ZIP veya JSON seçin.", parent=self)
            return

        # Veriyi yükle ve uygulamayı başlat
        data = load_data()
        if not data:
            messagebox.showerror("Hata",
                "Yedek yüklendi ancak veri okunamadı.", parent=self)
            return

        # Migrate kontrolleri
        if "ozel_aidatlar" not in data["bina"]:
            data["bina"]["ozel_aidatlar"] = {}
        if "baslangic" not in data["bina"]:
            data["bina"]["baslangic"] = buay()
        if "ortak_giderler" not in data:
            data["ortak_giderler"] = []

        messagebox.showinfo("Tamamlandı",
            "Veriler başarıyla geri yüklendi.", parent=self)
        self.destroy()
        self.callback_yuklendi(data)

    def _yeni_kurulum(self):
        self.destroy()
        SetupEkrani(self.parent, self.callback_kurulum)

    def _cikis(self):
        self.parent.destroy()

class SetupEkrani(tk.Toplevel):
    def __init__(self, parent, callback):
        super().__init__(parent)
        self.callback = callback
        self.title("Bina Yönetim Sistemi — İlk Kurulum")
        self.configure(bg=T["bg"])
        self.resizable(False, False)
        self.grab_set()
        self.update_idletasks()
        sw, sh = self.winfo_screenwidth(), self.winfo_screenheight()
        self.geometry(f"520x800+{(sw-520)//2}+{(sh-800)//2}")
        self._build()

    def _build(self):
        tk.Frame(self, bg=T["gold"], height=4).pack(fill="x")
        hdr = frm(self, bg=T["bg2"])
        hdr.pack(fill="x")
        tk.Frame(hdr, bg=T["bg2"], height=14).pack()
        lbl(hdr, "Bina", font=("Segoe UI Emoji",34),
            bg=T["bg2"], fg=T["gold"]).pack()
        lbl(hdr, "Bina Yönetim Sistemi",
            font=("Segoe UI",17,"bold"), bg=T["bg2"]).pack(pady=(4,2))
        lbl(hdr, "Lütfen bina bilgilerini girin",
            fg=T["text2"], bg=T["bg2"]).pack(pady=(0,14))

        form = frm(self, bg=T["bg"])
        form.pack(fill="both", expand=True, padx=40, pady=8)

        # Normal alanlar
        fields = [
            ("adi",    "Bina  Bina Adı",               "Örn: Gül Apartmanı"),
            ("daire",  "  Daire Sayısı",            "Örn: 12"),
            ("kat",    "Insaat  Kat Sayısı",             "Örn: 4"),
            ("aidat",  "Para  Aylık Aidat (₺)",         "Örn: 750"),
            ("faiz",   "Yukari  Geç Ödeme Faizi (%/ay)",  "Örn: 2"),
            ("songun", "Takvim  Son Ödeme Günü",           "Örn: 10"),
        ]
        self._vars = {}
        for key, ltext, ph in fields:
            lbl(form, ltext, fg=T["text2"],
                font=("Segoe UI",9,"bold"), bg=T["bg"]).pack(anchor="w", pady=(8,2))
            var = tk.StringVar(value=ph)
            e = ent(form, textvariable=var, width=42)
            e.pack(fill="x", ipady=7)
            def fi(ev, v=var, p=ph):
                if v.get()==p: v.set("")
            def fo(ev, v=var, p=ph):
                if not v.get(): v.set(p)
            e.bind("<FocusIn>", fi)
            e.bind("<FocusOut>", fo)
            self._vars[key] = (var, ph)

        # Başlangıç ayı seçici
        lbl(form, "  Aidat Başlangıç Ayı / Yılı",
            fg=T["text2"], font=("Segoe UI",9,"bold"),
            bg=T["bg"]).pack(anchor="w", pady=(12,4))

        bas_frm = frm(form, bg=T["bg3"],
            highlightthickness=1, highlightbackground=T["border"])
        bas_frm.pack(fill="x", ipady=6, pady=(0,4))

        inner_bas = frm(bas_frm, bg=T["bg3"])
        inner_bas.pack(padx=12, pady=6)

        lbl(inner_bas, "Sistem bu aydan itibaren aidat takibi yapar:",
            fg=T["text3"], font=("Segoe UI",8),
            bg=T["bg3"]).pack(anchor="w", pady=(0,6))

        self._bas_sec = AySecici(inner_bas, bg=T["bg3"])
        self._bas_sec.pack(anchor="w")

        btn(self, "  Sistemi Başlat", self._kaydet, "gold").pack(
            fill="x", padx=40, pady=16, ipady=10)

    def _kaydet(self):
        def val(k):
            v, ph = self._vars[k]
            s = v.get().strip()
            return "" if s == ph else s

        adi   = val("adi")
        daire = val("daire")
        aidat = val("aidat")
        if not adi or not daire or not aidat:
            messagebox.showerror("Eksik",
                "Bina adı, daire sayısı ve aidat zorunludur.", parent=self)
            return
        try:
            daire  = int(daire)
            aidat  = float(aidat)
            faiz   = float(val("faiz") or 0)
            kat    = int(val("kat") or 4)
            son_gun= int(val("songun") or 10)
        except ValueError:
            messagebox.showerror("Hata",
                "Sayısal alanlara geçerli değer girin.", parent=self)
            return

        bas = self._bas_sec.get()
        data = {
            "bina": {
                "adi": adi, "daire": daire, "kat": kat,
                "aidat": aidat, "faiz": faiz, "son_gun": son_gun,
                "baslangic": bas,
                "ozel_aidatlar": {},
            },
            "daireler": [
                {"no": i, "isim": "", "tel": "", "email": "",
                 "borc": 0.0, "faiz": 0.0, "son_odeme": ""}
                for i in range(1, daire+1)
            ],
            "odemeler": [], "gelirler": [], "giderler": [],
            "alacaklilar": [],
        }
        save_data(data)
        self.destroy()
        self.callback(data)

# ══════════════════════════════════════════════════════════════════════════════
#  ÖDEME DÜZENLEME DİYALOĞU
# ══════════════════════════════════════════════════════════════════════════════
class OdemeDuzenle(tk.Toplevel):
    """Mevcut bir ödemenin tutar, yöntem ve notunu düzenle."""
    def __init__(self, parent, data, odeme, callback=None):
        super().__init__(parent)
        self.data     = data
        self.odeme    = odeme
        self.callback = callback
        self.title(f"Ödeme Düzenle — Daire {odeme['daireNo']} / {ay_label(odeme['ay'])}")
        self.configure(bg=T["bg"])
        self.resizable(False, False)
        self.grab_set()
        self.update_idletasks()
        pr = parent.winfo_rootx(); pt = parent.winfo_rooty()
        self.geometry(f"420x460+{pr+60}+{pt+60}")
        self._build()

    def _build(self):
        o = self.odeme
        tk.Frame(self, bg=T["gold"], height=3).pack(fill="x")
        lbl(self, f"Duzenle  Ödeme Düzenle",
            font=("Segoe UI",13,"bold")).pack(pady=(14,4), padx=20, anchor="w")
        lbl(self, f"Daire {o['daireNo']}  •  {ay_label(o['ay'])}  •  Makbuz: {o['makbuzNo']}",
            fg=T["text2"], font=("Segoe UI",9)).pack(padx=20, anchor="w", pady=(0,10))
        sep(self)

        form = frm(self)
        form.pack(fill="x", padx=20, pady=8)

        # Tutar
        lbl(form, "Tutar (₺)", fg=T["text2"],
            font=("Segoe UI",9,"bold")).pack(anchor="w", pady=(0,3))
        self._tutar = tk.StringVar(value=str(o["tutar"]))
        ent(form, textvariable=self._tutar, width=30).pack(
            fill="x", ipady=7)

        # Yöntem
        lbl(form, "Ödeme Yöntemi", fg=T["text2"],
            font=("Segoe UI",9,"bold")).pack(anchor="w", pady=(12,3))
        self._yon = tk.StringVar(value=o.get("yontem","Havale/EFT"))
        cmb(form, self._yon,
            ["Havale/EFT","Nakit","Kredi Kartı"], width=28).pack(
            anchor="w", ipady=5)

        # Not
        lbl(form, "Not", fg=T["text2"],
            font=("Segoe UI",9,"bold")).pack(anchor="w", pady=(12,3))
        self._not = tk.StringVar(value=o.get("not_",""))
        ent(form, textvariable=self._not, width=38).pack(
            fill="x", ipady=7)

        # Tarih
        lbl(form, "Ödeme Tarihi (GG.AA.YYYY)",
            fg=T["text2"], font=("Segoe UI",9,"bold")).pack(anchor="w", pady=(12,3))
        try:
            mevcut_tarih = datetime.datetime.fromisoformat(
                o["tarih"]).strftime("%d.%m.%Y")
        except:
            mevcut_tarih = datetime.date.today().strftime("%d.%m.%Y")
        self._tarih = tk.StringVar(value=mevcut_tarih)
        ent(form, textvariable=self._tarih, width=20).pack(
            anchor="w", ipady=7)

        sep(self)
        btn_f = frm(self)
        btn_f.pack(fill="x", padx=20, pady=12)
        btn(btn_f, "X İptal",    self.destroy,    "gray").pack(side="right", padx=(6,0))
        btn(btn_f, "OK Kaydet",   self._kaydet,    "gold").pack(side="right")
        btn(btn_f, "Yazdir Makbuz", self._makbuz,    "gray").pack(side="left")
        btn(btn_f, "Sil Sil",    self._sil_odeme, "red" ).pack(side="left", padx=(8,0))

    def _sil_odeme(self):
        """Ödemeyi ve ilişkili gelir kaydını tamamen sil, daire borcuna geri ekle.
           Eğer bu ödeme eksik ödeme nedeniyle sonraki aya ek tutar eklediyse onu da düzelt."""
        o = self.odeme
        if not messagebox.askyesno(
            "Ödemeyi Sil",
            f"Daire {o['daireNo']} — {ay_label(o['ay'])} ödeme kaydı silinsin mi?\n"
            f"Tutar: {fmt(o['tutar'])}\n\n"
            "Bu işlem geri alınamaz. Gelir kaydı da silinecek.",
            parent=self
        ): return

        # Ödemeyi sil
        self.data["odemeler"] = [
            x for x in self.data["odemeler"] if x["id"] != o["id"]
        ]
        # İlişkili gelir kaydını sil
        self.data["gelirler"] = [
            g for g in self.data["gelirler"]
            if not (g.get("odeme_id") == o["id"] or g["id"] == o["id"]+1)
        ]
        # Daire borcuna geri ekle
        d = next((x for x in self.data["daireler"] if x["no"] == o["daireNo"]), None)
        if d:
            d["borc"] = round(d["borc"] + o["tutar"], 2)

        # 1. Eksik ödeme düzeltmesi: bu ay için beklenen tutar ile ödenen arasındaki
        #    fark bir sonraki ayın DAİREYE ÖZGÜ özel aidatına eklenmiş olabilir — geri al
        b = self.data["bina"]
        aidat_beklenen = ay_aidat(self.data, o["ay"], o["daireNo"])
        faiz_bek       = faiz_hesapla(aidat_beklenen, b["faiz"], o["ay"], b["son_gun"])
        tam_beklenen   = round(aidat_beklenen + faiz_bek, 2)
        eklenmis_eksik = round(tam_beklenen - o["tutar"], 2)

        if eklenmis_eksik > 0.01:
            # Bir sonraki ayı hesapla
            ay_y, ay_m = int(o["ay"][:4]), int(o["ay"][5:7])
            ay_m += 1
            if ay_m > 12: ay_m = 1; ay_y += 1
            sonraki = f"{ay_y}-{ay_m:02d}"
            # Daireye özgü ozel_aidatlar'dan çıkar
            d_obj = next((x for x in self.data["daireler"]
                          if x["no"] == o["daireNo"]), None)
            if d_obj is not None:
                daire_ozel = d_obj.get("ozel_aidatlar", {})
                if sonraki in daire_ozel:
                    duzeltilmis = round(daire_ozel[sonraki] - eklenmis_eksik, 2)
                    # Genel aidat değerini hesapla (daire özeli olmadan)
                    genel_aidat = ay_aidat(self.data, sonraki)  # daire_no olmadan
                    if abs(duzeltilmis - genel_aidat) < 0.01:
                        # Genel aidatla aynıysa daireye özgü kaydı sil
                        del daire_ozel[sonraki]
                    else:
                        daire_ozel[sonraki] = max(0.0, duzeltilmis)

        save_data(self.data)
        self.destroy()
        if self.callback: self.callback()

    def _kaydet(self):
        try:
            tutar = float(self._tutar.get())
        except:
            messagebox.showerror("Hata", "Geçerli tutar girin.", parent=self)
            return
        # Tarihi ISO'ya çevir
        try:
            t = datetime.datetime.strptime(self._tarih.get(), "%d.%m.%Y")
            tarih_iso = t.isoformat(timespec="seconds")
        except:
            tarih_iso = self.odeme["tarih"]

        eski_tutar = self.odeme["tutar"]
        self.odeme["tutar"]  = tutar
        self.odeme["yontem"] = self._yon.get()
        self.odeme["not_"]   = self._not.get().strip()
        self.odeme["tarih"]  = tarih_iso

        # Gelir kaydını güncelle (eşleşen id+1)
        for g in self.data["gelirler"]:
            if g["id"] == self.odeme["id"] + 1:
                g["tutar"] = tutar
                break

        # Daire borcunu fark kadar güncelle
        fark = tutar - eski_tutar
        d = next((x for x in self.data["daireler"]
                  if x["no"] == self.odeme["daireNo"]), None)
        if d:
            d["borc"] = max(0.0, round(d["borc"] - fark, 2))

        save_data(self.data)
        self.destroy()
        if self.callback: self.callback()

    def _makbuz(self):
        makbuz_yazdir(self.data, self.odeme)

# ══════════════════════════════════════════════════════════════════════════════
#  AYLIK ÖZEL AİDAT YÖNETİMİ
# ══════════════════════════════════════════════════════════════════════════════
class OzelAidatPencere(tk.Toplevel):
    """
    Belirli bir ay için özel aidat tutarı belirle.
    Geçmiş ve gelecek aylar listelenebilir.
    """
    def __init__(self, parent, data, callback=None):
        super().__init__(parent)
        self.data     = data
        self.callback = callback
        self.title("Aylık Özel Aidat Tutarları")
        self.configure(bg=T["bg"])
        self.geometry("680x560")
        self.minsize(560, 400)
        self.grab_set()
        self.update_idletasks()
        sw, sh = self.winfo_screenwidth(), self.winfo_screenheight()
        self.geometry(f"680x560+{(sw-680)//2}+{(sh-560)//2}")
        self._build()
        self._tree_doldur()

    def _build(self):
        b = self.data["bina"]
        tk.Frame(self, bg=T["gold"], height=3).pack(fill="x")

        hdr = frm(self, bg=T["bg2"])
        hdr.pack(fill="x")
        lbl_f = frm(hdr, bg=T["bg2"])
        lbl_f.pack(fill="x", padx=14, pady=10)
        lbl(lbl_f, "Takvim  Aylık Özel Aidat Tutarları",
            font=("Segoe UI",13,"bold"), bg=T["bg2"]).pack(side="left")
        lbl(lbl_f, f"  Varsayılan aidat: {fmt(b['aidat'])}",
            fg=T["text3"], bg=T["bg2"], font=("Segoe UI",9)).pack(side="left",padx=8)
        tk.Frame(hdr, bg=T["border"], height=1).pack(fill="x")

        # Form: Ay seçici + tutar gir
        form_wrap = frm(self, bg=T["bg3"],
            highlightthickness=1, highlightbackground=T["border"])
        form_wrap.pack(fill="x", padx=14, pady=10)
        form = frm(form_wrap, bg=T["bg3"])
        form.pack(fill="x", padx=14, pady=10)

        lbl(form, "Ay Seçin:", fg=T["text2"],
            bg=T["bg3"], font=("Segoe UI",9,"bold")).grid(
            row=0, column=0, sticky="w", padx=(0,8))
        self._ay_sec = AySecici(form, bg=T["bg3"])
        self._ay_sec.grid(row=0, column=1, sticky="w")

        lbl(form, "Aidat Tutarı (₺):", fg=T["text2"],
            bg=T["bg3"], font=("Segoe UI",9,"bold")).grid(
            row=0, column=2, sticky="w", padx=(20,8))
        self._tutar_var = tk.StringVar(value=str(b["aidat"]))
        ent(form, textvariable=self._tutar_var, width=12).grid(
            row=0, column=3, ipady=6)

        btn(form, "+ Ekle / Güncelle", self._ekle, "gold").grid(
            row=0, column=4, padx=(12,0))
        btn(form, "✕ Varsayılana Dön", self._sil_secili, "gray").grid(
            row=0, column=5, padx=(6,0))

        lbl(form,
            "i  Gelecek aya özel tutar eklerken 'Evet' seçeneğiyle o aydan sonrasına toplu uygulayabilirsiniz.",
            fg=T["text3"], font=("Segoe UI",8),
            bg=T["bg3"]).grid(row=1, column=0, columnspan=6, sticky="w", pady=(8,0))

        # Tablo
        tbl_wrap = frm(self)
        tbl_wrap.pack(fill="both", expand=True, padx=14, pady=(0,8))

        tbl_f, self._tree = scrolled(tbl_wrap, [
            ("Dönem",160), ("Durum",110), ("Özel Tutar",130),
            ("Varsayılan",130), ("Fark",100),
        ], height=14)
        tbl_f.pack(fill="both", expand=True)

        self._tree.bind("<<TreeviewSelect>>", self._on_sec)

        # Alt butonlar
        alt = frm(self)
        alt.pack(fill="x", padx=14, pady=(0,12))
        btn(alt, "Sil Seçili Özel Tutarı Sil", self._sil_secili, "red").pack(side="left")
        btn(alt, "OK Kapat", self.destroy, "gold").pack(side="right")

    def _tree_doldur(self):
        self._tree.delete(*self._tree.get_children())
        b    = self.data["bina"]
        ozel = b.get("ozel_aidatlar", {})
        bas  = b.get("baslangic", buay())

        # Geçmişten 5 yıl ileriye kadar
        today = datetime.date.today()
        bitis = f"{today.year + 5}-{today.month:02d}"
        tum   = ay_listesi(bas, bitis)

        for ay in reversed(tum):
            varsayilan = b["aidat"]
            ozel_val   = ozel.get(ay)
            today_str  = buay()

            if ay > today_str:
                if ozel_val is not None:
                    durum_txt = "o* İleri / Özel"
                    durum_tag = "ozel"
                else:
                    durum_txt = "o İleri Dönem"
                    durum_tag = "ileri"
            elif ozel_val is not None:
                durum_txt = "* Özel Tutar"
                durum_tag = "ozel"
            else:
                durum_txt = "Liste Varsayılan"
                durum_tag = "normal"

            goster = fmt(ozel_val) if ozel_val is not None else f"{fmt(varsayilan)} (varsayılan)"
            fark   = round((ozel_val if ozel_val is not None else varsayilan) - varsayilan, 2)
            fark_s = (f"+{fmt(fark)}" if fark > 0 else fmt(fark)) if fark != 0 else "—"

            self._tree.insert("", "end", iid=ay, tags=(durum_tag,),
                values=(ay_label(ay), durum_txt, goster, fmt(varsayilan), fark_s))

    def _on_sec(self, event):
        sel = self._tree.selection()
        if not sel: return
        ay  = sel[0]
        ozel = self.data["bina"].get("ozel_aidatlar", {})
        self._ay_sec.set(ay)
        val = ozel.get(ay, self.data["bina"]["aidat"])
        self._tutar_var.set(str(val))

    def _ekle(self):
        ay = self._ay_sec.get()
        try:
            tutar = float(self._tutar_var.get())
        except:
            messagebox.showerror("Hata", "Geçerli tutar girin.", parent=self)
            return

        if "ozel_aidatlar" not in self.data["bina"]:
            self.data["bina"]["ozel_aidatlar"] = {}

        # Seçilen ay gelecekte veya bugündeyse "devam aylarına da uygula?" sor
        sonraki_aylar = []
        if ay >= buay():
            cevap = messagebox.askyesnocancel(
                "Uygulama Kapsamı",
                f"{ay_label(ay)} için aidat {fmt(tutar)} olarak ayarlanıyor.\n\n"
                f"Bu tutarı {ay_label(ay)} ve sonrasındaki\ntüm aylara da uygulamak istiyor musunuz?\n\n"
                f"• Evet  → {ay_label(ay)}'dan itibaren tüm gelecek aylara uygular\n"
                f"• Hayır → Sadece {ay_label(ay)}'a uygular\n"
                f"• İptal → Hiçbir şey değişmez",
                parent=self,
            )
            if cevap is None:       # İptal
                return
            if cevap:               # Evet — tüm gelecek aylara uygula
                sonraki_aylar = _guncelle_gelecek_aylar(
                    self.data, tutar, baslangic_ay=ay
                )

        # Seçilen ayı her halükarda yaz
        self.data["bina"]["ozel_aidatlar"][ay] = tutar
        save_data(self.data)
        self._tree_doldur()
        if self.callback: self.callback()

        if sonraki_aylar:
            messagebox.showinfo(
                "Kaydedildi",
                f"{ay_label(ay)}'dan itibaren {len(sonraki_aylar)} aya\n"
                f"{fmt(tutar)} uygulandı.",
                parent=self,
            )
        else:
            messagebox.showinfo(
                "Kaydedildi",
                f"{ay_label(ay)} için aidat: {fmt(tutar)}",
                parent=self,
            )

    def _sil_secili(self):
        sel = self._tree.selection()
        if not sel:
            messagebox.showinfo("Seçim", "Lütfen bir ay seçin.", parent=self); return
        ay = sel[0]
        ozel = self.data["bina"].get("ozel_aidatlar", {})
        if ay in ozel:
            del ozel[ay]
            save_data(self.data)
            self._tree_doldur()
            if self.callback: self.callback()

# ══════════════════════════════════════════════════════════════════════════════
#  DAİRE DETAY PENCERESİ
# ══════════════════════════════════════════════════════════════════════════════
class DaireDetay(tk.Toplevel):
    def __init__(self, parent, app, daire_no):
        super().__init__(parent)
        self.app      = app
        self.daire_no = daire_no
        d = next((x for x in app.data["daireler"] if x["no"] == daire_no), {})
        isim = d.get("isim","")
        baslik = f"Daire {daire_no}" + (f"  —  {isim}" if isim else "")
        self.title(baslik + "  |  Aidat Geçmişi & Tahsilat")
        self.configure(bg=T["bg"])
        self.geometry("1020x720")
        self.minsize(800,520)
        sw, sh = self.winfo_screenwidth(), self.winfo_screenheight()
        self.geometry(f"1020x720+{(sw-1020)//2}+{(sh-720)//2}")
        self._build()
        self._guncelle()

    def _build(self):
        # ── Üst bar ──────────────────────────────────────────────────────────
        bar = frm(self, bg=T["bg2"])
        bar.pack(fill="x")
        tk.Frame(bar, bg=T["gold"], height=3).pack(fill="x")

        info = frm(bar, bg=T["bg2"])
        info.pack(side="left", padx=14, pady=8)
        self._lb_borc   = lbl(info,"", fg=T["red"],   bg=T["bg2"], font=("Segoe UI",10))
        self._lb_faiz   = lbl(info,"", fg=T["orange"],bg=T["bg2"], font=("Segoe UI",10))
        self._lb_top    = lbl(info,"", fg=T["text"],  bg=T["bg2"], font=("Segoe UI",10,"bold"))
        self._lb_borc.pack(side="left", padx=(0,16))
        self._lb_faiz.pack(side="left", padx=(0,16))
        self._lb_top.pack(side="left")

        btns = frm(bar, bg=T["bg2"])
        btns.pack(side="right", padx=14, pady=8)
        btn(btns,"Kisi Sakin Düzenle",        self._sakin, "gray").pack(side="left",padx=4)
        btn(btns,"OK Seçili Ayları Tahsil",  self._tahsil,"gold").pack(side="left",padx=4)

        tk.Frame(self, bg=T["border"], height=1).pack(fill="x")

        # ── Ana paned (dikey) ─────────────────────────────────────────────────
        pw = pane(self, orient="vertical")
        pw.pack(fill="both", expand=True)

        # Üst: tüm aylar
        ust = frm(pw)
        pw.add(ust, weight=58)

        uhdr = frm(ust, bg=T["bg2"])
        uhdr.pack(fill="x")
        tk.Frame(uhdr, bg=T["gold"], height=2).pack(fill="x")
        uh2 = frm(uhdr, bg=T["bg2"])
        uh2.pack(fill="x", padx=12, pady=6)
        lbl(uh2, "Takvim  Tüm Aidat Dönemleri",
            font=("Segoe UI",11,"bold"), bg=T["bg2"]).pack(side="left")
        lbl(uh2, "  Ctrl+tıkla = çoklu seçim  •  Çift tıkla ödendi = düzenle",
            fg=T["text3"], font=("Segoe UI",8), bg=T["bg2"]).pack(side="left",padx=6)

        ust_f, self._tree_ay = scrolled(ust, [
            ("Dönem",150), ("Durum",120), ("Aidat",110),
            ("Faiz",110), ("Toplam",110),
            ("Ödeme Tarihi",120), ("Yöntem",110),
        ], height=10)
        ust_f.pack(fill="both", expand=True, padx=8, pady=(4,8))
        self._tree_ay.bind("<Double-1>", self._ay_cift_tik)

        # Alt: geçmiş ödemeler
        alt = frm(pw)
        pw.add(alt, weight=42)

        ahdr = frm(alt, bg=T["bg2"])
        ahdr.pack(fill="x")
        tk.Frame(ahdr, bg=T["gold"], height=2).pack(fill="x")
        ah2 = frm(ahdr, bg=T["bg2"])
        ah2.pack(fill="x", padx=12, pady=6)
        lbl(ah2, "Makbuz  Geçmiş Ödemeler",
            font=("Segoe UI",11,"bold"), bg=T["bg2"]).pack(side="left")
        lbl(ah2, "  Çift tıkla = düzenle / makbuz",
            fg=T["text3"], font=("Segoe UI",8), bg=T["bg2"]).pack(side="left",padx=6)

        alt_f, self._tree_gec = scrolled(alt, [
            ("Dönem",150),("Tutar",110),("Yöntem",110),
            ("Tarih",120),("Makbuz No",140),("Not",160),
        ], height=7, selectmode="browse")
        alt_f.pack(fill="both", expand=True, padx=8, pady=(4,8))

        self._tree_gec.bind("<Double-1>", self._gec_cift_tik)

    # ── VERİ YENİLE ───────────────────────────────────────────────────────────
    def _guncelle(self):
        b = self.app.data["bina"]
        d = next((x for x in self.app.data["daireler"]
                  if x["no"] == self.daire_no), {})
        self._lb_borc.config(text=f"Anapara: {fmt(d.get('borc',0))}")
        self._lb_faiz.config(text=f"Faiz: {fmt(d.get('faiz',0))}")
        self._lb_top.config(text=f"Toplam: {fmt(d.get('borc',0)+d.get('faiz',0))}")

        # Tüm ay listesi (başlangıçtan 12 ay sonrasına kadar)
        bas = b.get("baslangic", buay())
        today = buay()
        ileri = (f"{int(today[:4])}-{int(today[5:])+12:02d}"
                 if int(today[5:]) <= 0 else
                 f"{int(today[:4]) + (int(today[5:])+11)//12}-"
                 f"{(int(today[5:])+11)%12+1:02d}")
        # Daha güvenli +12 ay hesabı
        ty, tm = int(today[:4]), int(today[5:])
        tm += 12
        if tm > 12: ty += tm // 12; tm = tm % 12 or 12
        ileri = f"{ty}-{tm:02d}"

        tum = ay_listesi(bas, ileri)

        self._tree_ay.delete(*self._tree_ay.get_children())
        for ay in reversed(tum):
            odeme = next((o for o in self.app.data["odemeler"]
                          if o["daireNo"] == self.daire_no and o["ay"] == ay), None)
            aidat_tut = ay_aidat(self.app.data, ay, self.daire_no)
            gec_faiz  = faiz_hesapla(aidat_tut, b["faiz"], ay, b["son_gun"])
            toplam    = aidat_tut + gec_faiz

            is_ileri = ay > today

            if odeme:
                tag = "odendi"
                self._tree_ay.insert("", "end", iid=ay, tags=(tag,),
                    values=(ay_label(ay), "v  Ödendi",
                            fmt(aidat_tut),
                            fmt(gec_faiz) if gec_faiz else "—",
                            fmt(odeme["tutar"]),
                            tarih_str(odeme["tarih"]),
                            odeme["yontem"]))
            elif is_ileri:
                tag = "ileri"
                self._tree_ay.insert("", "end", iid=ay, tags=(tag,),
                    values=(ay_label(ay), "o İleri Dönem",
                            fmt(aidat_tut), "—", "—", "—", "—"))
            else:
                tag = "borclu"
                self._tree_ay.insert("", "end", iid=ay, tags=(tag,),
                    values=(ay_label(ay), "✘  Ödenmedi",
                            fmt(aidat_tut),
                            fmt(gec_faiz) if gec_faiz else "—",
                            fmt(toplam), "—", "—"))

        # Geçmiş ödemeler
        self._tree_gec.delete(*self._tree_gec.get_children())
        gecmis = [o for o in self.app.data["odemeler"]
                  if o["daireNo"] == self.daire_no]
        for o in reversed(gecmis):
            self._tree_gec.insert("", "end", iid=str(o["id"]), tags=("normal",),
                values=(ay_label(o["ay"]), fmt(o["tutar"]),
                        o["yontem"], tarih_str(o["tarih"]),
                        o["makbuzNo"], o.get("not_","") or ""))

    # ── OLAYLAR ───────────────────────────────────────────────────────────────
    def _ay_cift_tik(self, event):
        """Ödenmiş aya çift tıkla = düzenle, ödenmemişe = hızlı tahsil."""
        item = self._tree_ay.identify_row(event.y)
        if not item: return
        odeme = next((o for o in self.app.data["odemeler"]
                      if o["daireNo"] == self.daire_no and o["ay"] == item), None)
        if odeme:
            OdemeDuzenle(self, self.app.data, odeme,
                         callback=lambda: (self._guncelle(), self.app.refresh_all()))
        else:
            if item > buay():
                messagebox.showinfo("İleri Dönem",
                    "Bu ay henüz gelmedi.\n"
                    "İleriye dönük tahsilat için seçip 'Tahsil Et' butonunu kullanabilirsiniz.",
                    parent=self)
            else:
                self._tree_ay.selection_set(item)
                self._tahsil()

    def _gec_cift_tik(self, event):
        """Geçmiş ödemelere çift tıkla = düzenle."""
        item = self._tree_gec.identify_row(event.y)
        if not item: return
        oid   = int(item)
        odeme = next((o for o in self.app.data["odemeler"] if o["id"] == oid), None)
        if odeme:
            OdemeDuzenle(self, self.app.data, odeme,
                         callback=lambda: (self._guncelle(), self.app.refresh_all()))

    def _tahsil(self):
        secim = self._tree_ay.selection()
        if not secim:
            messagebox.showinfo("Seçim",
                "Tahsil etmek istediğiniz ay(lar)ı seçin.\n"
                "(Ctrl+tıklama ile birden fazla ay seçebilirsiniz)", parent=self)
            return

        odenm = []
        for ay in secim:
            odeme = next((o for o in self.app.data["odemeler"]
                          if o["daireNo"] == self.daire_no and o["ay"] == ay), None)
            if not odeme:
                odenm.append(ay)

        if not odenm:
            messagebox.showinfo("Uyarı",
                "Seçilen tüm aylar zaten ödenmiş.\n"
                "Düzenlemek için çift tıklayın.", parent=self)
            return

        _TahsilatOnay(self, self.app, self.daire_no, odenm,
                      callback=lambda: (self._guncelle(), self.app.refresh_all()))

    def _sakin(self):
        d = next((x for x in self.app.data["daireler"]
                  if x["no"] == self.daire_no), {})
        _SakinDuzenle(self, self.app, self.daire_no, d,
                      callback=self._guncelle)

# ── TAHSİLAT ONAY ─────────────────────────────────────────────────────────────
class _TahsilatOnay(tk.Toplevel):
    def __init__(self, parent, app, daire_no, aylar, callback=None):
        super().__init__(parent)
        self.app      = app
        self.daire_no = daire_no
        self.aylar    = aylar
        self.callback = callback
        self.title("Ödeme Detayı")
        self.configure(bg=T["bg"])
        self.resizable(False, False)
        self.grab_set()
        self.update_idletasks()
        pr = parent.winfo_rootx(); pt = parent.winfo_rooty()
        pw = parent.winfo_width(); ph = parent.winfo_height()
        self.geometry(f"500x560+{pr+(pw-500)//2}+{pt+(ph-560)//2}")
        self._build()

    def _build(self):
        b = self.app.data["bina"]
        tk.Frame(self, bg=T["gold"], height=3).pack(fill="x")
        lbl(self, f"Kart  {len(self.aylar)} Ay için Aidat Tahsilat",
            font=("Segoe UI",13,"bold")).pack(pady=(14,6), padx=20, anchor="w")

        # ── Aidat satırları ───────────────────────────────────────────────────
        lbl(self, "Liste  Aidat Kalemleri", fg=T["text2"],
            font=("Segoe UI",9,"bold")).pack(padx=20, anchor="w", pady=(0,3))

        self._tutar_vars = {}
        scroll_f = frm(self, bg=T["bg3"],
            highlightthickness=1, highlightbackground=T["border"])
        scroll_f.pack(fill="x", padx=20, pady=(0,6))

        canvas = tk.Canvas(scroll_f, bg=T["bg3"],
                           height=min(len(self.aylar)*46+10, 180),
                           highlightthickness=0)
        vsb = ttk.Scrollbar(scroll_f, orient="vertical",
                            command=canvas.yview,
                            style="Dark.Vertical.TScrollbar")
        canvas.configure(yscrollcommand=vsb.set)
        vsb.pack(side="right", fill="y")
        canvas.pack(side="left", fill="both", expand=True)

        inner = frm(canvas, bg=T["bg3"])
        canvas.create_window((0,0), window=inner, anchor="nw")
        inner.bind("<Configure>",
                   lambda e: canvas.configure(scrollregion=canvas.bbox("all")))

        for ay in self.aylar:
            aidat_tut = ay_aidat(self.app.data, ay, self.daire_no)
            gec_faiz  = faiz_hesapla(aidat_tut, b["faiz"], ay, b["son_gun"])
            default_t = round(aidat_tut + gec_faiz, 2)

            row = frm(inner, bg=T["bg3"])
            row.pack(fill="x", padx=10, pady=4)

            lbl(row, f"• {ay_label(ay)}", fg=T["text"],
                bg=T["bg3"], font=("Segoe UI",10)).pack(side="left")

            if gec_faiz > 0:
                lbl(row, f"(+{fmt(gec_faiz)} faiz)",
                    fg=T["orange"], bg=T["bg3"],
                    font=("Segoe UI",8)).pack(side="left", padx=4)

            var = tk.StringVar(value=str(default_t))
            self._tutar_vars[ay] = var
            lbl(row, "₺", fg=T["text2"], bg=T["bg3"],
                font=("Segoe UI",10)).pack(side="right")
            ent(row, textvariable=var, width=10).pack(side="right", padx=4, ipady=4)
            lbl(row, "Tutar:", fg=T["text2"], bg=T["bg3"],
                font=("Segoe UI",8)).pack(side="right")

            tk.Frame(inner, bg=T["border"], height=1).pack(fill="x", padx=10)

        # ── Ortak gider kalemleri ─────────────────────────────────────────────
        # Bu daireye ait, ödenmemiş / kısmen ödenmemiş ortak gider alacaklıları
        self._og_vars = {}   # alacakli_id -> (BooleanVar, tutar_var, kalan)
        og_kayitlar = [
            a for a in self.app.data.get("alacaklilar", [])
            if a.get("tur") == "Ortak Gider"
            and a.get("daire_no") == self.daire_no
            and not a.get("odendi", False)
            and round(a["tutar"] - sum(p["tutar"] for p in a.get("odemeler", [])), 2) > 0
        ]

        if og_kayitlar:
            sep(self)
            lbl(self, "Arac  Ortak Gider Kalemleri  (seçip ödeyebilirsiniz)",
                fg=T["gold2"], font=("Segoe UI",9,"bold")).pack(
                padx=20, anchor="w", pady=(0,3))

            og_scroll_f = frm(self, bg=T["bg4"],
                highlightthickness=1, highlightbackground=T["border"])
            og_scroll_f.pack(fill="x", padx=20, pady=(0,6))

            og_canvas = tk.Canvas(og_scroll_f, bg=T["bg4"],
                                  height=min(len(og_kayitlar)*48+8, 160),
                                  highlightthickness=0)
            og_vsb = ttk.Scrollbar(og_scroll_f, orient="vertical",
                                   command=og_canvas.yview,
                                   style="Dark.Vertical.TScrollbar")
            og_canvas.configure(yscrollcommand=og_vsb.set)
            og_vsb.pack(side="right", fill="y")
            og_canvas.pack(side="left", fill="both", expand=True)

            og_inner = frm(og_canvas, bg=T["bg4"])
            og_canvas.create_window((0,0), window=og_inner, anchor="nw")
            og_inner.bind("<Configure>",
                lambda e: og_canvas.configure(scrollregion=og_canvas.bbox("all")))

            for a in og_kayitlar:
                odenen = sum(p["tutar"] for p in a.get("odemeler", []))
                kalan  = round(a["tutar"] - odenen, 2)
                aid    = a["id"]

                chk_var  = tk.BooleanVar(value=False)
                tut_var  = tk.StringVar(value=str(kalan))
                self._og_vars[aid] = (chk_var, tut_var, kalan, a)

                row = frm(og_inner, bg=T["bg4"])
                row.pack(fill="x", padx=10, pady=4)

                tk.Checkbutton(row, variable=chk_var,
                    bg=T["bg4"], fg=T["gold2"],
                    selectcolor=T["bg3"], activebackground=T["bg4"],
                    cursor="hand2",
                    command=self._hesapla_toplam).pack(side="left", padx=(0,4))

                isim = a.get("kisi") or a.get("aciklama","") or "Ortak Gider"
                lbl(row, f"Arac {isim}", fg=T["gold2"],
                    bg=T["bg4"], font=("Segoe UI",9,"bold")).pack(side="left")

                if odenen > 0:
                    lbl(row, f"  (kalan: {fmt(kalan)} / {fmt(a['tutar'])})",
                        fg=T["orange"], bg=T["bg4"],
                        font=("Segoe UI",8)).pack(side="left", padx=2)
                else:
                    lbl(row, f"  {fmt(kalan)}",
                        fg=T["text2"], bg=T["bg4"],
                        font=("Segoe UI",8)).pack(side="left", padx=2)

                lbl(row, "₺", fg=T["text2"], bg=T["bg4"],
                    font=("Segoe UI",10)).pack(side="right")
                ent(row, textvariable=tut_var, width=10).pack(
                    side="right", padx=4, ipady=4)
                lbl(row, "Öde:", fg=T["text2"], bg=T["bg4"],
                    font=("Segoe UI",8)).pack(side="right")
                tut_var.trace_add("write", lambda *_: self._hesapla_toplam())

                tk.Frame(og_inner, bg=T["border"], height=1).pack(
                    fill="x", padx=10)

        # ── Toplam ────────────────────────────────────────────────────────────
        self._lbl_top = lbl(self, "", fg=T["gold2"],
            font=("Segoe UI",13,"bold"))
        self._lbl_top.pack(padx=20, anchor="w", pady=4)
        self._hesapla_toplam()
        for v in self._tutar_vars.values():
            v.trace_add("write", lambda *_: self._hesapla_toplam())

        # ── Yöntem + not ──────────────────────────────────────────────────────
        form = frm(self)
        form.pack(fill="x", padx=20, pady=(4,8))

        lbl(form, "Ödeme Yöntemi", fg=T["text2"],
            font=("Segoe UI",9,"bold")).pack(anchor="w", pady=(0,3))
        self._yon = tk.StringVar(value="Havale/EFT")
        cmb(form, self._yon, ["Havale/EFT","Nakit","Kredi Kartı"],
            width=26).pack(anchor="w", ipady=5)

        lbl(form, "Not", fg=T["text2"],
            font=("Segoe UI",9,"bold")).pack(anchor="w", pady=(10,3))
        self._not = tk.StringVar()
        ent(form, textvariable=self._not, width=40).pack(fill="x", ipady=6)

        sep(self)
        btn_f = frm(self)
        btn_f.pack(fill="x", padx=20, pady=10)
        btn(btn_f,"X İptal",     self.destroy,  "gray").pack(side="right", padx=(6,0))
        btn(btn_f,"OK Tahsil Et", self._onayla,  "gold").pack(side="right")

    def _hesapla_toplam(self):
        total = 0
        for v in self._tutar_vars.values():
            try: total += float(v.get())
            except: pass
        for aid, (chk_var, tut_var, kalan, a) in getattr(self, "_og_vars", {}).items():
            if chk_var.get():
                try: total += float(tut_var.get())
                except: pass
        self._lbl_top.config(text=f"Toplam: {fmt(total)}")

    def _onayla(self):
        b   = self.app.data["bina"]
        yon = self._yon.get()
        not_= self._not.get().strip()
        total_odendi = 0

        for ay in self.aylar:
            try:
                tutar = float(self._tutar_vars[ay].get())
            except:
                messagebox.showerror("Hata", f"{ay_label(ay)} için geçerli tutar girin.")
                return

            aidat_beklenen = ay_aidat(self.app.data, ay, self.daire_no)
            faiz_bek       = faiz_hesapla(aidat_beklenen, b["faiz"], ay, b["son_gun"])
            tam_beklenen   = round(aidat_beklenen + faiz_bek, 2)

            mkb = yeni_makbuz_no(self.app.data)
            odeme = {
                "id"      : yeni_id(),
                "daireNo" : self.daire_no,
                "ay"      : ay, "tutar": tutar,
                "yontem"  : yon, "not_": not_,
                "tarih"   : simdi(),
                "makbuzNo": mkb,
            }
            self.app.data["odemeler"].append(odeme)
            d_obj = next((x for x in self.app.data["daireler"]
                          if x["no"] == self.daire_no), {})
            d_isim = d_obj.get("isim", "").strip()
            daire_etiketi = (f"Daire {self.daire_no} — {d_isim}"
                             if d_isim else f"Daire {self.daire_no}")
            self.app.data["gelirler"].append({
                "id"      : odeme["id"]+1, "tur": "Aidat",
                "tarih"   : simdi(), "tutar": tutar,
                "aciklama": f"{daire_etiketi} — {ay_label(ay)} aidatı",
                "odeme_id": odeme["id"],
            })
            makbuz_yazdir(self.app.data, odeme)
            total_odendi += tutar

            # Eksik ödeme varsa farkı bu daireye özgü bir sonraki aya ekle
            eksik = round(tam_beklenen - tutar, 2)
            if eksik > 0.01:
                ay_y, ay_m = int(ay[:4]), int(ay[5:7])
                ay_m += 1
                if ay_m > 12: ay_m = 1; ay_y += 1
                sonraki = f"{ay_y}-{ay_m:02d}"
                d_obj = next((x for x in self.app.data["daireler"]
                              if x["no"] == self.daire_no), None)
                if d_obj is not None:
                    daire_ozel = d_obj.setdefault("ozel_aidatlar", {})
                    mevcut_sonraki = ay_aidat(self.app.data, sonraki, self.daire_no)
                    daire_ozel[sonraki] = round(mevcut_sonraki + eksik, 2)

        # ── Seçilen ortak gider ödemelerini işle
        for aid, (chk_var, tut_var, kalan, a) in getattr(self, "_og_vars", {}).items():
            if not chk_var.get():
                continue
            try:
                og_tutar = float(tut_var.get())
                if og_tutar <= 0:
                    raise ValueError
            except:
                messagebox.showerror("Hata",
                    f"'{a.get('kisi','Ortak Gider')}' için geçerli tutar girin.")
                return
            og_tutar = min(og_tutar, kalan)   # kalandan fazlasını alma
            odeme_kayit = {
                "id"    : yeni_id(),
                "tutar" : round(og_tutar, 2),
                "tarih" : simdi(),
                "not_"  : not_ or yon,
            }
            a.setdefault("odemeler", []).append(odeme_kayit)
            yeni_odenen  = sum(p["tutar"] for p in a["odemeler"])
            yeni_kalan   = round(a["tutar"] - yeni_odenen, 2)
            if yeni_kalan <= 0.01:
                a["odendi"]       = True
                a["odeme_tarihi"] = simdi()
            total_odendi += og_tutar
            # Daire bazlı ortak gider ödemesini gelir hesabına yansıt
            gider_adi = a.get("kisi") or a.get("aciklama", "") or "Ortak Gider"
            d_obj2 = next((x for x in self.app.data["daireler"]
                           if x["no"] == self.daire_no), {})
            d_isim2 = d_obj2.get("isim", "").strip()
            daire_etiketi2 = (f"Daire {self.daire_no} — {d_isim2}"
                              if d_isim2 else f"Daire {self.daire_no}")
            self.app.data["gelirler"].append({
                "id"             : odeme_kayit["id"] + 1,
                "tur"            : "Ortak Gider Tahsilatı",
                "tarih"          : simdi(),
                "tutar"          : round(og_tutar, 2),
                "aciklama"       : f"{daire_etiketi2} — {gider_adi}",
                "odeme_id"       : odeme_kayit["id"],
            })
            # Ortak Gider ödemesi için ayrı makbuz oluştur
            og_makbuz_yazdir(self.app.data, a, odeme_kayit)

        # Daire borç güncelle
        d = next((x for x in self.app.data["daireler"]
                  if x["no"] == self.daire_no), None)
        if d:
            d["son_odeme"] = simdi()
            d["borc"] = max(0.0, round(d["borc"] - total_odendi, 2))
            d["faiz"] = max(0.0, round(d["faiz"], 2))

        save_data(self.app.data)
        self.destroy()
        if self.callback: self.callback()

# ── SAKİN DÜZENLE ─────────────────────────────────────────────────────────────
class _SakinDuzenle(tk.Toplevel):
    def __init__(self, parent, app, daire_no, daire, callback=None):
        super().__init__(parent)
        self.app=app; self.daire_no=daire_no; self.callback=callback
        self.title(f"Daire {daire_no} — Sakin Bilgisi")
        self.configure(bg=T["bg"]); self.resizable(False,False)
        self.grab_set()
        self.update_idletasks()
        pr=parent.winfo_rootx(); pt=parent.winfo_rooty()
        self.geometry(f"420x300+{pr+80}+{pt+80}")
        tk.Frame(self,bg=T["gold"],height=3).pack(fill="x")
        lbl(self,f"Kisi  Daire {daire_no} Sakin Bilgisi",
            font=("Segoe UI",13,"bold")).pack(pady=(14,6),padx=20,anchor="w")
        form=frm(self); form.pack(fill="x",padx=20)
        self._vars={}
        for key,txt,val in [
            ("isim","İsim Soyisim",daire.get("isim","")),
            ("tel", "Telefon",     daire.get("tel","")),
            ("email","E-posta",    daire.get("email","")),
        ]:
            lbl(form,txt,fg=T["text2"],font=("Segoe UI",9,"bold")).pack(anchor="w",pady=(10,2))
            var=tk.StringVar(value=val)
            ent(form,textvariable=var,width=36).pack(fill="x",ipady=7)
            self._vars[key]=var
        btn(self,"Kaydet Kaydet",self._kaydet,"gold").pack(pady=16,ipadx=20)

    def _kaydet(self):
        d=next((x for x in self.app.data["daireler"] if x["no"]==self.daire_no),None)
        if d:
            for k,v in self._vars.items(): d[k]=v.get().strip()
        save_data(self.app.data)
        self.destroy()
        if self.callback: self.callback()

# ══════════════════════════════════════════════════════════════════════════════
#  ANA UYGULAMA
# ══════════════════════════════════════════════════════════════════════════════
class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Bina Yönetim Sistemi v3.0")
        self.configure(bg=T["bg"])
        self.geometry("1320x860")
        self.minsize(1020,680)
        apply_style()
        self.update_idletasks()
        sw,sh=self.winfo_screenwidth(),self.winfo_screenheight()
        self.geometry(f"1320x860+{(sw-1320)//2}+{(sh-860)//2}")
        self.data=None
        data=load_data()
        # Uyarı yazısı şablonunu veri dizinine kopyala (varsa)
        _uyari_sablon_hedef = DATA_DIR / "uyari_sablonu.docx"
        if not _uyari_sablon_hedef.exists():
            import shutil
            for _aday in [
                Path(__file__).parent / "APARTMAN_YÖNETİMİNDEN_BİLGİLENDİRME_VE_TALEP_YAZISI.docx",
                Path.home() / "APARTMAN_YÖNETİMİNDEN_BİLGİLENDİRME_VE_TALEP_YAZISI.docx",
            ]:
                if _aday.exists():
                    shutil.copy(_aday, _uyari_sablon_hedef)
                    break
        if data:
            # Eski veri yapısını migrate et
            if "ozel_aidatlar" not in data["bina"]:
                data["bina"]["ozel_aidatlar"] = {}
            if "baslangic" not in data["bina"]:
                data["bina"]["baslangic"] = buay()
            if "ortak_giderler" not in data:
                data["ortak_giderler"] = []
            self.data=data; self._build()
        else:
            self.withdraw()
            VeriYokEkrani(self, self._ilk_kurulum, self._ilk_kurulum)

    def _ilk_kurulum(self, data):
        self.data=data; self.deiconify(); self._build()

    def _build(self):
        self._topbar(); self._navbar()
        self._content=frm(self)
        self._content.pack(fill="both",expand=True)
        self._show("dashboard")

    # ── TOPBAR ────────────────────────────────────────────────────────────────
    def _topbar(self):
        bar=frm(self,bg=T["bg2"],height=62)
        bar.pack(fill="x"); bar.pack_propagate(False)
        tk.Frame(bar,bg=T["gold"],width=4).pack(side="left",fill="y")
        left=frm(bar,bg=T["bg2"])
        left.pack(side="left",padx=14,pady=8)
        lbl(left,"Bina",font=("Segoe UI Emoji",22),bg=T["bg2"],fg=T["gold"]).pack(side="left",padx=(0,10))
        info=frm(left,bg=T["bg2"]); info.pack(side="left")
        b=self.data["bina"]
        lbl(info,b["adi"],font=("Segoe UI",13,"bold"),bg=T["bg2"]).pack(anchor="w")
        lbl(info,
            f"{b['daire']} Daire  •  Aidat: {fmt(b['aidat'])}  •  "
            f"Faiz: %{b['faiz']}  •  Son Gün: {b['son_gun']}  •  "
            f"Başlangıç: {ay_label(b['baslangic'])}",
            fg=T["text2"],font=("Segoe UI",9),bg=T["bg2"]).pack(anchor="w")
        right=frm(bar,bg=T["bg2"]); right.pack(side="right",padx=14)
        lbl(right,datetime.date.today().strftime("Takvim %B %Y"),
            bg=T["bg3"],fg=T["gold2"],font=("Segoe UI",9,"bold"),
            padx=10,pady=5).pack(side="left",padx=6)

    # ── NAVBAR ────────────────────────────────────────────────────────────────
    def _navbar(self):
        nav=frm(self,bg=T["bg2"]); nav.pack(fill="x")
        tk.Frame(nav,bg=T["border"],height=1).pack(fill="x")
        tabs=[
            ("dashboard",      "Grafik  Genel Bakış"),
            ("daireler",       "Ev  Daireler"),
            ("gelir_gider",    "Para  Gelir / Gider"),
            ("ortak_giderler", "Arac  Ortak Giderler"),
            ("alacaklilar",    "Liste  Alacaklılar"),
            ("raporlar",       "Yukari  Raporlar"),
            ("detay_rapor",    "Klasor  Detaylı Raporlar"),
            ("ayarlar",        "Ayar  Ayarlar"),
        ]
        self._nav_btns={}
        tb=frm(nav,bg=T["bg2"]); tb.pack(fill="x",padx=8)
        for key,label in tabs:
            b=tk.Button(tb,text=label,
                font=("Segoe UI",10),bg=T["bg2"],fg=T["text2"],
                activebackground=T["bg3"],relief="flat",bd=0,
                cursor="hand2",padx=16,pady=10,
                command=lambda k=key: self._show(k))
            b.pack(side="left")
            self._nav_btns[key]=b

    def _show(self, key, skip_sifre=False):
        # 2. Ayarlar sekmesine şifre koruması
        if key == "ayarlar" and not skip_sifre:
            if not self._sifre_dogrula():
                return
        for k,b in self._nav_btns.items():
            if k==key: b.config(fg=T["gold2"],bg=T["bg3"],font=("Segoe UI",10,"bold"))
            else:       b.config(fg=T["text2"],bg=T["bg2"],font=("Segoe UI",10))
        for w in self._content.winfo_children(): w.destroy()
        {
            "dashboard":      self._tab_dashboard,
            "daireler":       self._tab_daireler,
            "gelir_gider":    self._tab_gelir_gider,
            "ortak_giderler": self._tab_ortak_giderler,
            "alacaklilar":    self._tab_alacaklilar,
            "raporlar":       self._tab_raporlar,
            "detay_rapor":    self._tab_detay_rapor,
            "ayarlar":        self._tab_ayarlar,
        }.get(key, self._tab_dashboard)()
        self._aktif_tab = key

    def _sifre_dogrula(self):
        """Şifre doğrulama penceresi. Doğruysa True döner."""
        kayitli = self.data["bina"].get("ayar_sifre", "1234")
        result  = [False]

        dlg = tk.Toplevel(self)
        dlg.title("Kilitli  Ayarlar — Şifre Gerekli")
        dlg.configure(bg=T["bg"])
        dlg.resizable(False, False)
        dlg.grab_set()
        dlg.update_idletasks()
        sw, sh = dlg.winfo_screenwidth(), dlg.winfo_screenheight()
        dlg.geometry(f"360x250+{(sw-360)//2}+{(sh-250)//2}")

        tk.Frame(dlg, bg=T["gold"], height=3).pack(fill="x")
        lbl(dlg, "Kilitli  Ayarlara Giriş",
            font=("Segoe UI",13,"bold")).pack(pady=(16,4), padx=20, anchor="w")
        lbl(dlg, "Ayarlar sekmesine erişmek için şifrenizi girin.",
            fg=T["text2"], font=("Segoe UI",9)).pack(padx=20, anchor="w")
        lbl(dlg, "İlk şifre: 1234",
            fg=T["text3"], font=("Segoe UI",8)).pack(padx=20, anchor="w", pady=(2,12))

        sifre_var = tk.StringVar()
        e = ent(dlg, textvariable=sifre_var, width=26)
        e.config(show="●")
        e.pack(padx=20, ipady=8, fill="x")
        e.focus_set()

        msg_lbl = lbl(dlg, "", fg=T["red"], font=("Segoe UI",9))
        msg_lbl.pack(padx=20, pady=(4,0), anchor="w")

        def dogrula(event=None):
            if sifre_var.get() == kayitli:
                result[0] = True
                dlg.destroy()
            else:
                msg_lbl.config(text="X  Yanlış şifre, tekrar deneyin.")
                sifre_var.set("")
                e.focus_set()

        bf = frm(dlg); bf.pack(fill="x", padx=20, pady=14)
        btn(bf, "X İptal", dlg.destroy, "gray").pack(side="right", padx=(6,0))
        btn(bf, "OK Giriş",  dogrula,    "gold").pack(side="right")
        e.bind("<Return>", dogrula)
        dlg.wait_window()
        return result[0]

    def refresh_all(self):
        key = getattr(self,"_aktif_tab","dashboard")
        # Ayarlar aktifken refresh_all şifre sormadan yenilesin
        self._show(key, skip_sifre=(key=="ayarlar"))

    # ── DASHBOARD ─────────────────────────────────────────────────────────────
    def _tab_dashboard(self):
        outer=frm(self._content)
        outer.pack(fill="both",expand=True,padx=18,pady=14)
        b=self.data["bina"]; ba=buay()

        # 3. Bu ay içinde yapılan TÜM ödemeler (geçmiş aylar dahil)
        tah=sum(o["tutar"] for o in self.data["odemeler"]
                if o["tarih"][:7]==ba)   # ödemenin yapıldığı tarih bu ay
        bek=b["daire"]*b["aidat"]
        oran=int(tah/bek*100) if bek else 0

        # 2+3. Toplam borç = daire borçları + alacaklıların KALAN tutarı
        topB=sum(d["borc"] for d in self.data["daireler"])
        topF=sum(d["faiz"] for d in self.data["daireler"])
        topAlacak=sum(
            max(0.0, round(a["tutar"] - sum(p["tutar"] for p in a.get("odemeler",[])), 2))
            for a in self.data.get("alacaklilar",[])
            if not a.get("odendi", False)
        )
        topBorcToplam=topB+topF+topAlacak

        topG=sum(g["tutar"] for g in self.data["gelirler"])
        topGd=sum(g["tutar"] for g in self.data["giderler"])
        net=topG-topGd
        # 4. Ödenmemiş aidat toplamı
        tum_ay=ay_listesi(b.get("baslangic",ba))
        odenmemis_toplam=sum(
            ay_aidat(self.data,ay,d["no"])+faiz_hesapla(ay_aidat(self.data,ay,d["no"]),b["faiz"],ay,b["son_gun"])
            for ay in tum_ay
            for d in self.data["daireler"]
            if not any(o["daireNo"]==d["no"] and o["ay"]==ay for o in self.data["odemeler"])
        )

        stats=frm(outer); stats.pack(fill="x",pady=(0,12))
        cards=[
            ("Uyari  Toplam Borç",    fmt(topBorcToplam),
             f"Aidat:{fmt(topB+topF)}  Alacak:{fmt(topAlacak)}", "red"),
            ("Kart  Bu Ay Tahsilat", fmt(tah),
             f"%{oran}  •  Beklenen:{fmt(bek)}", "gold"),
            ("Yukari  Toplam Gelir",   fmt(topG), "Tüm zamanlar", "green"),
            ("Canta  Net Bakiye",     fmt(net),  "Gelir — Gider",
             "blue" if net>=0 else "red"),
            ("Saat  Ödenmemiş Aidat",fmt(odenmemis_toplam),
             "Tüm geçmiş ödenmemiş dönemler", "orange"),
        ]
        for i,(title,val,sub,color) in enumerate(cards):
            f,_,_=stat_card(stats,title,val,sub,color)
            f.grid(row=0,column=i,padx=5,sticky="nsew")
            stats.columnconfigure(i,weight=1)

        pw=pane(outer,orient="horizontal")
        pw.pack(fill="both",expand=True)

        sol=card_panel(pw,"Kart  Bu Ay Yapılan Ödemeler","Bu ay tahsil edilenler (tüm dönemler)")
        pw.add(sol,weight=1)
        si=frm(sol,bg=T["bg2"]); si.pack(fill="both",expand=True,padx=8,pady=8)
        sf,ts=scrolled(si,[("Daire",70),("Dönem",130),("Yöntem",100),("Tutar",100),("Tarih",110)])
        sf.pack(fill="both",expand=True)
        bu_ay_yapilan=[o for o in self.data["odemeler"] if o["tarih"][:7]==ba]
        for o in reversed(bu_ay_yapilan[-14:]):
            ts.insert("","end",tags=("odendi",),
                values=(f"Daire {o['daireNo']}",ay_label(o["ay"]),
                        o["yontem"],fmt(o["tutar"]),tarih_str(o["tarih"])))
        if not bu_ay_yapilan:
            ts.insert("","end",tags=("normal",),
                values=("Bu ay henüz ödeme yok","","","",""))

        sag=card_panel(pw,"Uyari  Gecikmiş Aidatlar","Tüm ödenmemiş dönemler — daire ve ay bazında")
        pw.add(sag,weight=1)
        ri=frm(sag,bg=T["bg2"]); ri.pack(fill="both",expand=True,padx=8,pady=8)
        rf,tr=scrolled(ri,[
            ("Daire",80),("Sakin",130),("Dönem",120),
            ("Aidat",100),("Faiz",90),("Toplam",100)])
        rf.pack(fill="both",expand=True)
        tum_ay=ay_listesi(b.get("baslangic",ba))
        satirsayac=0
        for ay in reversed(tum_ay):
            for d in self.data["daireler"]:
                odendi=any(o["daireNo"]==d["no"] and o["ay"]==ay
                           for o in self.data["odemeler"])
                if not odendi:
                    aidat_tut=ay_aidat(self.data,ay,d["no"])
                    faiz_tut=faiz_hesapla(aidat_tut,b["faiz"],ay,b["son_gun"])
                    isim=d.get("isim") or "—"
                    tr.insert("","end",tags=("borclu",),
                        values=(f"Daire {d['no']}",isim,ay_label(ay),
                                fmt(aidat_tut),
                                fmt(faiz_tut) if faiz_tut else "—",
                                fmt(aidat_tut+faiz_tut)))
                    satirsayac+=1
        if satirsayac==0:
            tr.insert("","end",tags=("odendi",),
                values=("OK Gecikmiş aidat yok","","","","",""))

    # ── DAİRELER ──────────────────────────────────────────────────────────────
    def _tab_daireler(self):
        outer=frm(self._content)
        outer.pack(fill="both",expand=True,padx=18,pady=14)

        # ── Başlık + üst butonlar ─────────────────────────────────────────
        hdr=frm(outer); hdr.pack(fill="x",pady=(0,8))
        lbl(hdr,"Ev  Daire Listesi  —  Çift tıkla: detay / tahsilat ekranı",
            font=("Segoe UI",13,"bold")).pack(side="left")
        bf=frm(hdr); bf.pack(side="right")
        btn(bf," Sakin Listesi",self._sakin_listesi,"blue").pack(side="left",padx=4)
        btn(bf,"Takvim Özel Aidat Yönetimi",
            lambda: OzelAidatPencere(self,self.data,callback=self.refresh_all),
            "purple").pack(side="left",padx=4)
        btn(bf,"Yenile Faiz Hesapla",self._faiz,"orange").pack(side="left",padx=4)

        pw=pane(outer,orient="vertical")
        pw.pack(fill="both",expand=True)

        # ── Üst: daire listesi ────────────────────────────────────────────
        ust=frm(pw); pw.add(ust,weight=60)
        tbl_f,self._tv=scrolled(ust,[
            ("Daire",70),("Sakin",150),("Bu Ay",100),
            ("Borçlu Ay",90),("Anapara",110),
            ("Faiz",100),("Toplam Borç",115),("Son Ödeme",115),
        ],height=14)
        tbl_f.pack(fill="both",expand=True)
        self._tv.bind("<Double-1>",self._daire_cift)
        self._tv.bind("<<TreeviewSelect>>", self._daire_sec_guncelle)

        # ── Alt: iki sekmeli panel ────────────────────────────────────────
        alt_nb_frm = frm(pw); pw.add(alt_nb_frm, weight=40)
        alt_nb = ttk.Notebook(alt_nb_frm, style="Dark.TNotebook")
        alt_nb.pack(fill="both", expand=True)

        # ── ALT SEKME 1: Genel İşlemler ──────────────────────────────────
        t_genel = frm(alt_nb, bg=T["bg2"])
        alt_nb.add(t_genel, text="    Seçili Daire  ")

        gi = frm(t_genel, bg=T["bg2"])
        gi.pack(fill="both", expand=True, padx=14, pady=10)

        # Daire bilgisi etiketi
        self._daire_bilgi_lbl = lbl(gi,
            "Listeden bir daire seçin.",
            fg=T["text3"], bg=T["bg2"], font=("Segoe UI",10))
        self._daire_bilgi_lbl.pack(anchor="w", pady=(0,8))

        gb = frm(gi, bg=T["bg2"]); gb.pack(anchor="w")
        btn(gb,"Klasor Detay & Tahsilat", self._daire_ac,  "gold").pack(side="left",padx=(0,8))
        btn(gb,"Kisi Sakin Düzenle",    self._sakin_ac,  "gray").pack(side="left",padx=(0,8))

        # ── ALT SEKME 2: Makbuz Merkezi ───────────────────────────────────
        t_makbuz = frm(alt_nb, bg=T["bg2"])
        alt_nb.add(t_makbuz, text="  Yazdir  Makbuz Merkezi  ")

        # Üst: seçili daire + filtre çubuğu
        mk_top = frm(t_makbuz, bg=T["bg3"],
            highlightthickness=1, highlightbackground=T["border"])
        mk_top.pack(fill="x", padx=10, pady=(10,0))
        mk_top_i = frm(mk_top, bg=T["bg3"])
        mk_top_i.pack(fill="x", padx=12, pady=8)

        # Sol: daire bilgisi
        self._mk_daire_lbl = lbl(mk_top_i,
            "Listeden bir daire seçin",
            fg=T["gold2"], bg=T["bg3"], font=("Segoe UI",10,"bold"))
        self._mk_daire_lbl.pack(side="left")

        # Sağ: tür filtresi
        fil_frm = frm(mk_top_i, bg=T["bg3"])
        fil_frm.pack(side="right")
        lbl(fil_frm,"Göster:",fg=T["text3"],bg=T["bg3"],
            font=("Segoe UI",8,"bold")).pack(side="left",padx=(0,4))
        self._mk_filtre = tk.StringVar(value="Tümü")
        for val in ["Tümü","Aidat","Ortak Gider"]:
            tk.Radiobutton(fil_frm, text=val, variable=self._mk_filtre, value=val,
                bg=T["bg3"], fg=T["text"], selectcolor=T["bg4"],
                activebackground=T["bg3"], font=("Segoe UI",9),
                cursor="hand2",
                command=self._mk_listele).pack(side="left", padx=4)

        # Makbuz listesi
        mk_tbl_frm = frm(t_makbuz, bg=T["bg2"])
        mk_tbl_frm.pack(fill="both", expand=True, padx=10, pady=(6,0))

        mk_list_hdr = frm(mk_tbl_frm, bg=T["bg2"]); mk_list_hdr.pack(fill="x")
        tk.Frame(mk_list_hdr, bg=T["gold"], height=2).pack(fill="x")
        mk_lhf = frm(mk_list_hdr, bg=T["bg2"]); mk_lhf.pack(fill="x", padx=10, pady=5)
        lbl(mk_lhf,"Makbuz  Makbuz Listesi",
            font=("Segoe UI",10,"bold"),bg=T["bg2"]).pack(side="left")
        lbl(mk_lhf,"  Çift tıkla veya butona bas → makbuzu aç",
            fg=T["text3"],bg=T["bg2"],font=("Segoe UI",8)).pack(side="left",padx=6)

        mk_tf, self._mk_tree = scrolled(mk_tbl_frm, [
            ("Tür",        100),
            ("Dönem/Gider",180),
            ("Tutar",       95),
            ("Tarih",      110),
            ("Makbuz No",  150),
            ("Not",        160),
        ], height=7)
        mk_tf.pack(fill="both", expand=True)
        self._mk_tree.bind("<Double-1>", lambda e: self._mk_yazdir())

        # Alt buton çubuğu
        mk_btn_row = frm(t_makbuz, bg=T["bg2"])
        mk_btn_row.pack(fill="x", padx=10, pady=(6,10))

        btn(mk_btn_row,"Yazdir  Makbuzu Yazdır",
            self._mk_yazdir, "gold").pack(side="left", ipadx=10, ipady=4)
        btn(mk_btn_row,"Duzenle  Düzenle / Sil",
            self._mk_gecmis_duzenle, "blue").pack(side="left", padx=6, ipadx=8, ipady=4)
        btn(mk_btn_row,"Yazdir  Tüm Aidatlar",
            self._mk_tumunu_yazdir_aidat, "gray").pack(side="left", padx=6, ipadx=8, ipady=4)
        btn(mk_btn_row,"Yazdir  Tüm Ortak Giderler",
            self._mk_tumunu_yazdir_og, "purple").pack(side="left", ipadx=8, ipady=4)

        # ── ALT SEKME 3: Uyarı Yazısı ────────────────────────────────────
        t_uyari = frm(alt_nb, bg=T["bg2"])
        alt_nb.add(t_uyari, text="  Mektup  Uyarı Yazısı  ")

        uw_top = frm(t_uyari, bg=T["bg3"],
            highlightthickness=1, highlightbackground=T["border"])
        uw_top.pack(fill="x", padx=10, pady=(10,0))
        uw_top_i = frm(uw_top, bg=T["bg3"])
        uw_top_i.pack(fill="x", padx=12, pady=8)

        self._uw_daire_lbl = lbl(uw_top_i,
            "Listeden bir daire seçin",
            fg=T["gold2"], bg=T["bg3"], font=("Segoe UI",10,"bold"))
        self._uw_daire_lbl.pack(side="left")

        self._uw_borc_lbl = lbl(uw_top_i,
            "",
            fg=T["red"], bg=T["bg3"], font=("Segoe UI",10))
        self._uw_borc_lbl.pack(side="left", padx=(16,0))

        uw_info_frm = frm(t_uyari, bg=T["bg2"])
        uw_info_frm.pack(fill="x", padx=10, pady=(8,2))
        lbl(uw_info_frm,
            "Seçili daire için şablon Word belgesi doldurularak PDF olarak kaydedilir.",
            fg=T["text3"], bg=T["bg2"], font=("Segoe UI",9)).pack(side="left")

        uw_btn_row = frm(t_uyari, bg=T["bg2"])
        uw_btn_row.pack(fill="x", padx=10, pady=(6,10))
        btn(uw_btn_row, "Dosya Uyarı Yazısı Oluştur (PDF)",
            self._daire_uyari_yazisi_pdf, "gold").pack(side="left", ipadx=10, ipady=4)
        btn(uw_btn_row, "Not Word Olarak Kaydet",
            self._daire_uyari_yazisi_docx, "blue").pack(side="left", padx=8, ipadx=8, ipady=4)

        self._daire_doldur()

    # ── MAkbuz Merkezi yardımcıları ───────────────────────────────────────
    def _daire_sec_guncelle(self, event=None):
        """Daire listesinde seçim değişince genel bilgi etiketlerini ve makbuz listesini güncelle."""
        sel = self._tv.selection()
        if not sel:
            if hasattr(self,"_daire_bilgi_lbl"):
                self._daire_bilgi_lbl.config(text="Listeden bir daire seçin.")
            if hasattr(self,"_mk_daire_lbl"):
                self._mk_daire_lbl.config(text="Listeden bir daire seçin")
            return
        no = int(sel[0])
        d  = next((x for x in self.data["daireler"] if x["no"] == no), {})
        isim = d.get("isim","") or "—"
        # Genel sekme etiketi
        if hasattr(self,"_daire_bilgi_lbl"):
            b   = self.data["bina"]
            ba  = buay()
            tum = ay_listesi(b.get("baslangic",ba), ba)
            borclu = sum(1 for ay in tum
                         if not any(o["daireNo"]==no and o["ay"]==ay
                                    for o in self.data["odemeler"]))
            durum = f"✘ {borclu} ödenmemiş ay" if borclu else "v Tüm aidatlar ödendi"
            self._daire_bilgi_lbl.config(
                text=f"Daire {no}  —  {isim}  •  {durum}")
        # Makbuz sekmesi etiketi
        if hasattr(self,"_mk_daire_lbl"):
            self._mk_daire_lbl.config(text=f"Ev  Daire {no}  —  {isim}")
        self._mk_listele()
        # Uyarı yazısı sekmesi etiketleri
        if hasattr(self,"_uw_daire_lbl"):
            self._uw_daire_lbl.config(text=f"Ev  Daire {no}  —  {isim}")
        if hasattr(self,"_uw_borc_lbl"):
            _b  = self.data["bina"]
            _ba = buay()
            _tum = ay_listesi(_b.get("baslangic", _ba), _ba)
            _bugun2   = datetime.date.today()
            _son_gun2 = _b.get("son_gun", 10)
            _odenmemis = [
                ay for ay in _tum
                if not any(o["daireNo"] == no and o["ay"] == ay
                           for o in self.data["odemeler"])
                and not (ay == buay() and _bugun2.day < _son_gun2)
            ]
            _anapara = round(sum(ay_aidat(self.data, ay, no) for ay in _odenmemis), 2)
            _faiz    = round(sum(faiz_hesapla(ay_aidat(self.data, ay, no),
                                 _b["faiz"], ay, _b["son_gun"])
                                 for ay in _odenmemis), 2)
            _toplam  = round(_anapara + _faiz, 2)
            if _toplam > 0:
                self._uw_borc_lbl.config(text=f"Toplam Borç: {fmt(_toplam)}")
            else:
                self._uw_borc_lbl.config(text="v Borç yok")

    def _mk_seçili_daire_no(self):
        sel = self._tv.selection()
        return int(sel[0]) if sel else None

    def _mk_listele(self):
        """Makbuz listesini seçili daireye + filtreye göre doldur."""
        if not hasattr(self,"_mk_tree"): return
        self._mk_tree.delete(*self._mk_tree.get_children())
        no = self._mk_seçili_daire_no()
        if no is None: return

        filtre = self._mk_filtre.get()

        # Aidat ödemeleri
        if filtre in ("Tümü","Aidat"):
            aidatlar = [o for o in self.data["odemeler"] if o["daireNo"] == no]
            for o in reversed(aidatlar):
                self._mk_tree.insert("","end",
                    iid=f"a_{o['id']}", tags=("odendi",),
                    values=(
                        "Kart Aidat",
                        ay_label(o["ay"]),
                        fmt(o["tutar"]),
                        tarih_str(o["tarih"]),
                        o["makbuzNo"],
                        o.get("not_","") or "—",
                    ))

        # Ortak Gider ödemeleri
        if filtre in ("Tümü","Ortak Gider"):
            og_al = [a for a in self.data.get("alacaklilar",[])
                     if a.get("tur")=="Ortak Gider" and a.get("daire_no")==no]
            for a in og_al:
                gider_adi = a.get("kisi") or a.get("aciklama","") or "Ortak Gider"
                for p in reversed(a.get("odemeler",[])):
                    mkb_no = f"OG-{datetime.date.today().year}-{p['id'] % 10000:04d}"
                    self._mk_tree.insert("","end",
                        iid=f"og_{a['id']}_{p['id']}", tags=("ozel",),
                        values=(
                            "Arac Ortak Gider",
                            gider_adi,
                            fmt(p["tutar"]),
                            tarih_str(p["tarih"]),
                            mkb_no,
                            p.get("not_","") or "—",
                        ))

    def _mk_yazdir(self):
        """Seçili makbuzu yazdır."""
        if not hasattr(self,"_mk_tree"): return
        sel = self._mk_tree.selection()
        if not sel:
            messagebox.showinfo("Seçim","Lütfen bir makbuz seçin.",parent=self)
            return
        iid = sel[0]
        if iid.startswith("a_"):
            oid = int(iid[2:])
            odeme = next((o for o in self.data["odemeler"] if o["id"]==oid), None)
            if odeme:
                makbuz_yazdir(self.data, odeme)
        elif iid.startswith("og_"):
            parts = iid.split("_")
            # og_{alacakli_id}_{odeme_id}
            al_id = int(parts[1])
            p_id  = int(parts[2])
            al = next((a for a in self.data.get("alacaklilar",[]) if a["id"]==al_id), None)
            if al:
                p_kaydi = next((p for p in al.get("odemeler",[]) if p["id"]==p_id), None)
                if p_kaydi:
                    og_makbuz_yazdir(self.data, al, p_kaydi)

    def _mk_gecmis_duzenle(self):
        """Makbuz merkezinden seçili ortak gider ödemesini düzenle/sil."""
        if not hasattr(self, "_mk_tree"): return
        sel = self._mk_tree.selection()
        if not sel:
            messagebox.showinfo("Seçim", "Lütfen bir kayıt seçin.", parent=self); return
        iid = sel[0]
        if iid.startswith("og_"):
            parts = iid.split("_")
            al_id = int(parts[1])
            p_id  = int(parts[2])
            al = next((a for a in self.data.get("alacaklilar",[]) if a["id"]==al_id), None)
            if al:
                p_kaydi = next((p for p in al.get("odemeler",[]) if p["id"]==p_id), None)
                if p_kaydi:
                    AlacakliGecmisPencere(self, al, data=self.data,
                        callback=lambda: self._mk_listele())
                    return
        elif iid.startswith("a_"):
            messagebox.showinfo("Bilgi",
                "Aidat odemeleri icin Daire Detay ekranini kullanin.",
                parent=self)
            return
        messagebox.showinfo("Secim","Lutfen bir ortak gider odeme kaydi secin.",parent=self)

    def _mk_tumunu_yazdir_aidat(self):
        """Seçili dairedeki tüm aidat makbuzlarını toplu HTML sayfasında aç."""
        no = self._mk_seçili_daire_no()
        if no is None:
            messagebox.showinfo("Seçim","Lütfen önce listeden bir daire seçin.",parent=self)
            return
        aidatlar = [o for o in self.data["odemeler"] if o["daireNo"]==no]
        if not aidatlar:
            messagebox.showinfo("Bilgi","Bu daireye ait aidat makbuzu yok.",parent=self)
            return
        d    = next((x for x in self.data["daireler"] if x["no"]==no), {})
        b    = self.data["bina"]
        isim = d.get("isim","") or "—"
        rows = ""
        for o in sorted(aidatlar, key=lambda x: x["ay"]):
            rows += f"""
            <div class="makbuz">
              <div class="mbadge">Kart AİDAT MAKBUZU — {o['makbuzNo']}</div>
              <table>
                <tr><td>Daire</td><td><b>Daire {o['daireNo']}</b></td></tr>
                <tr><td>Sakin</td><td>{isim}</td></tr>
                <tr><td>Dönem</td><td>{ay_label(o['ay'])}</td></tr>
                <tr><td>Ödeme Tarihi</td><td>{tarih_str(o['tarih'])}</td></tr>
                <tr><td>Yöntem</td><td>{o['yontem']}</td></tr>
                {"<tr><td>Not</td><td>" + (o.get('not_') or '') + "</td></tr>" if o.get('not_') else ""}
              </table>
              <div class="tutar"><p>ÖDENEN TUTAR</p><h3>{fmt(o['tutar'])}</h3></div>
            </div>"""
        html = _toplu_makbuz_html(b, isim, no, "Aidat Makbuzları", rows)
        _makbuz_tarayici_ac(html)

    def _mk_tumunu_yazdir_og(self):
        """Seçili dairedeki tüm ortak gider ödeme makbuzlarını toplu HTML sayfasında aç."""
        no = self._mk_seçili_daire_no()
        if no is None:
            messagebox.showinfo("Seçim","Lütfen önce listeden bir daire seçin.",parent=self)
            return
        og_al = [a for a in self.data.get("alacaklilar",[])
                 if a.get("tur")=="Ortak Gider" and a.get("daire_no")==no
                 and a.get("odemeler")]
        if not og_al:
            messagebox.showinfo("Bilgi","Bu daireye ait ortak gider ödemesi yok.",parent=self)
            return
        d    = next((x for x in self.data["daireler"] if x["no"]==no), {})
        b    = self.data["bina"]
        isim = d.get("isim","") or "—"
        rows = ""
        for a in og_al:
            gider_adi = a.get("kisi") or a.get("aciklama","") or "Ortak Gider"
            toplam    = a["tutar"]
            for p in sorted(a.get("odemeler",[]), key=lambda x: x["tarih"]):
                mkb_no = f"OG-{datetime.date.today().year}-{p['id'] % 10000:04d}"
                odenen_o = sum(pp["tutar"] for pp in a.get("odemeler",[])
                               if pp["tarih"] <= p["tarih"])
                kalan_o  = round(toplam - odenen_o, 2)
                rows += f"""
            <div class="makbuz og">
              <div class="mbadge og-badge">Arac ORTAK GİDER MAKBUZU — {mkb_no}</div>
              <div class="gider-adi">{gider_adi}</div>
              <table>
                <tr><td>Daire</td><td><b>Daire {no}</b></td></tr>
                <tr><td>Sakin</td><td>{isim}</td></tr>
                <tr><td>Ödeme Tarihi</td><td>{tarih_str(p['tarih'])}</td></tr>
                <tr><td>Gider Toplam</td><td>{fmt(toplam)}</td></tr>
                {"<tr><td>Not</td><td>" + (p.get('not_') or '') + "</td></tr>" if p.get('not_') else ""}
              </table>
              <div class="tutar og-tutar"><p>BU ÖDEMEDE ÖDENEN</p><h3>{fmt(p['tutar'])}</h3></div>
              {"<div class='kalan-kart kalan-var'><p>Kalan Borç</p><h4>" + fmt(kalan_o) + "</h4></div>" if kalan_o > 0.01 else "<div class='kalan-kart kalan-tamam'><p>OK Borç Tamamen Kapatıldı</p><h4>v Sıfırlandı</h4></div>"}
            </div>"""
        html = _toplu_makbuz_html(b, isim, no, "Ortak Gider Makbuzları", rows)
        _makbuz_tarayici_ac(html)

    def _daire_doldur(self):
        if not hasattr(self,"_tv"): return
        self._tv.delete(*self._tv.get_children())
        ba=buay(); b=self.data["bina"]
        # Yalnızca geçmiş + bu ay (ileri dönem dahil edilmez)
        tum=ay_listesi(b.get("baslangic",buay()), ba)
        for d in self.data["daireler"]:
            odeme_bu_ay=next((o for o in self.data["odemeler"]
                              if o["daireNo"]==d["no"] and o["ay"]==ba),None)
            # Ödenmemiş ayları bul
            odenmemis_aylar=[
                ay for ay in tum
                if not any(o["daireNo"]==d["no"] and o["ay"]==ay
                           for o in self.data["odemeler"])
            ]
            borclu_ay=len(odenmemis_aylar)
            # Gerçek anapara ve faiz: her ödenmemiş ay için hesapla
            anapara=sum(
                ay_aidat(self.data, ay, d["no"])
                for ay in odenmemis_aylar
            )
            faiz_top=sum(
                faiz_hesapla(ay_aidat(self.data, ay, d["no"]),
                             b["faiz"], ay, b["son_gun"])
                for ay in odenmemis_aylar
            )
            anapara=round(anapara, 2)
            faiz_top=round(faiz_top, 2)
            top=round(anapara+faiz_top, 2)

            if odeme_bu_ay: tag,durum="odendi","v Ödendi"
            elif borclu_ay>0: tag,durum="borclu","✘ Borçlu"
            else: tag,durum="bekliyor","Uyari Bekliyor"

            self._tv.insert("","end",iid=str(d["no"]),tags=(tag,),
                values=(f"Daire {d['no']}",d.get("isim") or "—",durum,
                        f"{borclu_ay} ay" if borclu_ay else "—",
                        fmt(anapara) if anapara else "—",
                        fmt(faiz_top) if faiz_top else "—",
                        fmt(top) if top else "—",
                        tarih_str(d.get("son_odeme","")) if d.get("son_odeme") else "—"))

    def _daire_cift(self,e):
        item=self._tv.identify_row(e.y)
        if item: DaireDetay(self,self,int(item))

    def _daire_ac(self):
        sel=self._tv.selection()
        if not sel: messagebox.showinfo("Seçim","Lütfen bir daire seçin."); return
        DaireDetay(self,self,int(sel[0]))

    def _sakin_ac(self):
        sel=self._tv.selection()
        if not sel: messagebox.showinfo("Seçim","Lütfen bir daire seçin."); return
        no=int(sel[0])
        d=next((x for x in self.data["daireler"] if x["no"]==no),{})
        _SakinDuzenle(self,self,no,d,callback=self._daire_doldur)

    def _sakin_listesi(self):
        """Tüm bina sakinlerinin listesini göster."""
        SakinListePencere(self, self.data)

    def _faiz(self):
        b=self.data["bina"]; ba=buay()
        today=datetime.date.today()
        ay_y,ay_m=int(ba[:4]),int(ba[5:])
        son_gun=b["son_gun"]
        gecikti=(today.year>ay_y or
                 (today.year==ay_y and today.month>ay_m) or
                 (today.year==ay_y and today.month==ay_m and today.day>son_gun))
        if not gecikti:
            messagebox.showinfo("Bilgi","Son ödeme tarihi henüz geçmedi."); return
        if not b["faiz"]:
            messagebox.showinfo("Bilgi","Faiz oranı 0."); return
        sayac=0
        for d in self.data["daireler"]:
            if not any(o["daireNo"]==d["no"] and o["ay"]==ba for o in self.data["odemeler"]):
                d["borc"]=round(d["borc"]+b["aidat"],2)
                d["faiz"]=round(d["faiz"]+b["aidat"]*b["faiz"]/100,2)
                sayac+=1
        save_data(self.data); self._daire_doldur()
        messagebox.showinfo("Tamam",f"{sayac} daire için faiz uygulandı.")

    # ── GELİR / GİDER ─────────────────────────────────────────────────────────
    def _tab_gelir_gider(self):
        outer=frm(self._content)
        outer.pack(fill="both",expand=True,padx=18,pady=14)
        nb=ttk.Notebook(outer,style="Dark.TNotebook")
        nb.pack(fill="both",expand=True)
        for tip,label,turler,color in [
            ("gelir","  Yukari  Gelirler  ",
             ["Aidat","Kira","Banka Faizi","Diğer"],"green"),
            ("gider","  Asagi  Giderler  ",
             ["Elektrik","Su","Doğalgaz","Asansör Bakım",
              "Temizlik","Güvenlik","Tamir/Tadilat","Sigorta","Diğer"],"red"),
        ]:
            tab=frm(nb,bg=T["bg2"]); nb.add(tab,text=label)
            self._ie_panel(tab,tip,turler,color)

    def _ie_panel(self,parent,tip,turler,color):
        fw=frm(parent,bg=T["bg3"],
               highlightthickness=1,highlightbackground=T["border"])
        fw.pack(fill="x",padx=12,pady=12)
        form=frm(fw,bg=T["bg3"]); form.pack(fill="x",padx=12,pady=10)

        # Başlık satırı
        lbl(form,"Tür",         fg=T["text2"],bg=T["bg3"],font=("Segoe UI",9,"bold")).grid(row=0,column=0,sticky="w",padx=4)
        lbl(form,"Tutar (₺)",   fg=T["text2"],bg=T["bg3"],font=("Segoe UI",9,"bold")).grid(row=0,column=1,sticky="w",padx=4)
        lbl(form,"Fatura Tarihi",fg=T["text2"],bg=T["bg3"],font=("Segoe UI",9,"bold")).grid(row=0,column=2,sticky="w",padx=4)
        lbl(form,"Açıklama",    fg=T["text2"],bg=T["bg3"],font=("Segoe UI",9,"bold")).grid(row=0,column=3,sticky="w",padx=4)

        tv=tk.StringVar(value=turler[0])
        cmb(form,tv,turler,width=16).grid(row=1,column=0,padx=4,ipady=5)

        tuv=tk.StringVar()
        ent(form,textvariable=tuv,width=12).grid(row=1,column=1,padx=4,ipady=6)

        # Fatura tarihi — varsayılan bugün, GG.AA.YYYY formatında
        ftv=tk.StringVar(value=datetime.date.today().strftime("%d.%m.%Y"))
        ft_ent=ent(form,textvariable=ftv,width=13)
        ft_ent.grid(row=1,column=2,padx=4,ipady=6)
        lbl(form,"(GG.AA.YYYY)",fg=T["text3"],bg=T["bg3"],
            font=("Segoe UI",7)).grid(row=2,column=2,sticky="w",padx=4)

        av=tk.StringVar()
        ent(form,textvariable=av,width=28).grid(row=1,column=3,padx=4,ipady=6)

        pw=pane(parent,orient="vertical")
        pw.pack(fill="both",expand=True,padx=12,pady=(0,12))
        tw=frm(pw,bg=T["bg3"],highlightthickness=1,highlightbackground=T["border"])
        pw.add(tw,weight=1)
        tf,tree=scrolled(tw,[
            ("Fatura Tarihi",120),("Kayıt Tarihi",120),
            ("Tür",120),("Açıklama",280),("Tutar",110)],height=14)
        tf.pack(fill="both",expand=True,padx=6,pady=6)

        def ekle(t=tip,tv2=tv,tuv2=tuv,fv2=ftv,av2=av,tr=tree):
            try: tutar=float(tuv2.get())
            except: messagebox.showerror("Hata","Geçerli tutar girin."); return
            # Fatura tarihi parse
            try:
                ft_obj=datetime.datetime.strptime(fv2.get().strip(),"%d.%m.%Y")
                ft_iso=ft_obj.isoformat(timespec="seconds")
            except:
                messagebox.showerror("Hata","Fatura tarihi GG.AA.YYYY formatında olmalı.\nÖrnek: 15.03.2025")
                return
            k={"id":yeni_id(),"tur":tv2.get(),"tutar":tutar,
               "aciklama":av2.get().strip(),
               "tarih":simdi(),          # kayıt zamanı
               "fatura_tarihi":ft_iso}   # fatura/işlem tarihi
            self.data[f"{t}ler"].append(k)
            save_data(self.data)
            tuv2.set(""); av2.set("")
            fv2.set(datetime.date.today().strftime("%d.%m.%Y"))
            self._ie_guncelle(tr,t)

        def sil(t=tip,tr=tree):
            sel=tr.selection()
            if not sel: return
            if not messagebox.askyesno("Sil","Seçili kayıt silinsin mi?"): return
            oid=int(sel[0])
            self.data[f"{t}ler"]=[x for x in self.data[f"{t}ler"] if x["id"]!=oid]
            save_data(self.data); self._ie_guncelle(tr,t)

        btn(form,"+ Ekle",ekle,color).grid(row=1,column=4,padx=(10,4))
        btn_r=frm(parent,bg=T["bg2"]); btn_r.pack(fill="x",padx=12,pady=(0,8))
        btn(btn_r,"Sil Seçili Kaydı Sil",sil,"red").pack(side="left")
        self._ie_guncelle(tree,tip)

    def _ie_guncelle(self,tree,tip):
        tree.delete(*tree.get_children())
        tag="gelir" if tip=="gelir" else "gider"
        for k in reversed(self.data[f"{tip}ler"]):
            # Fatura tarihi: yeni kayıtlarda "fatura_tarihi", eskilerinde "tarih" kullan
            ft = k.get("fatura_tarihi") or k.get("tarih","")
            tree.insert("","end",iid=str(k["id"]),tags=(tag,),
                values=(tarih_str(ft),
                        tarih_str(k.get("tarih","")),
                        k["tur"],
                        k.get("aciklama") or "—",
                        fmt(k["tutar"])))

    # ── ORTAK GİDERLER ────────────────────────────────────────────────────────
    def _tab_ortak_giderler(self):
        outer = frm(self._content)
        outer.pack(fill="both", expand=True, padx=18, pady=14)

        # Başlık + buton satırı
        hdr = frm(outer); hdr.pack(fill="x", pady=(0,10))
        lbl(hdr, "Arac  Ortak Giderler / Tadilat Takibi",
            font=("Segoe UI",13,"bold")).pack(side="left")
        bf = frm(hdr); bf.pack(side="right")
        btn(bf, "+  Yeni Ortak Gider",
            lambda: OrtakGiderPencere(self, self.data,
                                      callback=self.refresh_all),
            "gold").pack(side="left", ipadx=12, ipady=5)

        # ── İki sekmeli yapı
        nb = ttk.Notebook(outer, style="Dark.TNotebook")
        nb.pack(fill="both", expand=True)

        # ── SEKME 1: Gider Listesi ──────────────────────────────────────────
        t1 = frm(nb, bg=T["bg2"]); nb.add(t1, text="  Liste  Ortak Gider Listesi  ")

        pw = pane(t1, orient="vertical")
        pw.pack(fill="both", expand=True, padx=4, pady=4)

        ust = frm(pw); pw.add(ust, weight=55)

        hdr2 = frm(ust, bg=T["bg2"]); hdr2.pack(fill="x")
        tk.Frame(hdr2, bg=T["gold"], height=2).pack(fill="x")
        h2f = frm(hdr2, bg=T["bg2"]); h2f.pack(fill="x", padx=10, pady=6)
        lbl(h2f, "Liste  Kayıtlı Ortak Giderler",
            font=("Segoe UI",11,"bold"), bg=T["bg2"]).pack(side="left")
        lbl(h2f, "  Çift tıkla = düzenle",
            fg=T["text3"], bg=T["bg2"], font=("Segoe UI",8)).pack(side="left", padx=6)

        tf, self._og_tree = scrolled(ust, [
            ("Tarih",      110), ("İş Adı",  200), ("Toplam",   100),
            ("Daire Başı", 100), ("Kapsam",   90), ("Ödeme",     90),
            ("Durum",      120),
        ], height=9)
        tf.pack(fill="both", expand=True, padx=8, pady=(4,4))
        self._og_tree.bind("<Double-1>", self._og_cift_tik)
        self._og_tree.bind("<<TreeviewSelect>>", self._og_sec)

        btn_row = frm(ust); btn_row.pack(fill="x", padx=8, pady=(0,4))
        btn(btn_row, "Duzenle Düzenle", self._og_duzenle, "blue").pack(side="left", padx=(0,8), ipadx=8, ipady=4)
        btn(btn_row, "Sil Sil",    self._og_sil,     "red" ).pack(side="left", ipadx=8, ipady=4)

        # Alt bölme: taksit detayı (sol) + daire özet (sağ) — seçili gidere göre
        alt = frm(pw); pw.add(alt, weight=45)

        ahdr = frm(alt, bg=T["bg2"]); ahdr.pack(fill="x")
        tk.Frame(ahdr, bg=T["gold"], height=2).pack(fill="x")
        lbl(frm(ahdr, bg=T["bg2"]).pack(fill="x",padx=10,pady=6) or ahdr,
            "Takvim  Detay (Seçili Kayıt) — Taksitler ve Daire Ödemeleri",
            font=("Segoe UI",10,"bold"), bg=T["bg2"]).pack(anchor="w", padx=10, pady=6)

        alt_pw = pane(alt, orient="horizontal")
        alt_pw.pack(fill="both", expand=True, padx=8, pady=(0,8))

        taksit_frm = frm(alt_pw); alt_pw.add(taksit_frm, weight=40)
        lbl(taksit_frm, "Taksitler", fg=T["text2"], bg=T["bg3"],
            font=("Segoe UI",9,"bold")).pack(anchor="w", padx=4, pady=(4,2))
        tf2, self._og_det_tree = scrolled(taksit_frm, [
            ("Ay", 130), ("Tutar / Daire", 140),
        ], height=6)
        tf2.pack(fill="both", expand=True)

        daire_frm = frm(alt_pw); alt_pw.add(daire_frm, weight=60)
        lbl(daire_frm, "Daire Bazlı Ödeme Durumu", fg=T["text2"], bg=T["bg3"],
            font=("Segoe UI",9,"bold")).pack(anchor="w", padx=4, pady=(4,2))
        tf3, self._og_daire_tree = scrolled(daire_frm, [
            ("Daire",  70), ("Sakin",  110), ("Toplam", 85),
            ("Ödenen", 85), ("Kalan",  85),  ("Durum",  85),
        ], height=6)
        tf3.pack(fill="both", expand=True)
        btn_og1_row = frm(daire_frm, bg=T["bg3"]); btn_og1_row.pack(anchor="w", padx=4, pady=(4,2))
        btn(btn_og1_row, "Kart Ödeme Yap",    self._og_daire_odeme_yap,  "gold").pack(side="left", ipadx=8, ipady=3)
        btn(btn_og1_row, " Geçmiş/Düzenle", self._og_daire1_gecmis, "blue").pack(side="left", padx=6, ipadx=6, ipady=3)

        # ── SEKME 2: Daire Bazlı Ödeme Takibi ──────────────────────────────
        t2 = frm(nb, bg=T["bg2"]); nb.add(t2, text="  Ev  Daire Bazlı Ödemeler  ")

        t2hdr = frm(t2, bg=T["bg2"]); t2hdr.pack(fill="x", padx=10, pady=8)
        lbl(t2hdr, "Ortak Gider:", fg=T["text2"], bg=T["bg2"],
            font=("Segoe UI",9,"bold")).pack(side="left")
        self._og_sec_var = tk.StringVar(value="")
        self._og_sec_cb  = ttk.Combobox(t2hdr, textvariable=self._og_sec_var,
            style="Dark.TCombobox", font=("Segoe UI",10), width=44, state="readonly")
        self._og_sec_cb.pack(side="left", padx=8, ipady=5)
        btn(t2hdr, "Ara Göster", self._og_daire_yenile, "blue").pack(side="left", ipadx=8, ipady=4)

        # Özet şerit
        self._og_ozet_frm = frm(t2, bg=T["bg3"],
            highlightthickness=1, highlightbackground=T["border"])
        self._og_ozet_frm.pack(fill="x", padx=10, pady=(0,6))

        # Daire listesi
        lhdr = frm(t2, bg=T["bg2"]); lhdr.pack(fill="x", padx=0)
        tk.Frame(lhdr, bg=T["gold"], height=2).pack(fill="x")
        lhf  = frm(lhdr, bg=T["bg2"]); lhf.pack(fill="x", padx=10, pady=6)
        lbl(lhf, "Ev  Tüm Daireler — Ödeme Takibi",
            font=("Segoe UI",11,"bold"), bg=T["bg2"]).pack(side="left")

        tf4, self._og_daire2_tree = scrolled(t2, [
            ("Daire",  75), ("Sakin",  150), ("Toplam",  100),
            ("Ödenen", 100), ("Kalan", 100), ("Durum",   110),
        ], height=16)
        tf4.pack(fill="both", expand=True, padx=8, pady=(4,4))
        self._og_daire2_tree.bind("<Double-1>", lambda e: self._og_daire2_odeme_yap())

        btn_row2 = frm(t2); btn_row2.pack(fill="x", padx=8, pady=(0,8))
        btn(btn_row2, "Kart Ödeme Yap",    self._og_daire2_odeme_yap, "gold").pack(side="left", ipadx=8, ipady=4)
        btn(btn_row2, " Ödeme Geçmişi",self._og_daire2_gecmis,   "blue").pack(side="left", padx=8, ipadx=8, ipady=4)

        self._og_guncelle()
        self._og_daire_listesi_doldur()

    def _og_guncelle(self):
        if not hasattr(self, "_og_tree"):
            return
        self._og_tree.delete(*self._og_tree.get_children())
        for g in reversed(self.data.get("ortak_giderler", [])):
            al_kayitlar = [a for a in self.data.get("alacaklilar", [])
                           if a.get("ortak_gider_id") == g["id"]]
            if al_kayitlar:
                tam_odenen   = sum(1 for a in al_kayitlar
                    if a.get("odendi") or
                    round(a["tutar"] - sum(p["tutar"] for p in a.get("odemeler",[])), 2) <= 0)
                kismi_odenen = sum(1 for a in al_kayitlar
                    if not (a.get("odendi") or
                    round(a["tutar"] - sum(p["tutar"] for p in a.get("odemeler",[])), 2) <= 0)
                    and sum(p["tutar"] for p in a.get("odemeler",[])) > 0)
                if tam_odenen == len(al_kayitlar):
                    durum_txt, durum_tag = "OK Tümü Ödendi", "odendi"
                elif tam_odenen > 0 or kismi_odenen > 0:
                    durum_txt, durum_tag = f"! Kısmi ({tam_odenen}/{len(al_kayitlar)})", "bekliyor"
                else:
                    durum_txt, durum_tag = "... Bekliyor", "borclu"
            else:
                durum_txt, durum_tag = "Liste Kaydedildi", "normal"

            daire_sayisi = len(g.get("daire_nos", []))
            kapsam_txt = "Tüm Daireler" if g.get("kapsam") == "tum" else f"{daire_sayisi} Daire"
            odeme_txt  = "Tek Seferlik" if g.get("odeme_sekli") == "tek" else "Taksitli"
            self._og_tree.insert("", "end", iid=str(g["id"]), tags=(durum_tag,),
                values=(tarih_str(g.get("tarih", "")), g.get("ad", ""),
                        fmt(g.get("toplam", 0)), fmt(g.get("daire_basi", 0)),
                        kapsam_txt, odeme_txt, durum_txt))

    def _og_secili_gider(self):
        sel = self._og_tree.selection()
        if not sel:
            return None
        gid = int(sel[0])
        return next((g for g in self.data.get("ortak_giderler", [])
                     if g["id"] == gid), None)

    def _og_sec(self, event=None):
        """Seçili giderin taksit + daire detayını doldur."""
        if not hasattr(self, "_og_det_tree"):
            return
        self._og_det_tree.delete(*self._og_det_tree.get_children())
        if hasattr(self, "_og_daire_tree"):
            self._og_daire_tree.delete(*self._og_daire_tree.get_children())
        g = self._og_secili_gider()
        if not g:
            return
        # Taksit listesi — toplam tutarı da göster
        taksitler = g.get("taksitler", [])
        toplam_taksit = sum(t["tutar"] for t in taksitler)
        for t in taksitler:
            self._og_det_tree.insert("", "end", tags=("normal",),
                values=(ay_label(t["ay"]), fmt(t["tutar"])))
        # Toplam satırı
        if len(taksitler) > 1:
            self._og_det_tree.insert("", "end", tags=("gelir",),
                values=("── Toplam", fmt(toplam_taksit)))
        # Daire ödeme durumu
        if hasattr(self, "_og_daire_tree"):
            self._og_doldur_daire_ozet(g, self._og_daire_tree)

    def _og_cift_tik(self, event):
        item = self._og_tree.identify_row(event.y)
        if not item:
            return
        self._og_tree.selection_set(item)
        self._og_duzenle()

    def _og_duzenle(self):
        g = self._og_secili_gider()
        if not g:
            messagebox.showinfo("Seçim", "Lütfen bir kayıt seçin.", parent=self)
            return
        OrtakGiderPencere(self, self.data, callback=self.refresh_all, duzenle=g)

    def _og_uygula(self):
        g = self._og_secili_gider()
        if not g:
            messagebox.showinfo("Seçim", "Lütfen bir kayıt seçin.", parent=self)
            return
        if g.get("uygulandi"):
            messagebox.showinfo("Bilgi",
                "Bu kayıt zaten aidatlara uygulanmış.\n"
                "Önce geri alıp tekrar uygulayabilirsiniz.", parent=self)
            return
        _ortak_gider_aidata_uygula(self.data, g)
        save_data(self.data)
        self.refresh_all()
        messagebox.showinfo("Tamam",
            f"'{g['ad']}' aidatlara uygulandı.", parent=self)

    def _og_geri_al(self):
        g = self._og_secili_gider()
        if not g:
            messagebox.showinfo("Seçim", "Lütfen bir kayıt seçin.", parent=self)
            return
        if not g.get("uygulandi"):
            messagebox.showinfo("Bilgi", "Bu kayıt henüz uygulanmamış.", parent=self)
            return
        if not messagebox.askyesno("Geri Al",
            f"'{g['ad']}' aidatlardan geri alınsın mı?", parent=self):
            return
        _ortak_gider_aidat_geri_al(self.data, g)
        save_data(self.data)
        self.refresh_all()

    def _og_sil(self):
        g = self._og_secili_gider()
        if not g:
            messagebox.showinfo("Seçim", "Lütfen bir kayıt seçin.", parent=self)
            return
        if not messagebox.askyesno("Sil",
            f"'{g['ad']}' silinsin mi?\n"
            + ("Aidatlara uygulanmış tutarlar da geri alınacak." if g.get("uygulandi") else ""),
            parent=self):
            return
        if g.get("uygulandi"):
            _ortak_gider_aidat_geri_al(self.data, g)
        self.data["ortak_giderler"] = [
            x for x in self.data["ortak_giderler"] if x["id"] != g["id"]]
        # Gider hesabından da kaldır
        self.data["giderler"] = [
            x for x in self.data["giderler"] if x.get("ortak_gider_id") != g["id"]]
        save_data(self.data)
        self.refresh_all()

    # ── Ortak gider daire yardımcıları ───────────────────────────────────────

    def _og_doldur_daire_ozet(self, g, tree):
        """Verilen gidere bağlı alacaklı kayıtlarını daire bazlı tree'ye doldur."""
        tree.delete(*tree.get_children())
        al_by_daire = {a["daire_no"]: a
                       for a in self.data.get("alacaklilar", [])
                       if a.get("ortak_gider_id") == g["id"] and "daire_no" in a}
        for dno in sorted(g.get("daire_nos", [])):
            d    = next((x for x in self.data["daireler"] if x["no"] == dno), {})
            isim = d.get("isim", "") or "—"
            a    = al_by_daire.get(dno)
            if a:
                odenen = sum(p["tutar"] for p in a.get("odemeler", []))
                kalan  = round(a["tutar"] - odenen, 2)
                tam    = a.get("odendi", False) or kalan <= 0
                if tam:
                    durum, tag = "v Ödendi", "odendi"
                elif odenen > 0:
                    durum, tag = "! Kısmi", "bekliyor"
                else:
                    durum, tag = "... Bekliyor", "borclu"
                tree.insert("", "end", iid=f"d{dno}_{g['id']}", tags=(tag,),
                    values=(f"Daire {dno}", isim, fmt(a["tutar"]),
                            fmt(odenen) if odenen > 0 else "—",
                            fmt(kalan)  if kalan  > 0 else "v",
                            durum))
            else:
                tree.insert("", "end", iid=f"d{dno}_{g['id']}", tags=("normal",),
                    values=(f"Daire {dno}", isim, "—", "—", "—", "—"))

    def _og_daire_odeme_yap(self):
        """Sekme-1 daire listesinden seçili daire için ödeme yap."""
        if not hasattr(self, "_og_daire_tree"):
            return
        sel = self._og_daire_tree.selection()
        if not sel:
            messagebox.showinfo("Seçim", "Lütfen bir daire seçin."); return
        iid = sel[0]   # "d{dno}_{gid}"
        parts = iid.lstrip("d").split("_")
        if len(parts) < 2:
            return
        dno = int(parts[0])
        gid = int(parts[1])
        a = next((x for x in self.data.get("alacaklilar", [])
                  if x.get("ortak_gider_id") == gid and x.get("daire_no") == dno), None)
        if not a:
            messagebox.showinfo("Bilgi", "Bu daire için alacaklı kaydı bulunamadı."); return
        kalan = round(a["tutar"] - sum(p["tutar"] for p in a.get("odemeler", [])), 2)
        if kalan <= 0:
            messagebox.showinfo("Tamam", "Bu dairenin borcu zaten kapatılmış."); return
        AlacakliOdemePencere(self, self.data, a,
            callback=lambda: (self.refresh_all(),))

    def _og_daire1_gecmis(self):
        """Sekme-1 daire listesinden seçili daire için geçmiş/düzenle penceresi aç."""
        if not hasattr(self, "_og_daire_tree"):
            return
        sel = self._og_daire_tree.selection()
        if not sel:
            messagebox.showinfo("Seçim", "Lütfen bir daire seçin."); return
        iid = sel[0]
        parts = iid.lstrip("d").split("_")
        if len(parts) < 2:
            return
        dno = int(parts[0])
        gid = int(parts[1])
        a = next((x for x in self.data.get("alacaklilar", [])
                  if x.get("ortak_gider_id") == gid and x.get("daire_no") == dno), None)
        if not a:
            messagebox.showinfo("Bilgi", "Bu daire için alacaklı kaydı bulunamadı."); return
        AlacakliGecmisPencere(self, a, data=self.data,
                              callback=lambda: (self._og_guncelle(), self._og_sec()))

    def _og_daire_listesi_doldur(self):
        """Sekme-2 combobox'ını uygulanan ortak giderlerle doldur."""
        if not hasattr(self, "_og_sec_cb"):
            return
        giderler = [g for g in self.data.get("ortak_giderler", []) if g.get("uygulandi")]
        secenekler = [f"{g['ad']}  ({tarih_str(g.get('tarih',''))})" for g in giderler]
        self._og_sec_cb["values"] = secenekler
        self._og_sec_giderler = giderler   # indeks eşleştirme için
        if secenekler:
            self._og_sec_cb.current(0)
            self._og_daire_yenile()

    def _og_daire_yenile(self):
        """Sekme-2 daire listesini ve özet şeridini yenile."""
        if not hasattr(self, "_og_daire2_tree"):
            return
        self._og_daire2_tree.delete(*self._og_daire2_tree.get_children())
        # Özet şerit temizle
        for w in self._og_ozet_frm.winfo_children():
            w.destroy()

        giderler = getattr(self, "_og_sec_giderler", [])
        idx = self._og_sec_cb.current()
        if idx < 0 or idx >= len(giderler):
            return
        g = giderler[idx]

        # Özet
        al_all = [a for a in self.data.get("alacaklilar", [])
                  if a.get("ortak_gider_id") == g["id"]]
        toplam_borc  = sum(a["tutar"] for a in al_all)
        toplam_odenan= sum(sum(p["tutar"] for p in a.get("odemeler",[])) for a in al_all)
        toplam_kalan = round(toplam_borc - toplam_odenan, 2)
        tam_sayi     = sum(1 for a in al_all
            if a.get("odendi") or
            round(a["tutar"] - sum(p["tutar"] for p in a.get("odemeler",[])), 2) <= 0)

        for col, (ltext, val, fg) in enumerate([
            ("İş Adı",        g.get("ad",""),              T["text"]),
            ("Toplam Borç",   fmt(toplam_borc),             T["text"]),
            ("Toplam Ödenen", fmt(toplam_odenan),           T["green"]),
            ("Kalan",         fmt(toplam_kalan),            T["red"] if toplam_kalan > 0 else T["green"]),
            ("Tamamlanan",    f"{tam_sayi}/{len(al_all)} Daire", T["gold2"]),
        ]):
            cf = frm(self._og_ozet_frm, bg=T["bg3"]); cf.pack(side="left", expand=True, padx=8, pady=6)
            lbl(cf, ltext, fg=T["text3"], bg=T["bg3"], font=("Segoe UI",8)).pack(anchor="w")
            lbl(cf, val,   fg=fg,         bg=T["bg3"], font=("Segoe UI",10,"bold")).pack(anchor="w")

        # Tüm daireleri listele (sadece bu giderin kapsamındakileri değil, tümünü göster)
        al_by_daire = {a["daire_no"]: a for a in al_all if "daire_no" in a}
        daire_nos_in_gider = set(g.get("daire_nos", []))

        for d in self.data["daireler"]:
            dno  = d["no"]
            isim = d.get("isim", "") or "—"
            a    = al_by_daire.get(dno)
            if a:
                odenen = sum(p["tutar"] for p in a.get("odemeler", []))
                kalan  = round(a["tutar"] - odenen, 2)
                tam    = a.get("odendi", False) or kalan <= 0
                if tam:
                    durum, tag = "v Ödendi", "odendi"
                elif odenen > 0:
                    durum, tag = "! Kısmi", "bekliyor"
                else:
                    durum, tag = "... Bekliyor", "borclu"
                self._og_daire2_tree.insert("", "end",
                    iid=f"d2_{dno}_{g['id']}", tags=(tag,),
                    values=(f"Daire {dno}", isim, fmt(a["tutar"]),
                            fmt(odenen) if odenen > 0 else "—",
                            fmt(kalan)  if kalan  > 0 else "v", durum))
            elif dno in daire_nos_in_gider:
                # Bu gidere dahil ama alacaklı kaydı yok (uygulama yapılmamış)
                self._og_daire2_tree.insert("", "end",
                    iid=f"d2_{dno}_{g['id']}", tags=("normal",),
                    values=(f"Daire {dno}", isim, "—", "—", "—", "Kapsam dışı"))
            else:
                # Bu giderin kapsamı dışındaki daire
                self._og_daire2_tree.insert("", "end",
                    iid=f"d2_{dno}_{g['id']}", tags=("normal",),
                    values=(f"Daire {dno}", isim, "—", "—", "—", "—"))

    def _og_daire2_secili_alacakli(self):
        """Sekme-2 tree'den seçili daire için alacaklı kaydını döndür."""
        if not hasattr(self, "_og_daire2_tree"):
            return None
        sel = self._og_daire2_tree.selection()
        if not sel:
            return None
        iid = sel[0]   # "d2_{dno}_{gid}"
        parts = iid.lstrip("d").lstrip("2").lstrip("_").split("_")
        # iid formatı: d2_{dno}_{gid}
        raw = iid[3:]   # "d2_" sonrası
        spl = raw.split("_")
        if len(spl) < 2:
            return None
        dno = int(spl[0])
        gid = int(spl[1])
        return next((x for x in self.data.get("alacaklilar", [])
                     if x.get("ortak_gider_id") == gid and x.get("daire_no") == dno), None)

    def _og_daire2_odeme_yap(self):
        a = self._og_daire2_secili_alacakli()
        if not a:
            messagebox.showinfo("Seçim", "Lütfen ödeme yapılacak daireyi seçin."); return
        kalan = round(a["tutar"] - sum(p["tutar"] for p in a.get("odemeler", [])), 2)
        if kalan <= 0:
            messagebox.showinfo("Tamam", "Bu dairenin borcu zaten kapatılmış."); return
        AlacakliOdemePencere(self, self.data, a,
            callback=lambda: self.refresh_all())

    def _og_daire2_gecmis(self):
        a = self._og_daire2_secili_alacakli()
        if not a:
            messagebox.showinfo("Seçim", "Lütfen bir daire seçin."); return
        AlacakliGecmisPencere(self, a, data=self.data,
                              callback=lambda: self.refresh_all())

    # ── ALACAKLILAR ───────────────────────────────────────────────────────────
    def _tab_alacaklilar(self):
        # Eski widget referanslarını temizle (refresh_all ile yeniden kurulumda çakışma önlenir)
        for attr in ("_al_tree", "_al_filtre", "_al_ozet_lbl",
                     "_al_tur", "_al_kisi", "_al_tutar", "_al_aciklama", "_al_tarih"):
            if hasattr(self, attr):
                delattr(self, attr)

        outer = frm(self._content)
        outer.pack(fill="both", expand=True, padx=18, pady=14)

        hdr = frm(outer); hdr.pack(fill="x", pady=(0,10))
        lbl(hdr, "Liste  Alacaklılar", font=("Segoe UI",13,"bold")).pack(side="left")
        al = self.data.get("alacaklilar", [])
        top_kalan = sum(max(0, round(a["tutar"]-sum(p["tutar"] for p in a.get("odemeler",[])),2))
                        for a in al if not (a.get("odendi",False) or
                        round(a["tutar"]-sum(p["tutar"] for p in a.get("odemeler",[])),2)<=0)
                        and a.get("tur") != "Ortak Gider")
        top_tamam = sum(a["tutar"] for a in al if (a.get("odendi",False) or
                        round(a["tutar"]-sum(p["tutar"] for p in a.get("odemeler",[])),2)<=0)
                        and a.get("tur") != "Ortak Gider")
        self._al_ozet_lbl = lbl(hdr,
            f"Toplam Kalan: {fmt(top_kalan)}  •  Tamamlanan: {fmt(top_tamam)}",
            fg=T["red"], font=("Segoe UI",10,"bold"))
        self._al_ozet_lbl.pack(side="right")

        pw = pane(outer, orient="vertical")
        pw.pack(fill="both", expand=True)

        # ── ÜST: Giriş formu ─────────────────────────────────────────────
        form_pnl = card_panel(pw, "+  Yeni Alacaklı Kaydı", "Gelen fatura veya nakit harcama")
        pw.add(form_pnl, weight=24)
        form_inner = frm(form_pnl, bg=T["bg2"])
        form_inner.pack(fill="both", expand=True, padx=12, pady=10)
        form = frm(form_inner, bg=T["bg2"]); form.pack(fill="x")
        for col, txt in [(0,"Tür"),(1,"Alacaklı / Kişi"),(2,"Toplam Tutar (₺)"),
                         (3,"Açıklama"),(4,"Fatura Tarihi")]:
            lbl(form, txt, fg=T["text2"], bg=T["bg2"],
                font=("Segoe UI",9,"bold")).grid(row=0,column=col,sticky="w",padx=6,pady=(0,3))
        self._al_tur = tk.StringVar(value="Fatura")
        cmb(form, self._al_tur,
            ["Fatura","Nakit Harcama","Bakım/Onarım","Temizlik","Güvenlik","Sigorta","Diğer"],
            width=14).grid(row=1,column=0,padx=6,ipady=5,sticky="ew")
        self._al_kisi = tk.StringVar()
        ent(form, textvariable=self._al_kisi, width=18).grid(row=1,column=1,padx=6,ipady=6,sticky="ew")
        self._al_tutar = tk.StringVar()
        ent(form, textvariable=self._al_tutar, width=13).grid(row=1,column=2,padx=6,ipady=6,sticky="ew")
        self._al_aciklama = tk.StringVar()
        ent(form, textvariable=self._al_aciklama, width=28).grid(row=1,column=3,padx=6,ipady=6,sticky="ew")
        self._al_tarih = tk.StringVar(value=datetime.date.today().strftime("%d.%m.%Y"))
        ent(form, textvariable=self._al_tarih, width=13).grid(row=1,column=4,padx=6,ipady=6,sticky="ew")
        btn(form, "+ Ekle", self._al_ekle, "gold").grid(row=1,column=5,padx=(10,4))
        lbl(form_inner,"Fatura Tarihi: GG.AA.YYYY  •  Çift tıkla → Ödeme yap  •  Kısmi ödemeler takip edilir.",
            fg=T["text3"],bg=T["bg2"],font=("Segoe UI",8)).pack(anchor="w",pady=(6,0))

        # ── ALT: Liste ───────────────────────────────────────────────────
        tbl_pnl = card_panel(pw, "Liste  Alacaklı Listesi", "Çift tıkla → Ödeme yap")
        pw.add(tbl_pnl, weight=76)
        tbl_inner = frm(tbl_pnl, bg=T["bg2"])
        tbl_inner.pack(fill="both", expand=True, padx=8, pady=8)

        fil = frm(tbl_inner, bg=T["bg2"]); fil.pack(fill="x", pady=(0,6))
        lbl(fil,"Filtre:",fg=T["text2"],bg=T["bg2"],font=("Segoe UI",9,"bold")).pack(side="left")
        self._al_filtre = tk.StringVar(value="Tümü")
        for val in ["Tümü","Bekleyen","Kısmi Ödendi","Tamamlandı"]:
            tk.Radiobutton(fil,text=val,variable=self._al_filtre,value=val,
                bg=T["bg2"],fg=T["text"],selectcolor=T["bg3"],activebackground=T["bg2"],
                font=("Segoe UI",9),cursor="hand2",
                command=lambda: self._al_guncelle()).pack(side="left",padx=6)

        tbl_f, self._al_tree = scrolled(tbl_inner, [
            ("Fatura Tarihi",110),("Grup",110),("Tür / Açıklama",190),
            ("Daire",70),("Toplam",100),("Ödenen",100),
            ("Kalan",100),("Durum",100),
        ], height=10)
        tbl_f.pack(fill="both", expand=True)
        self._al_tree.bind("<Double-1>", self._al_cift_tik)

        alt_f = frm(tbl_inner,bg=T["bg2"]); alt_f.pack(fill="x",pady=(8,0))
        btn(alt_f,"Kart Ödeme Yap",       self._al_odeme_yap,    "gold"  ).pack(side="left",padx=(0,6))
        btn(alt_f," Ödeme Geçmişi",   self._al_gecmis_goster,"blue"  ).pack(side="left",padx=(0,6))
        btn(alt_f,"OK Tamamen Ödendi",   self._al_tam_odendi,   "green" ).pack(side="left",padx=(0,6))
        btn(alt_f,"↩️ Ödemeleri Sıfırla",self._al_sifirla,     "orange").pack(side="left",padx=(0,6))
        btn(alt_f,"Sil Sil",            self._al_sil,          "red"   ).pack(side="left")
        self._al_guncelle()

    @staticmethod
    def _al_kalan(a):
        odenen = sum(p["tutar"] for p in a.get("odemeler",[]))
        return round(a["tutar"] - odenen, 2)

    def _al_secili(self):
        if not hasattr(self,"_al_tree"): return None
        sel = self._al_tree.selection()
        if not sel: return None
        return next((x for x in self.data.get("alacaklilar",[]) if x["id"]==int(sel[0])),None)

    def _al_guncelle(self):
        if not hasattr(self,"_al_tree"): return
        self._al_tree.delete(*self._al_tree.get_children())
        filtre = self._al_filtre.get() if hasattr(self,"_al_filtre") else "Tümü"
        for a in reversed(self.data.get("alacaklilar",[])):
            odenen = sum(p["tutar"] for p in a.get("odemeler",[]))
            kalan  = round(a["tutar"]-odenen,2)
            tam    = a.get("odendi",False) or kalan<=0
            if tam:          durum,tag = "v Tamamlandı","odendi"
            elif odenen>0:   durum,tag = "! Kısmi","bekliyor"
            else:            durum,tag = "... Bekliyor","borclu"
            if filtre=="Bekleyen"     and (tam or odenen>0): continue
            if filtre=="Kısmi Ödendi" and (tam or odenen==0): continue
            if filtre=="Tamamlandı"   and not tam: continue
                        # Ortak Gider kayıtları Alacaklılar sekmesinde gösterilmez
            if a.get("tur") == "Ortak Gider":
                continue
            ft = a.get("fatura_tarihi") or a.get("kayit_tarihi","")
            grup = a["tur"]
            daire_txt = "—"
            aciklama_txt = (a.get("kisi") or "") + (" — " + a.get("aciklama","") if a.get("aciklama") else "") or "—"
            self._al_tree.insert("","end",iid=str(a["id"]),tags=(tag,),
                values=(tarih_str(ft), grup, aciklama_txt, daire_txt,
                        fmt(a["tutar"]),
                        fmt(odenen) if odenen>0 else "—",
                        fmt(kalan) if kalan>0 else "v Tamam",
                        durum))
        al = self.data.get("alacaklilar",[])
        tk2 = sum(max(0,self._al_kalan(a)) for a in al if not a.get("odendi",False) and self._al_kalan(a)>0 and a.get("tur") != "Ortak Gider")
        tt  = sum(a["tutar"] for a in al if (a.get("odendi",False) or self._al_kalan(a)<=0) and a.get("tur") != "Ortak Gider")
        if hasattr(self,"_al_ozet_lbl"):
            self._al_ozet_lbl.config(text=f"Toplam Kalan: {fmt(tk2)}  •  Tamamlanan: {fmt(tt)}")

    def _al_cift_tik(self, event):
        item = self._al_tree.identify_row(event.y)
        if not item: return
        self._al_tree.selection_set(item)
        self._al_odeme_yap()

    def _al_odeme_yap(self):
        a = self._al_secili()
        if not a: messagebox.showinfo("Seçim","Lütfen bir alacaklı seçin."); return
        kalan = self._al_kalan(a)
        if kalan<=0: messagebox.showinfo("Tamam","Bu alacaklının borcu kapatılmış."); return
        AlacakliOdemePencere(self, self.data, a,
            callback=lambda:(self._al_guncelle(),self.refresh_all()))

    def _al_gecmis_goster(self):
        a = self._al_secili()
        if not a: messagebox.showinfo("Seçim","Lütfen bir alacaklı seçin."); return
        # Ortak Gider ise data ve callback ilet (düzenleme destekli)
        if a.get("tur") == "Ortak Gider":
            AlacakliGecmisPencere(self, a, data=self.data,
                callback=lambda: (self._al_guncelle(), self.refresh_all()))
        else:
            AlacakliGecmisPencere(self, a)

    def _al_tam_odendi(self):
        a = self._al_secili()
        if not a: messagebox.showinfo("Seçim","Lütfen bir alacaklı seçin."); return
        kalan = self._al_kalan(a)
        if kalan>0:
            if not messagebox.askyesno("Onayla",
                f"Kalan {fmt(kalan)} ödenmiş kabul edilsin mi?\n"
                "Kalan tutara eşit ödeme kaydı eklenecek."): return
            a.setdefault("odemeler",[]).append({
                "id":yeni_id(),"tutar":kalan,"tarih":simdi(),"not_":"Tamamen ödendi"})
        a["odendi"]=True; a["odeme_tarihi"]=simdi()
        save_data(self.data); self._al_guncelle(); self.refresh_all()

    def _al_sifirla(self):
        a = self._al_secili()
        if not a: messagebox.showinfo("Seçim","Lütfen bir alacaklı seçin."); return
        if not messagebox.askyesno("Sıfırla",
            f"'{a.get('kisi') or a['tur']}' için tüm ödeme geçmişi silinsin mi?"): return
        a["odemeler"]=[]; a["odendi"]=False; a["odeme_tarihi"]=""
        save_data(self.data); self._al_guncelle(); self.refresh_all()

    def _al_sil(self):
        a = self._al_secili()
        if not a: messagebox.showinfo("Seçim","Lütfen bir alacaklı seçin."); return
        if not messagebox.askyesno("Sil",
            f"'{a.get('kisi') or a['tur']}' kaydı ve tüm ödeme geçmişi silinsin mi?\n"
            f"Toplam: {fmt(a['tutar'])}"): return
        self.data["alacaklilar"]=[x for x in self.data["alacaklilar"] if x["id"]!=a["id"]]
        save_data(self.data); self._al_guncelle(); self.refresh_all()

    def _al_ekle(self):
        tur=self._al_tur.get(); kisi=self._al_kisi.get().strip()
        aciklama=self._al_aciklama.get().strip()
        try:
            tutar=float(self._al_tutar.get())
            if tutar<=0: raise ValueError
        except: messagebox.showerror("Hata","Geçerli ve pozitif tutar girin."); return
        try:
            ft=datetime.datetime.strptime(self._al_tarih.get().strip(),"%d.%m.%Y")
            ft_iso=ft.isoformat(timespec="seconds")
        except: messagebox.showerror("Hata","Fatura tarihi GG.AA.YYYY formatında olmalı."); return
        kayit={"id":yeni_id(),"tur":tur,"kisi":kisi,"aciklama":aciklama,
               "tutar":tutar,"fatura_tarihi":ft_iso,"kayit_tarihi":simdi(),
               "odendi":False,"odeme_tarihi":"","odemeler":[]}
        self.data.setdefault("alacaklilar",[]).append(kayit)
        save_data(self.data)
        self._al_kisi.set(""); self._al_tutar.set("")
        self._al_aciklama.set("")
        self._al_tarih.set(datetime.date.today().strftime("%d.%m.%Y"))
        self._al_guncelle(); self.refresh_all()

    def _tab_raporlar(self):
        outer=frm(self._content)
        outer.pack(fill="both",expand=True,padx=18,pady=14)
        hdr=frm(outer); hdr.pack(fill="x",pady=(0,10))
        lbl(hdr,"Yukari  Raporlar",font=("Segoe UI",13,"bold")).pack(side="left")
        btn(hdr,"Yazdir Özet Raporu Yazdır",self._rapor_yazdir,"gold").pack(side="right")

        pw=pane(outer,orient="vertical")
        pw.pack(fill="both",expand=True)

        # 7. Gider Dağılımı ÜSTTE
        alt=card_panel(pw,"Para  Gider Dağılımı","Kategoriye göre harcama")
        pw.add(alt,weight=30)
        ai=frm(alt,bg=T["bg2"]); ai.pack(fill="both",expand=True,padx=12,pady=8)
        gtur={}
        for g in self.data["giderler"]: gtur[g["tur"]]=gtur.get(g["tur"],0)+g["tutar"]
        if gtur:
            COLS=["#e74c3c","#f39c12","#3498db","#2ecc71","#9b59b6","#1abc9c","#e91e63","#f1c40f"]
            mx=max(gtur.values())
            for i,(tur,v) in enumerate(sorted(gtur.items(),key=lambda x:-x[1])[:8]):
                row=frm(ai,bg=T["bg2"]); row.pack(fill="x",pady=2)
                lbl(row,tur,fg=T["text2"],bg=T["bg2"],font=("Segoe UI",9),width=20,anchor="e").pack(side="left",padx=(0,8))
                c=tk.Canvas(row,bg=T["bg3"],height=20,highlightthickness=0)
                c.pack(side="left",fill="x",expand=True)
                c.update_idletasks()
                tw2=c.winfo_width() or 400
                fw=int(tw2*v/mx); col=COLS[i%len(COLS)]
                c.after(50,lambda cv=c,fw2=fw,cl=col: cv.create_rectangle(0,0,fw2,20,fill=cl,outline=""))
                lbl(row,fmt(v),fg=T["text2"],bg=T["bg2"],font=("Segoe UI",9,"bold"),width=12).pack(side="left",padx=8)
        else:
            lbl(ai,"Gider kaydı yok.",fg=T["text3"],bg=T["bg2"]).pack(pady=20)

        # Aylık Özet ORTADA
        ust=card_panel(pw,"Liste  Aylık Özet","Gelir, gider, tahsilat")
        pw.add(ust,weight=40)
        ui=frm(ust,bg=T["bg2"]); ui.pack(fill="both",expand=True,padx=8,pady=8)
        tf,to=scrolled(ui,[("Ay",160),("Gelir",140),("Gider",140),("Net",140),("Tahsilat %",110)],height=8)
        tf.pack(fill="both",expand=True)
        b=self.data["bina"]
        aylar=set()
        for g in self.data["gelirler"]: aylar.add(g["tarih"][:7])
        for g in self.data["giderler"]: aylar.add(g["tarih"][:7])
        for o in self.data["odemeler"]: aylar.add(o["ay"])
        for ay in sorted(aylar,reverse=True):
            gel=sum(g["tutar"] for g in self.data["gelirler"] if g["tarih"].startswith(ay))
            gid=sum(g["tutar"] for g in self.data["giderler"] if g["tarih"].startswith(ay))
            net=gel-gid
            tah=sum(o["tutar"] for o in self.data["odemeler"] if o["ay"]==ay)
            oran=int(tah/(b["daire"]*b["aidat"])*100) if b["aidat"] else 0
            tag="gelir" if net>=0 else "gider"
            to.insert("","end",tags=(tag,),
                values=(ay_label(ay),fmt(gel),fmt(gid),fmt(net),f"%{oran}"))

        # 2. Alacaklılar özet tablosu ALTTA
        alacak_pnl=card_panel(pw,"Liste  Alacaklılar Özeti","Fatura ve nakit harcamalar")
        pw.add(alacak_pnl,weight=30)
        alai=frm(alacak_pnl,bg=T["bg2"]); alai.pack(fill="both",expand=True,padx=8,pady=8)
        altf,altr=scrolled(alai,[
            ("Fatura Tarihi",110),("Grup",110),("Tür / Açıklama",190),
            ("Daire",70),("Toplam",100),("Ödenen",100),("Kalan",100),("Durum",100)
        ],height=6)
        altf.pack(fill="both",expand=True)
        al_list=self.data.get("alacaklilar",[])
        if al_list:
            for a in reversed(al_list):
                odenen=sum(p["tutar"] for p in a.get("odemeler",[]))
                kalan=round(a["tutar"]-odenen,2)
                tam=a.get("odendi",False) or kalan<=0
                if tam:       durum,tag="v Tamamlandı","odendi"
                elif odenen>0:durum,tag="! Kısmi","bekliyor"
                else:         durum,tag="... Bekliyor","borclu"
                ft=a.get("fatura_tarihi") or a.get("kayit_tarihi","")
                if a.get("tur")=="Ortak Gider":
                    grup="Arac Ortak Gider"
                    dno=a.get("daire_no"); daire_txt=f"Daire {dno}" if dno else "—"
                    acik_txt=a.get("kisi") or a.get("aciklama") or "—"
                else:
                    grup=a["tur"]; daire_txt="—"
                    acik_txt=(a.get("kisi") or "")+((" — "+a.get("aciklama","")) if a.get("aciklama") else "") or "—"
                altr.insert("","end",tags=(tag,),
                    values=(tarih_str(ft),grup,acik_txt,daire_txt,
                            fmt(a["tutar"]),
                            fmt(odenen) if odenen>0 else "—",
                            fmt(kalan) if kalan>0 else "v Tamam",
                            durum))
        else:
            altr.insert("","end",tags=("normal",),
                values=("Alacaklı kaydı yok","","","","","","",""))

    def _rapor_yazdir(self):
        b=self.data["bina"]; ba=buay()

        # ── Genel Bakış kartlarındaki değerleri hesapla ───────────────────
        tah_buay = sum(o["tutar"] for o in self.data["odemeler"]
                       if o["tarih"][:7] == ba)
        bek      = b["daire"] * b["aidat"]
        oran_buay= int(tah_buay / bek * 100) if bek else 0

        topB     = sum(d["borc"] for d in self.data["daireler"])
        topF     = sum(d["faiz"] for d in self.data["daireler"])
        topAlacak= sum(
            max(0.0, round(a["tutar"] - sum(p["tutar"] for p in a.get("odemeler",[])), 2))
            for a in self.data.get("alacaklilar", [])
            if not a.get("odendi", False)
        )
        topBorcToplam = topB + topF + topAlacak

        topG  = sum(g["tutar"] for g in self.data["gelirler"])
        topGd = sum(g["tutar"] for g in self.data["giderler"])
        net   = topG - topGd

        tum_ay = ay_listesi(b.get("baslangic", ba))
        odenmemis_toplam = sum(
            ay_aidat(self.data, ay, d["no"])
            + faiz_hesapla(ay_aidat(self.data, ay, d["no"]), b["faiz"], ay, b["son_gun"])
            for ay in tum_ay
            for d in self.data["daireler"]
            if not any(o["daireNo"] == d["no"] and o["ay"] == ay
                       for o in self.data["odemeler"])
        )

        net_renk = "#2ecc71" if net >= 0 else "#e74c3c"

        stat_cards_html = f"""
<div class="stat-grid">
  <div class="stat-card stat-red">
    <div class="stat-title">Uyari  Toplam Borç</div>
    <div class="stat-val">{fmt(topBorcToplam)}</div>
    <div class="stat-sub">Aidat: {fmt(topB+topF)} &nbsp;|&nbsp; Alacak: {fmt(topAlacak)}</div>
  </div>
  <div class="stat-card stat-gold">
    <div class="stat-title">Kart  Bu Ayki Tahsilat</div>
    <div class="stat-val">{fmt(tah_buay)}</div>
    <div class="stat-sub">%{oran_buay} &nbsp;|&nbsp; Beklenen: {fmt(bek)}</div>
  </div>
  <div class="stat-card stat-green">
    <div class="stat-title">Yukari  Toplam Gelir</div>
    <div class="stat-val">{fmt(topG)}</div>
    <div class="stat-sub">Tüm zamanlar</div>
  </div>
  <div class="stat-card" style="border-top:3px solid {net_renk}">
    <div class="stat-title">Canta  Net Bakiye</div>
    <div class="stat-val" style="color:{net_renk}">{fmt(net)}</div>
    <div class="stat-sub">Gelir — Gider</div>
  </div>
  <div class="stat-card stat-orange">
    <div class="stat-title">Saat  Ödenmemiş Borç</div>
    <div class="stat-val">{fmt(odenmemis_toplam)}</div>
    <div class="stat-sub">Tüm geçmiş ödenmemiş dönemler</div>
  </div>
</div>"""

        # ── Aylık özet tablosu ────────────────────────────────────────────
        aylar=set()
        for g in self.data["gelirler"]: aylar.add(g["tarih"][:7])
        for g in self.data["giderler"]: aylar.add(g["tarih"][:7])
        for o in self.data["odemeler"]: aylar.add(o["ay"])
        rows=""
        for ay in sorted(aylar, reverse=True):
            gel=sum(g["tutar"] for g in self.data["gelirler"] if g["tarih"].startswith(ay))
            gid=sum(g["tutar"] for g in self.data["giderler"] if g["tarih"].startswith(ay))
            net_ay=gel-gid
            tah=sum(o["tutar"] for o in self.data["odemeler"] if o["ay"]==ay)
            oran=int(tah/(b["daire"]*b["aidat"])*100) if b["aidat"] else 0
            nc="#2ecc71" if net_ay>=0 else "#e74c3c"
            rows+=(f"<tr><td>{ay_label(ay)}</td>"
                   f"<td style='color:#2ecc71'>{fmt(gel)}</td>"
                   f"<td style='color:#e74c3c'>{fmt(gid)}</td>"
                   f"<td style='color:{nc};font-weight:bold'>{fmt(net_ay)}</td>"
                   f"<td>%{oran}</td></tr>")

        html=f"""<!DOCTYPE html><html lang="tr"><head><meta charset="UTF-8">
<title>Özet Rapor — {b['adi']}</title>
<style>
body{{font-family:Arial,sans-serif;padding:30px;background:#f0f0f0;}}
.card{{background:#fff;padding:28px;border-radius:8px;max-width:820px;margin:0 auto;}}
h2{{font-size:18px;margin-bottom:4px;}}
.meta{{color:#888;font-size:12px;margin-bottom:20px;}}
.stat-grid{{display:grid;grid-template-columns:repeat(5,1fr);gap:12px;margin-bottom:24px;}}
.stat-card{{background:#f9f9f9;border-radius:6px;padding:14px 12px;border-top:3px solid #999;}}
.stat-red{{border-top-color:#e74c3c;}}
.stat-gold{{border-top-color:#c9a84c;}}
.stat-green{{border-top-color:#2ecc71;}}
.stat-orange{{border-top-color:#f39c12;}}
.stat-title{{font-size:11px;color:#666;margin-bottom:6px;}}
.stat-val{{font-size:18px;font-weight:bold;color:#111;}}
.stat-sub{{font-size:10px;color:#999;margin-top:4px;}}
h3{{font-size:13px;color:#444;margin:20px 0 6px;}}
table{{width:100%;border-collapse:collapse;font-size:13px;}}
th{{background:#111;color:#fff;padding:9px;text-align:center;}}
td{{padding:8px;text-align:center;border-bottom:1px solid #eee;}}
.np{{text-align:center;margin-bottom:20px;}}
button{{padding:9px 22px;background:#111;color:#fff;border:none;border-radius:4px;cursor:pointer;}}
@media print{{
  .np{{display:none;}}
  body{{background:#fff;padding:0;}}
  .stat-grid{{grid-template-columns:repeat(5,1fr);}}
  .stat-card{{border:1px solid #ddd;}}
}}
</style></head><body>
<div class="np"><button onclick="window.print()">Yazdir Yazdır / PDF</button></div>
<div class="card">
  <h2>Bina {b['adi']} — Özet Rapor</h2>
  <p class="meta">Oluşturma tarihi: {datetime.datetime.now().strftime('%d.%m.%Y %H:%M')}</p>
  {stat_cards_html}
  <h3>Liste Aylık Özet</h3>
  <table>
    <tr><th>Ay</th><th>Gelir</th><th>Gider</th><th>Net</th><th>Tahsilat %</th></tr>
    {rows}
  </table>
</div></body></html>"""
        tmp=tempfile.NamedTemporaryFile(suffix=".html",delete=False,mode="w",encoding="utf-8")
        tmp.write(html); tmp.close()
        webbrowser.open(f"file:///{tmp.name}")

    # ── DETAYLI RAPORLAR ──────────────────────────────────────────────────────
    def _tab_detay_rapor(self):
        outer = frm(self._content)
        outer.pack(fill="both", expand=True, padx=18, pady=14)

        # Başlık + filtre bar
        hdr = frm(outer); hdr.pack(fill="x", pady=(0,10))
        lbl(hdr, "Klasor  Detaylı Raporlar",
            font=("Segoe UI",13,"bold")).pack(side="left")

        # Ay filtresi
        fil = frm(hdr); fil.pack(side="right")
        lbl(fil, "Dönem:", fg=T["text2"],
            font=("Segoe UI",9,"bold")).pack(side="left", padx=(0,6))
        self._dr_bas = AySecici(fil, initial=self.data["bina"].get("baslangic", buay()))
        self._dr_bas.pack(side="left")
        lbl(fil, " — ", fg=T["text2"]).pack(side="left", padx=4)
        self._dr_son = AySecici(fil, initial=buay())
        self._dr_son.pack(side="left")
        btn(fil, "Ara Filtrele", lambda: self._dr_yenile(nb), "gold").pack(
            side="left", padx=(10,0))
        btn(fil, "Yazdir Tümünü Yazdır", lambda: self._dr_yazdir(), "blue").pack(
            side="left", padx=(6,0))

        # Notebook — 4 sekme
        nb = ttk.Notebook(outer, style="Dark.TNotebook")
        nb.pack(fill="both", expand=True)

        # ── Sekme 1: Aidat Ödeyenler ──────────────────────────────────────
        t1 = frm(nb, bg=T["bg2"]); nb.add(t1, text="  OK  Aidat Ödeyenler  ")

        t1_hdr = frm(t1, bg=T["bg2"]); t1_hdr.pack(fill="x", padx=12, pady=8)
        self._dr_lbl_odeyenler = lbl(t1_hdr, "", fg=T["green"],
            bg=T["bg2"], font=("Segoe UI",10,"bold"))
        self._dr_lbl_odeyenler.pack(side="left")
        btn(t1_hdr, "Yazdir Bu Listeyi Yazdır",
            lambda: self._dr_yazdir("odeyenler"), "green").pack(side="right")

        t1pw = pane(t1, orient="vertical"); t1pw.pack(fill="both", expand=True, padx=8, pady=(0,8))

        # Üst: Dönemsel özet
        t1ust = card_panel(t1pw, "Takvim  Dönemsel Tahsilat Özeti", "Ay bazında ödeme oranı")
        t1pw.add(t1ust, weight=40)
        t1ui = frm(t1ust, bg=T["bg2"]); t1ui.pack(fill="both", expand=True, padx=8, pady=6)
        t1f, self._dr_tree_ay_ozet = scrolled(t1ui, [
            ("Dönem",130),("Toplam Daire",110),("Ödeyen",100),
            ("Ödemeyen",110),("Tahsilat %",110),("Tahsil Edilen",130),("Beklenen",130),
        ], height=6)
        t1f.pack(fill="both", expand=True)

        # Alt: Ödeyen daireler listesi
        t1alt = card_panel(t1pw, "Ev  Ödeme Yapan Daireler", "Seçili dönem içindeki tüm ödemeler")
        t1pw.add(t1alt, weight=60)
        t1ai = frm(t1alt, bg=T["bg2"]); t1ai.pack(fill="both", expand=True, padx=8, pady=6)
        t1af, self._dr_tree_odeyenler = scrolled(t1ai, [
            ("Daire",80),("Sakin",150),("Dönem",120),
            ("Tutar",110),("Yöntem",110),("Tarih",110),("Makbuz",130),
        ], height=10)
        t1af.pack(fill="both", expand=True)

        # ── Sekme 2: Aidat Ödemeyenler ────────────────────────────────────
        t2 = frm(nb, bg=T["bg2"]); nb.add(t2, text="  X  Aidat Ödemeyenler  ")

        t2_hdr = frm(t2, bg=T["bg2"]); t2_hdr.pack(fill="x", padx=12, pady=8)
        self._dr_lbl_odemeyenler = lbl(t2_hdr, "", fg=T["red"],
            bg=T["bg2"], font=("Segoe UI",10,"bold"))
        self._dr_lbl_odemeyenler.pack(side="left")
        btn(t2_hdr, "Yazdir Bu Listeyi Yazdır",
            lambda: self._dr_yazdir("odemeyenler"), "red").pack(side="right")

        t2pw = pane(t2, orient="vertical"); t2pw.pack(fill="both", expand=True, padx=8, pady=(0,8))

        # Üst: Ay bazında ödemeyen özet
        t2ust = card_panel(t2pw, "Takvim  Dönemsel Borç Özeti", "Ay bazında ödenmemiş aidatlar")
        t2pw.add(t2ust, weight=40)
        t2ui = frm(t2ust, bg=T["bg2"]); t2ui.pack(fill="both", expand=True, padx=8, pady=6)
        t2f, self._dr_tree_borclu_ay = scrolled(t2ui, [
            ("Dönem",130),("Ödemeyen Daire",130),("Toplam Borç",130),
            ("Toplam Faiz",120),("Kaybedilen Gelir",140),
        ], height=6)
        t2f.pack(fill="both", expand=True)

        # Alt: Borçlu daireler detay
        t2alt = card_panel(t2pw, "Ev  Borçlu Daire Detayı", "Seçili dönem içinde hiç ödeme yapmayan veya eksik ödeyen daireler")
        t2pw.add(t2alt, weight=60)
        t2ai = frm(t2alt, bg=T["bg2"]); t2ai.pack(fill="both", expand=True, padx=8, pady=6)
        t2af, self._dr_tree_odemeyenler = scrolled(t2ai, [
            ("Daire",80),("Sakin",150),("Dönem",120),
            ("Aidat",110),("Faiz",100),("Toplam Borç",120),("Son Ödeme",110),
        ], height=10)
        t2af.pack(fill="both", expand=True)

        # ── Sekme 3: Gelir Raporu ──────────────────────────────────────────
        t3 = frm(nb, bg=T["bg2"]); nb.add(t3, text="  Yukari  Gelir Raporu  ")

        t3_hdr = frm(t3, bg=T["bg2"]); t3_hdr.pack(fill="x", padx=12, pady=8)
        self._dr_lbl_gelir = lbl(t3_hdr, "", fg=T["green"],
            bg=T["bg2"], font=("Segoe UI",10,"bold"))
        self._dr_lbl_gelir.pack(side="left")
        btn(t3_hdr, "Yazdir Gelir Raporunu Yazdır",
            lambda: self._dr_yazdir("gelir"), "green").pack(side="right")

        t3pw = pane(t3, orient="vertical"); t3pw.pack(fill="both", expand=True, padx=8, pady=(0,8))

        # Üst: Kategoriye göre gelir özeti
        t3ust = card_panel(t3pw, "Grafik  Gelir Kategorisi Özeti", "Tür bazında toplam")
        t3pw.add(t3ust, weight=38)
        t3ui = frm(t3ust, bg=T["bg2"]); t3ui.pack(fill="both", expand=True, padx=8, pady=6)
        t3f, self._dr_tree_gelir_kat = scrolled(t3ui, [
            ("Kategori",160),("Kayıt Sayısı",120),("Toplam",130),("Oran %",100),
        ], height=6)
        t3f.pack(fill="both", expand=True)

        # Alt: Tüm gelir kayıtları
        t3alt = card_panel(t3pw, "Liste  Tüm Gelir Kayıtları", "")
        t3pw.add(t3alt, weight=62)
        t3ai = frm(t3alt, bg=T["bg2"]); t3ai.pack(fill="both", expand=True, padx=8, pady=6)
        t3af, self._dr_tree_gelirler = scrolled(t3ai, [
            ("Tarih",110),("Kategori",130),("Açıklama",280),("Tutar",120),
        ], height=10)
        t3af.pack(fill="both", expand=True)

        # ── Sekme 4: Gider Raporu ──────────────────────────────────────────
        t4 = frm(nb, bg=T["bg2"]); nb.add(t4, text="  Asagi  Gider Raporu  ")

        t4_hdr = frm(t4, bg=T["bg2"]); t4_hdr.pack(fill="x", padx=12, pady=8)
        self._dr_lbl_gider = lbl(t4_hdr, "", fg=T["red"],
            bg=T["bg2"], font=("Segoe UI",10,"bold"))
        self._dr_lbl_gider.pack(side="left")
        btn(t4_hdr, "Yazdir Gider Raporunu Yazdır",
            lambda: self._dr_yazdir("gider"), "red").pack(side="right")

        t4pw = pane(t4, orient="vertical"); t4pw.pack(fill="both", expand=True, padx=8, pady=(0,8))

        # Üst: Kategoriye göre gider özeti + bar grafik
        t4ust = card_panel(t4pw, "Grafik  Gider Kategorisi Özeti", "Tür bazında toplam")
        t4pw.add(t4ust, weight=38)
        t4ui = frm(t4ust, bg=T["bg2"]); t4ui.pack(fill="both", expand=True, padx=8, pady=6)
        t4f, self._dr_tree_gider_kat = scrolled(t4ui, [
            ("Kategori",160),("Kayıt Sayısı",120),("Toplam",130),("Oran %",100),
        ], height=6)
        t4f.pack(fill="both", expand=True)

        # Alt: Tüm gider kayıtları
        t4alt = card_panel(t4pw, "Liste  Tüm Gider Kayıtları", "")
        t4pw.add(t4alt, weight=62)
        t4ai = frm(t4alt, bg=T["bg2"]); t4ai.pack(fill="both", expand=True, padx=8, pady=6)
        t4af, self._dr_tree_giderler = scrolled(t4ai, [
            ("Tarih",110),("Kategori",130),("Açıklama",280),("Tutar",120),
        ], height=10)
        t4af.pack(fill="both", expand=True)

        # ── Sekme 5: Alacaklılar ──────────────────────────────────────────
        t5 = frm(nb, bg=T["bg2"]); nb.add(t5, text="  Liste  Alacaklılar  ")

        t5_hdr = frm(t5, bg=T["bg2"]); t5_hdr.pack(fill="x", padx=12, pady=8)
        self._dr_lbl_alacak = lbl(t5_hdr, "", fg=T["orange"],
            bg=T["bg2"], font=("Segoe UI",10,"bold"))
        self._dr_lbl_alacak.pack(side="left")
        btn(t5_hdr, "Yazdir Alacaklılar Raporunu Yazdır",
            lambda: self._dr_yazdir("alacaklilar"), "orange").pack(side="right")

        t5pw = pane(t5, orient="vertical"); t5pw.pack(fill="both", expand=True, padx=8, pady=(0,8))

        # Üst: Tür bazında özet
        t5ust = card_panel(t5pw, "Grafik  Alacaklı Özeti", "Tür bazında toplam / ödenen / kalan")
        t5pw.add(t5ust, weight=35)
        t5ui = frm(t5ust, bg=T["bg2"]); t5ui.pack(fill="both", expand=True, padx=8, pady=6)
        t5f, self._dr_tree_alacak_ozet = scrolled(t5ui, [
            ("Tür",130),("Kayıt",70),("Toplam Fatura",130),
            ("Toplam Ödenen",130),("Toplam Kalan",130),("Tamamlanan",100),
        ], height=5)
        t5f.pack(fill="both", expand=True)

        # Alt: Tüm alacaklılar
        t5alt = card_panel(t5pw, "Liste  Tüm Alacaklı Kayıtları", "Çift tıkla → ödeme yap")
        t5pw.add(t5alt, weight=65)
        t5ai = frm(t5alt, bg=T["bg2"]); t5ai.pack(fill="both", expand=True, padx=8, pady=6)
        t5af, self._dr_tree_alacaklar = scrolled(t5ai, [
            ("Fatura Tarihi",110),("Grup",110),("Tür / Açıklama",190),
            ("Daire",70),("Toplam",100),("Ödenen",100),("Kalan",100),("Durum",100),
        ], height=10)
        t5af.pack(fill="both", expand=True)

        # ── Sekme 6: Ortak Giderler ───────────────────────────────────────
        t6 = frm(nb, bg=T["bg2"]); nb.add(t6, text="  Arac  Ortak Giderler  ")

        t6_hdr = frm(t6, bg=T["bg2"]); t6_hdr.pack(fill="x", padx=12, pady=8)
        self._dr_lbl_og = lbl(t6_hdr, "", fg=T["gold2"],
            bg=T["bg2"], font=("Segoe UI",10,"bold"))
        self._dr_lbl_og.pack(side="left")
        btn(t6_hdr, "Yazdir Ortak Giderler Raporunu Yazdır",
            lambda: self._dr_yazdir("ortak_giderler"), "gold").pack(side="right")

        t6pw = pane(t6, orient="vertical"); t6pw.pack(fill="both", expand=True, padx=8, pady=(0,8))

        # Üst: Özet kart listesi
        t6ust = card_panel(t6pw, "Grafik  Ortak Gider Özeti", "Gider adı bazında toplam / ödenen / kalan")
        t6pw.add(t6ust, weight=35)
        t6ui = frm(t6ust, bg=T["bg2"]); t6ui.pack(fill="both", expand=True, padx=8, pady=6)
        t6f, self._dr_tree_og_ozet = scrolled(t6ui, [
            ("Gider Adı",170),(("Tarih"),100),(("Kapsam"),90),
            ("Toplam Fatura",120),(("Ödenen"),110),(("Kalan"),110),(("Durum"),110),
        ], height=5)
        t6f.pack(fill="both", expand=True)

        # Alt: Daire bazlı ödeme takibi
        t6alt = card_panel(t6pw, "Ev  Daire Bazlı Ortak Gider Ödemeleri", "Seçili dönem içindeki tüm ortak gider kayıtları")
        t6pw.add(t6alt, weight=65)
        t6ai = frm(t6alt, bg=T["bg2"]); t6ai.pack(fill="both", expand=True, padx=8, pady=6)
        t6af, self._dr_tree_og_daire = scrolled(t6ai, [
            ("Gider",150),(("Daire"),70),(("Sakin"),130),
            ("Toplam",100),(("Ödenen"),100),(("Kalan"),100),(("Durum"),100),
        ], height=10)
        t6af.pack(fill="both", expand=True)

        # ── Sekme 7: Uyarı Yazısı ─────────────────────────────────────────
        t7 = frm(nb, bg=T["bg2"]); nb.add(t7, text="  Mektup  Uyarı Yazısı  ")

        t7_hdr = frm(t7, bg=T["bg2"]); t7_hdr.pack(fill="x", padx=12, pady=8)
        lbl(t7_hdr, "Borçlu daireler için Word uyarı yazısı oluşturun.",
            fg=T["text2"], bg=T["bg2"], font=("Segoe UI",10)).pack(side="left")
        btn(t7_hdr, "Yazdir Tümü → PDF", self._uyari_tumu_pdf, "gold").pack(side="right", padx=(0,4))
        btn(t7_hdr, "Klasor Tümü → Word", self._uyari_tumu_yazdir, "purple").pack(side="right", padx=(0,4))

        t7_card = card_panel(t7, "Ev  Borçlu Daireler", "Seçili daireye uyarı yazısı oluşturun")
        t7_card.pack(fill="both", expand=True, padx=8, pady=(0,8))
        t7_ic = frm(t7_card, bg=T["bg2"]); t7_ic.pack(fill="both", expand=True, padx=8, pady=6)

        t7_tf, self._uw_tree = scrolled(t7_ic, [
            ("Daire",  80),
            ("Sakin", 180),
            ("Anapara",120),
            ("Faiz",  100),
            ("Toplam Borç",120),
            ("Son Ödeme",  120),
        ], height=14)
        t7_tf.pack(fill="both", expand=True)

        t7_btn_row = frm(t7_card, bg=T["bg2"]); t7_btn_row.pack(fill="x", padx=8, pady=(4,10))
        btn(t7_btn_row, "Dosya Seçili → PDF",
            self._uyari_secili_yazdir, "blue").pack(side="left", ipadx=10, ipady=4)
        btn(t7_btn_row, "Not Seçili → Word",
            self._uyari_secili_word, "gray").pack(side="left", padx=6, ipadx=8, ipady=4)
        btn(t7_btn_row, "Yazdir Tümü → PDF (Toplu)",
            self._uyari_tumu_pdf, "gold").pack(side="left", padx=6, ipadx=8, ipady=4)

        # İlk yükleme
        self._dr_yenile(nb)

    def _dr_yenile(self, nb=None):
        """Detaylı raporlar sekmesindeki tüm verileri yenile."""
        bas = self._dr_bas.get()
        son = self._dr_son.get()
        if bas > son:
            messagebox.showwarning("Tarih Hatası",
                "Başlangıç ayı, bitiş ayından büyük olamaz.", parent=self)
            return

        b         = self.data["bina"]
        daireler  = self.data["daireler"]
        odemeler  = self.data["odemeler"]
        gelirler  = self.data["gelirler"]
        giderler  = self.data["giderler"]
        donem_aylar = ay_listesi(bas, son)

        # ── 1. Ödeyen / Ödemeyen analizi ──────────────────────────────────
        # Dönem özet tablosu
        self._dr_tree_ay_ozet.delete(*self._dr_tree_ay_ozet.get_children())
        self._dr_tree_odeyenler.delete(*self._dr_tree_odeyenler.get_children())
        self._dr_tree_borclu_ay.delete(*self._dr_tree_borclu_ay.get_children())
        self._dr_tree_odemeyenler.delete(*self._dr_tree_odemeyenler.get_children())

        toplam_odeyenler = 0
        toplam_odemeyenler = 0
        toplam_tahsilat = 0
        toplam_borclu_tutar = 0

        for ay in donem_aylar:
            aidat_tut = ay_aidat(self.data, ay)
            beklenen  = b["daire"] * aidat_tut

            # Bu ayda ödeme yapan daireler
            ay_odemeleri = [o for o in odemeler if o["ay"] == ay]
            odeyenler_no  = {o["daireNo"] for o in ay_odemeleri}
            odemeyenler_d = [d for d in daireler if d["no"] not in odeyenler_no]

            ay_tahsilat = sum(o["tutar"] for o in ay_odemeleri)
            oran = int(ay_tahsilat / beklenen * 100) if beklenen else 0

            borç_tutar = sum(ay_aidat(self.data, ay, d["no"]) for d in odemeyenler_d)
            borç_faiz  = sum(
                faiz_hesapla(ay_aidat(self.data, ay, d["no"]), b["faiz"], ay, b["son_gun"])
                for d in odemeyenler_d
            )

            # Özet satırı
            self._dr_tree_ay_ozet.insert("", "end",
                tags=("odendi" if oran >= 100 else ("bekliyor" if oran > 0 else "borclu"),),
                values=(
                    ay_label(ay),
                    b["daire"],
                    len(odeyenler_no),
                    len(odemeyenler_d),
                    f"%{oran}",
                    fmt(ay_tahsilat),
                    fmt(beklenen),
                ))

            # Ödeyenler detay
            for o in ay_odemeleri:
                d = next((x for x in daireler if x["no"] == o["daireNo"]), {})
                self._dr_tree_odeyenler.insert("", "end",
                    tags=("odendi",),
                    values=(
                        f"Daire {o['daireNo']}",
                        d.get("isim") or "—",
                        ay_label(ay),
                        fmt(o["tutar"]),
                        o["yontem"],
                        tarih_str(o["tarih"]),
                        o["makbuzNo"],
                    ))
                toplam_tahsilat += o["tutar"]
                toplam_odeyenler += 1

            # Borçlu ay satırı
            if odemeyenler_d:
                self._dr_tree_borclu_ay.insert("", "end",
                    tags=("borclu",),
                    values=(
                        ay_label(ay),
                        len(odemeyenler_d),
                        fmt(borç_tutar),
                        fmt(borç_faiz),
                        fmt(borç_tutar + borç_faiz),
                    ))
                # Ödemeyen daireler detay
                for d in odemeyenler_d:
                    aidat_d = ay_aidat(self.data, ay, d["no"])
                    faiz = faiz_hesapla(aidat_d, b["faiz"], ay, b["son_gun"])
                    self._dr_tree_odemeyenler.insert("", "end",
                        tags=("borclu",),
                        values=(
                            f"Daire {d['no']}",
                            d.get("isim") or "—",
                            ay_label(ay),
                            fmt(aidat_d),
                            fmt(faiz) if faiz else "—",
                            fmt(aidat_d + faiz),
                            tarih_str(d.get("son_odeme","")) if d.get("son_odeme") else "—",
                        ))
                    toplam_borclu_tutar += aidat_d + faiz
                    toplam_odemeyenler += 1

        # Özet etiketleri
        self._dr_lbl_odeyenler.config(
            text=f"OK  Dönemde {toplam_odeyenler} ödeme  •  Toplam Tahsilat: {fmt(toplam_tahsilat)}")
        self._dr_lbl_odemeyenler.config(
            text=f"X  Dönemde {toplam_odemeyenler} ödenmemiş aidat  •  Toplam Borç: {fmt(toplam_borclu_tutar)}")

        # ── 2. Gelir Analizi ──────────────────────────────────────────────
        self._dr_tree_gelir_kat.delete(*self._dr_tree_gelir_kat.get_children())
        self._dr_tree_gelirler.delete(*self._dr_tree_gelirler.get_children())

        filtreli_gelirler = [
            g for g in gelirler
            if bas <= g["tarih"][:7] <= son
        ]
        toplam_gelir = sum(g["tutar"] for g in filtreli_gelirler)

        # Kategori özeti
        kat_gelir = {}
        for g in filtreli_gelirler:
            kat_gelir[g["tur"]] = kat_gelir.get(g["tur"], [0, 0])
            kat_gelir[g["tur"]][0] += g["tutar"]
            kat_gelir[g["tur"]][1] += 1
        for tur, (toplam, sayi) in sorted(kat_gelir.items(), key=lambda x: -x[1][0]):
            oran = int(toplam / toplam_gelir * 100) if toplam_gelir else 0
            self._dr_tree_gelir_kat.insert("", "end",
                tags=("gelir",),
                values=(tur, sayi, fmt(toplam), f"%{oran}"))

        # Detay kayıtlar
        for g in reversed(filtreli_gelirler):
            self._dr_tree_gelirler.insert("", "end",
                tags=("gelir",),
                values=(tarih_str(g["tarih"]), g["tur"],
                        g.get("aciklama") or "—", fmt(g["tutar"])))

        self._dr_lbl_gelir.config(
            text=f"Yukari  {len(filtreli_gelirler)} kayıt  •  Toplam Gelir: {fmt(toplam_gelir)}")

        # ── 3. Gider Analizi ──────────────────────────────────────────────
        self._dr_tree_gider_kat.delete(*self._dr_tree_gider_kat.get_children())
        self._dr_tree_giderler.delete(*self._dr_tree_giderler.get_children())

        filtreli_giderler = [
            g for g in giderler
            if bas <= g["tarih"][:7] <= son
        ]
        toplam_gider = sum(g["tutar"] for g in filtreli_giderler)

        # Kategori özeti
        kat_gider = {}
        for g in filtreli_giderler:
            kat_gider[g["tur"]] = kat_gider.get(g["tur"], [0, 0])
            kat_gider[g["tur"]][0] += g["tutar"]
            kat_gider[g["tur"]][1] += 1
        for tur, (toplam, sayi) in sorted(kat_gider.items(), key=lambda x: -x[1][0]):
            oran = int(toplam / toplam_gider * 100) if toplam_gider else 0
            self._dr_tree_gider_kat.insert("", "end",
                tags=("gider",),
                values=(tur, sayi, fmt(toplam), f"%{oran}"))

        # Detay kayıtlar
        for g in reversed(filtreli_giderler):
            self._dr_tree_giderler.insert("", "end",
                tags=("gider",),
                values=(tarih_str(g["tarih"]), g["tur"],
                        g.get("aciklama") or "—", fmt(g["tutar"])))

        self._dr_lbl_gider.config(
            text=f"Asagi  {len(filtreli_giderler)} kayıt  •  Toplam Gider: {fmt(toplam_gider)}")

        # ── 4. Alacaklılar Analizi ────────────────────────────────────────
        if not hasattr(self, "_dr_tree_alacak_ozet"):
            self._uw_tree_doldur(); return

        self._dr_tree_alacak_ozet.delete(*self._dr_tree_alacak_ozet.get_children())
        self._dr_tree_alacaklar.delete(*self._dr_tree_alacaklar.get_children())

        al_list = self.data.get("alacaklilar", [])
        # Dönem filtresi
        filtreli_al = []
        for a in al_list:
            ft = (a.get("fatura_tarihi") or a.get("kayit_tarihi",""))[:7]
            if not ft or (bas <= ft <= son):
                filtreli_al.append(a)

        # Tür bazında özet
        tur_ozet = {}
        for a in filtreli_al:
            t2 = a["tur"]
            if t2 not in tur_ozet:
                tur_ozet[t2] = {"sayi":0,"toplam":0.0,"odenen":0.0,"tamam":0}
            od = sum(p["tutar"] for p in a.get("odemeler",[]))
            tur_ozet[t2]["sayi"]   += 1
            tur_ozet[t2]["toplam"] += a["tutar"]
            tur_ozet[t2]["odenen"] += od
            if a.get("odendi",False) or round(a["tutar"]-od,2) <= 0:
                tur_ozet[t2]["tamam"] += 1

        for t2, d2 in sorted(tur_ozet.items(), key=lambda x: -x[1]["toplam"]):
            kalan = round(d2["toplam"] - d2["odenen"], 2)
            self._dr_tree_alacak_ozet.insert("","end",
                tags=("borclu" if kalan>0 else "odendi",),
                values=(t2, d2["sayi"], fmt(d2["toplam"]),
                        fmt(d2["odenen"]), fmt(kalan), d2["tamam"]))

        # Detay listesi
        toplam_al_kalan = 0.0
        for a in reversed(filtreli_al):
            od    = sum(p["tutar"] for p in a.get("odemeler",[]))
            kalan = round(a["tutar"]-od, 2)
            tam   = a.get("odendi",False) or kalan<=0
            if tam:       durum,tag = "v Tamamlandı","odendi"
            elif od > 0:  durum,tag = "! Kısmi",      "bekliyor"
            else:         durum,tag = "... Bekliyor",    "borclu"
            if not tam:   toplam_al_kalan += kalan
            ft = a.get("fatura_tarihi") or a.get("kayit_tarihi","")
            if a.get("tur") == "Ortak Gider":
                grup = "Arac Ortak Gider"
                dno = a.get("daire_no"); daire_txt = f"Daire {dno}" if dno else "—"
                acik_txt = a.get("kisi") or a.get("aciklama") or "—"
            else:
                grup = a["tur"]; daire_txt = "—"
                acik_txt = (a.get("kisi") or "") + ((" — " + a.get("aciklama","")) if a.get("aciklama") else "") or "—"
            self._dr_tree_alacaklar.insert("","end",tags=(tag,),
                values=(tarih_str(ft), grup, acik_txt, daire_txt,
                        fmt(a["tutar"]),
                        fmt(od) if od>0 else "—",
                        fmt(kalan) if kalan>0 else "v Tamam",
                        durum))

        if hasattr(self, "_dr_lbl_alacak"):
            self._dr_lbl_alacak.config(
                text=f"Liste  {len(filtreli_al)} kayıt  •  Toplam Kalan: {fmt(toplam_al_kalan)}")

        # ── 5. Ortak Giderler Analizi ─────────────────────────────────────
        if not hasattr(self, "_dr_tree_og_ozet"):
            self._uw_tree_doldur(); return

        self._dr_tree_og_ozet.delete(*self._dr_tree_og_ozet.get_children())
        self._dr_tree_og_daire.delete(*self._dr_tree_og_daire.get_children())

        og_list = self.data.get("ortak_giderler", [])
        al_list_og = [a for a in self.data.get("alacaklilar", []) if a.get("tur") == "Ortak Gider"]

        # Dönem filtresine göre giderleri seç
        filtreli_og = []
        for g in og_list:
            tarih_og = g.get("tarih", "")[:7]
            if not tarih_og or (bas <= tarih_og <= son):
                filtreli_og.append(g)

        toplam_og_kalan = 0.0
        toplam_og_fatura = 0.0

        for g in reversed(filtreli_og):
            al_g = [a for a in al_list_og if a.get("ortak_gider_id") == g["id"]]
            if al_g:
                fatura_top = sum(a["tutar"] for a in al_g)
                odenen_top = sum(sum(p["tutar"] for p in a.get("odemeler",[])) for a in al_g)
                kalan_top  = round(fatura_top - odenen_top, 2)
                tam_sayi   = sum(1 for a in al_g if a.get("odendi") or
                                 round(a["tutar"]-sum(p["tutar"] for p in a.get("odemeler",[])),2)<=0)
                if tam_sayi == len(al_g):
                    durum_txt, durum_tag = "OK Tümü Ödendi", "odendi"
                elif tam_sayi > 0 or odenen_top > 0:
                    durum_txt, durum_tag = f"! Kısmi ({tam_sayi}/{len(al_g)})", "bekliyor"
                else:
                    durum_txt, durum_tag = "... Bekliyor", "borclu"
                toplam_og_kalan  += kalan_top
                toplam_og_fatura += fatura_top
            else:
                fatura_top = 0; odenen_top = 0; kalan_top = 0
                durum_txt, durum_tag = "Liste Kaydedildi", "normal"

            kapsam_txt = "Tüm Daireler" if g.get("kapsam") == "tum" else f"{len(g.get('daire_nos',[]))} Daire"
            self._dr_tree_og_ozet.insert("", "end", tags=(durum_tag,),
                values=(g.get("ad",""), tarih_str(g.get("tarih","")),
                        kapsam_txt, fmt(fatura_top),
                        fmt(odenen_top) if odenen_top > 0 else "—",
                        fmt(kalan_top) if kalan_top > 0 else "v",
                        durum_txt))

            # Daire bazlı satırlar
            for a in al_g:
                d = next((x for x in self.data["daireler"] if x["no"] == a.get("daire_no")), {})
                isim = d.get("isim","") or "—"
                od_a  = sum(p["tutar"] for p in a.get("odemeler",[]))
                kal_a = round(a["tutar"] - od_a, 2)
                tam_a = a.get("odendi", False) or kal_a <= 0
                if tam_a:   dtag, dtxt = "odendi",  "v Ödendi"
                elif od_a>0: dtag, dtxt = "bekliyor", "! Kısmi"
                else:        dtag, dtxt = "borclu",  "... Bekliyor"
                self._dr_tree_og_daire.insert("", "end", tags=(dtag,),
                    values=(g.get("ad",""),
                            f"Daire {a.get('daire_no','?')}",
                            isim,
                            fmt(a["tutar"]),
                            fmt(od_a) if od_a > 0 else "—",
                            fmt(kal_a) if kal_a > 0 else "v",
                            dtxt))

        if hasattr(self, "_dr_lbl_og"):
            self._dr_lbl_og.config(
                text=f"Arac  {len(filtreli_og)} gider  •  Toplam Kalan: {fmt(toplam_og_kalan)}")

        self._uw_tree_doldur()

    def _dr_yazdir(self, mod=None):
        """Detaylı raporlar sekmesinden HTML rapor oluştur ve tarayıcıda aç."""
        b        = self.data["bina"]
        bas      = self._dr_bas.get()
        son      = self._dr_son.get()
        donem    = f"{ay_label(bas)} – {ay_label(son)}"
        tarih_now = datetime.datetime.now().strftime("%d.%m.%Y %H:%M")

        def tablo(basliklar, satirlar, renk="#111"):
            th = "".join(f"<th>{h}</th>" for h in basliklar)
            td_rows = "".join(
                "<tr>" + "".join(f"<td>{v}</td>" for v in sat) + "</tr>"
                for sat in satirlar
            )
            return (
                f"<table><thead><tr style='background:{renk};color:#fff'>{th}</tr></thead>"
                f"<tbody>{td_rows}</tbody></table>"
            )

        bolumler = []

        # ── 1. Aidat Ödeyenler ────────────────────────────────────────────
        if mod in (None, "odeyenler"):
            # Dönemsel özet
            sat_ozet = []
            for iid in self._dr_tree_ay_ozet.get_children():
                sat_ozet.append(self._dr_tree_ay_ozet.item(iid)["values"])
            # Detay
            sat_det = []
            for iid in self._dr_tree_odeyenler.get_children():
                sat_det.append(self._dr_tree_odeyenler.item(iid)["values"])

            bolumler.append(
                "<h2 style='color:#2ecc71'>OK Aidat Ödeyenler</h2>" +
                "<h3>Dönemsel Tahsilat Özeti</h3>" +
                tablo(["Dönem","Toplam Daire","Ödeyen","Ödemeyen","Tahsilat %",
                        "Tahsil Edilen","Beklenen"], sat_ozet, "#1a5c2a") +
                "<h3>Ödeme Yapan Daireler</h3>" +
                tablo(["Daire","Sakin","Dönem","Tutar","Yöntem","Tarih","Makbuz"],
                      sat_det, "#2ecc71")
            )
            if mod == "odeyenler":
                bolumler = [bolumler[-1]]

        # ── 2. Aidat Ödemeyenler ──────────────────────────────────────────
        if mod in (None, "odemeyenler"):
            sat_bay = []
            for iid in self._dr_tree_borclu_ay.get_children():
                sat_bay.append(self._dr_tree_borclu_ay.item(iid)["values"])
            sat_bod = []
            for iid in self._dr_tree_odemeyenler.get_children():
                sat_bod.append(self._dr_tree_odemeyenler.item(iid)["values"])

            bolumler.append(
                "<h2 style='color:#e74c3c'>X Aidat Ödemeyenler</h2>" +
                "<h3>Dönemsel Borç Özeti</h3>" +
                tablo(["Dönem","Ödemeyen Daire","Toplam Borç","Toplam Faiz",
                        "Kaybedilen Gelir"], sat_bay, "#7a1a1a") +
                "<h3>Borçlu Daire Detayı</h3>" +
                tablo(["Daire","Sakin","Dönem","Aidat","Faiz","Toplam Borç","Son Ödeme"],
                      sat_bod, "#e74c3c")
            )
            if mod == "odemeyenler":
                bolumler = [bolumler[-1]]

        # ── 3. Gelir Raporu ───────────────────────────────────────────────
        if mod in (None, "gelir"):
            sat_gk = []
            for iid in self._dr_tree_gelir_kat.get_children():
                sat_gk.append(self._dr_tree_gelir_kat.item(iid)["values"])
            sat_gd = []
            for iid in self._dr_tree_gelirler.get_children():
                sat_gd.append(self._dr_tree_gelirler.item(iid)["values"])

            # Toplam satırı
            toplam_gelir_tut = sum(
                float(str(s[2]).replace("₺","").replace(".","").replace(",",".").replace("−","").strip() or 0)
                for s in sat_gk
            )
            toplam_gelir_kayit = sum(int(s[1]) for s in sat_gk)
            toplam_satir = (
                "<tfoot><tr style='background:#1a5c2a;color:#fff;font-weight:bold'>"
                f"<td>TOPLAM</td><td>{toplam_gelir_kayit} kayıt</td>"
                f"<td>{fmt(toplam_gelir_tut)}</td><td>%100</td>"
                "</tr></tfoot>"
            )
            ozet_tablo = (
                "<table><thead><tr style='background:#1a5c2a;color:#fff'>"
                + "".join(f"<th>{h}</th>" for h in ["Kategori","Kayıt Sayısı","Toplam","Oran %"])
                + "</tr></thead><tbody>"
                + "".join("<tr>" + "".join(f"<td>{v}</td>" for v in s) + "</tr>" for s in sat_gk)
                + "</tbody>" + toplam_satir + "</table>"
            )

            bolumler.append(
                "<h2 style='color:#2ecc71'>Yukari Gelir Raporu</h2>" +
                "<h3>Gelir Kategorisi Özeti</h3>" +
                ozet_tablo +
                "<div class='no-print'><h3>Tüm Gelir Kayıtları</h3>" +
                tablo(["Tarih","Kategori","Açıklama","Tutar"], sat_gd, "#2ecc71") +
                "</div>"
            )
            if mod == "gelir":
                bolumler = [bolumler[-1]]

        # ── 4. Gider Raporu ───────────────────────────────────────────────
        if mod in (None, "gider"):
            sat_gik = []
            for iid in self._dr_tree_gider_kat.get_children():
                sat_gik.append(self._dr_tree_gider_kat.item(iid)["values"])
            sat_gid = []
            for iid in self._dr_tree_giderler.get_children():
                sat_gid.append(self._dr_tree_giderler.item(iid)["values"])

            bolumler.append(
                "<h2 style='color:#e74c3c'>Asagi Gider Raporu</h2>" +
                "<h3>Gider Kategorisi Özeti</h3>" +
                tablo(["Kategori","Kayıt Sayısı","Toplam","Oran %"], sat_gik, "#7a1a1a") +
                "<h3>Tüm Gider Kayıtları</h3>" +
                tablo(["Tarih","Kategori","Açıklama","Tutar"], sat_gid, "#e74c3c")
            )
            if mod == "gider":
                bolumler = [bolumler[-1]]

        # ── 5. Alacaklılar ────────────────────────────────────────────────
        if mod in (None, "alacaklilar"):
            sat_ao = []
            for iid in self._dr_tree_alacak_ozet.get_children():
                sat_ao.append(self._dr_tree_alacak_ozet.item(iid)["values"])
            sat_ad = []
            for iid in self._dr_tree_alacaklar.get_children():
                sat_ad.append(self._dr_tree_alacaklar.item(iid)["values"])

            bolumler.append(
                "<h2 style='color:#f39c12'>Liste Alacaklılar Raporu</h2>" +
                "<h3>Alacaklı Özeti</h3>" +
                tablo(["Tür","Kayıt","Toplam Fatura","Toplam Ödenen",
                        "Toplam Kalan","Tamamlanan"], sat_ao, "#7a4a00") +
                "<h3>Tüm Alacaklı Kayıtları</h3>" +
                tablo(["Fatura Tarihi","Grup","Tür / Açıklama","Daire",
                        "Toplam","Ödenen","Kalan","Durum"], sat_ad, "#f39c12")
            )
            if mod == "alacaklilar":
                bolumler = [bolumler[-1]]

        # ── 6. Ortak Giderler ─────────────────────────────────────────────
        if mod in (None, "ortak_giderler"):
            sat_oo = []
            for iid in self._dr_tree_og_ozet.get_children():
                sat_oo.append(self._dr_tree_og_ozet.item(iid)["values"])
            sat_od = []
            for iid in self._dr_tree_og_daire.get_children():
                sat_od.append(self._dr_tree_og_daire.item(iid)["values"])

            bolumler.append(
                "<h2 style='color:#c9a84c'>Arac Ortak Giderler Raporu</h2>" +
                "<h3>Ortak Gider Özeti</h3>" +
                tablo(["Gider Adı","Tarih","Kapsam","Toplam Fatura",
                        "Ödenen","Kalan","Durum"], sat_oo, "#4a3a00") +
                "<h3>Daire Bazlı Ortak Gider Ödemeleri</h3>" +
                tablo(["Gider","Daire","Sakin","Toplam","Ödenen","Kalan","Durum"],
                      sat_od, "#c9a84c")
            )
            if mod == "ortak_giderler":
                bolumler = [bolumler[-1]]

        baslik_map = {
            "odeyenler":    "Aidat Ödeyenler",
            "odemeyenler":  "Aidat Ödemeyenler",
            "gelir":        "Gelir Raporu",
            "gider":        "Gider Raporu",
            "alacaklilar":  "Alacaklılar Raporu",
            "ortak_giderler": "Ortak Giderler Raporu",
        }
        sayfa_baslik = baslik_map.get(mod, "Detaylı Rapor — Tüm Sekmeler")

        icerik = "\n<hr style='margin:32px 0;border:none;border-top:1px solid #ddd'>\n".join(bolumler)

        html = f"""<!DOCTYPE html><html lang="tr"><head><meta charset="UTF-8">
<title>{sayfa_baslik} — {b['adi']}</title>
<style>
  body{{font-family:Arial,sans-serif;background:#f0f0f0;padding:30px;}}
  .card{{background:#fff;max-width:960px;margin:0 auto;padding:32px;
         border-radius:8px;box-shadow:0 4px 20px rgba(0,0,0,.12);}}
  h1{{font-size:20px;margin-bottom:4px;}}
  h2{{font-size:16px;margin-top:28px;margin-bottom:6px;}}
  h3{{font-size:13px;color:#555;margin:16px 0 4px;}}
  table{{width:100%;border-collapse:collapse;font-size:13px;margin-bottom:8px;}}
  th{{padding:8px 10px;text-align:left;}}
  td{{padding:7px 10px;border-bottom:1px solid #eee;}}
  tr:hover td{{background:#fafafa;}}
  .meta{{color:#888;font-size:12px;margin-bottom:20px;}}
  .np{{text-align:center;margin-bottom:20px;}}
  button{{padding:9px 24px;background:#111;color:#fff;border:none;
          border-radius:4px;cursor:pointer;font-size:14px;margin:0 4px;}}
  @media print{{.np{{display:none;}}.no-print{{display:none;}}body{{background:#fff;padding:0;}}
                h2{{page-break-before:auto;}} table{{page-break-inside:auto;}}}}
</style></head><body>
<div class="np">
  <button onclick="window.print()">Yazdir Yazdır / PDF</button>
</div>
<div class="card">
  <h1>Bina {b['adi']} — {sayfa_baslik}</h1>
  <p class="meta">Dönem: {donem} &nbsp;|&nbsp; Oluşturma: {tarih_now}</p>
  {icerik}
</div>
</body></html>"""

        tmp = tempfile.NamedTemporaryFile(suffix=".html", delete=False,
                                          mode="w", encoding="utf-8")
        tmp.write(html); tmp.close()
        webbrowser.open(f"file:///{tmp.name}")

    def _uw_tree_doldur(self):
        """Uyarı yazısı sekmesindeki borçlu daire listesini anlık hesaplayarak doldur."""
        if not hasattr(self, "_uw_tree"):
            return
        self._uw_tree.delete(*self._uw_tree.get_children())
        b        = self.data["bina"]
        ba       = buay()
        tum_aylar = ay_listesi(b.get("baslangic", ba), ba)
        _bugun   = datetime.date.today()
        _son_gun = b.get("son_gun", 10)
        for d in sorted(self.data["daireler"], key=lambda x: x["no"]):
            no = d["no"]
            odenmemis = [
                ay for ay in tum_aylar
                if not any(o["daireNo"] == no and o["ay"] == ay
                           for o in self.data["odemeler"])
                and not (ay == buay() and _bugun.day < _son_gun)
            ]
            ana  = round(sum(ay_aidat(self.data, ay, no) for ay in odenmemis), 2)
            faiz = round(sum(faiz_hesapla(ay_aidat(self.data, ay, no),
                             b["faiz"], ay, b["son_gun"])
                             for ay in odenmemis), 2)
            if ana <= 0 and faiz <= 0:
                continue
            isim   = d.get("isim","") or "—"
            son_od = d.get("son_odeme","") or "—"
            self._uw_tree.insert("", "end", iid=str(no), tags=("borclu",),
                values=(f"Daire {no}", isim, fmt(ana), fmt(faiz),
                        fmt(round(ana+faiz,2)), son_od))

    # ── Uyarı Yazısı metodları ─────────────────────────────────────────────
    def _uyari_sablon_doc(self, daire_no, ana_para):
        """
        Şablon Word dosyasını bellek üzerinde düzenleyerek ZIP baytı döndür.
        Daire No ve Ana Para yer tutucularını değiştir.
        Ana Para XML içinde birden fazla run'a bölünmüş olduğundan literal değiştirme yapılır.
        """
        import zipfile, io
        LQ = "\u201c"
        RQ = "\u201d"

        # Her zaman gömülü şablonu kullan — böylece DATA_DIR'deki eski/bozuk
        # dosyadan bağımsız olarak doğru XML yapısı garantilenir.
        import base64 as _b64
        sablon_yolu = DATA_DIR / "uyari_sablonu.docx"
        _B64_SABLON = self._uyari_gomulu_b64()
        sablon_bytes = _b64.b64decode(_B64_SABLON)
        # DATA_DIR'e de yaz (güncel tut)
        try:
            sablon_yolu.write_bytes(sablon_bytes)
        except Exception:
            pass

        _rpr = (
            '<w:rPr><w:rFonts w:ascii="Roboto" w:hAnsi="Roboto"/>'
            '<w:color w:val="7C7C7C"/></w:rPr>'
        )
        # Ana Para XML'de 4 run'a bölünmüş: A / na<space> / P / ara" ...
        _ANA_OLD_PARTS = [
            '<w:r w:rsidR="00947FF7">',
            _rpr,
            '<w:t>A</w:t></w:r>',
            '<w:r>',
            _rpr,
            '<w:t xml:space="preserve">na </w:t></w:r>',
            '<w:r w:rsidR="00947FF7">',
            _rpr,
            '<w:t>P</w:t></w:r>',
            '<w:r>',
            _rpr,
            '<w:t>ara' + RQ,
        ]
        ANA_PARA_SPLIT = "".join(_ANA_OLD_PARTS)

        buf_in  = io.BytesIO(sablon_bytes)
        buf_out = io.BytesIO()
        with zipfile.ZipFile(buf_in,  "r") as zin, \
             zipfile.ZipFile(buf_out, "w", zipfile.ZIP_DEFLATED) as zout:
            for item in zin.infolist():
                data_bytes = zin.read(item.filename)
                if item.filename == "word/document.xml":
                    xml_str = data_bytes.decode("utf-8")
                    # 1. Daire No (akilli tırnak)
                    xml_str = xml_str.replace(
                        LQ + "Daire No" + RQ,
                        LQ + str(daire_no) + RQ)
                    xml_str = xml_str.replace('"Daire No"', '"' + str(daire_no) + '"')
                    # 2. Ana Para (XML'de bölünmüş run → tek run)
                    ANA_PARA_NEW = '<w:r>' + _rpr + '<w:t xml:space="preserve">' + str(ana_para) + RQ
                    xml_str = xml_str.replace(ANA_PARA_SPLIT, ANA_PARA_NEW)
                    # Yedek: düz metin
                    xml_str = xml_str.replace(
                        LQ + "Ana Para" + RQ, LQ + str(ana_para) + RQ)
                    xml_str = xml_str.replace('"Ana Para"', '"' + str(ana_para) + '"')
                    xml_str = xml_str.replace("Ana Para", str(ana_para))
                    data_bytes = xml_str.encode("utf-8")
                zout.writestr(item, data_bytes)
        return buf_out.getvalue()

    def _uyari_gomulu_b64(self):
        """Gömülü şablon DOCX'in base64 verisi."""
        return (
            "UEsDBBQABgAIAAAAIQDfpNJsWgEAACAFAAATAAgCW0NvbnRlbnRfVHlwZXNdLnhtbCCiBAIooAACAAAAAAAAAAAA"
            "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA"
            "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA"
            "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA"
            "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA"
            "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA"
            "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA"
            "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA"
            "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAC0lMtuwjAQRfeV+g+Rt1Vi6KKqKgKLPpYt"
            "UukHGHsCVv2Sx7z+vhMCUVUBkQpsIiUz994zVsaD0dqabAkRtXcl6xc9loGTXmk3K9nX5C1/ZBkm4ZQw3kHJNoBs"
            "NLy9GUw2ATAjtcOSzVMKT5yjnIMVWPgAjiqVj1Ykeo0zHoT8FjPg973eA5feJXApT7UHGw5eoBILk7LXNX1uSCIY"
            "ZNlz01hnlUyEYLQUiep86dSflHyXUJBy24NzHfCOGhg/mFBXjgfsdB90NFEryMYipndhqYuvfFRcebmwpCxO2xzg"
            "9FWlJbT62i1ELwGRztyaoq1Yod2e/ygHpo0BvDxF49sdDymR4BoAO+dOhBVMP69G8cu8E6Si3ImYGrg8RmvdCZFo"
            "A6F59s/m2NqciqTOcfQBaaPjP8ber2ytzmngADHp039dm0jWZ88H9W2gQB3I5tv7bfgDAAD//wMAUEsDBBQABgAI"
            "AAAAIQAekRq37wAAAE4CAAALAAgCX3JlbHMvLnJlbHMgogQCKKAAAgAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA"
            "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA"
            "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA"
            "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA"
            "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA"
            "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA"
            "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA"
            "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA"
            "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAArJLBasMwDEDvg/2D0b1R2sEYo04vY9DbGNkHCFtJTBPb2GrX/v082NgC"
            "XelhR8vS05PQenOcRnXglF3wGpZVDYq9Cdb5XsNb+7x4AJWFvKUxeNZw4gyb5vZm/cojSSnKg4tZFYrPGgaR+IiY"
            "zcAT5SpE9uWnC2kiKc/UYySzo55xVdf3mH4zoJkx1dZqSFt7B6o9Rb6GHbrOGX4KZj+xlzMtkI/C3rJdxFTqk7gy"
            "jWop9SwabDAvJZyRYqwKGvC80ep6o7+nxYmFLAmhCYkv+3xmXBJa/ueK5hk/Nu8hWbRf4W8bnF1B8wEAAP//AwBQ"
            "SwMEFAAGAAgAAAAhACgqyzneBwAAICQAABEAAAB3b3JkL2RvY3VtZW50LnhtbOxazW7jRhK+L7Dv0NBpF/CYkqh/"
            "xA4kSzKMGTvG2MnCubXIltRRs5toNm1LpzxFTrvw3nIdX+bk00rzInmSVDVJSbbGA1qTeDJY27BE9s/X1V1ffVWU"
            "/M2314Egl0xHXMm9Qmm3WCBMesrncrRX+P68/6pRIJGh0qdCSbZXmLKo8O3+3//2zVXLV14cMGkIQMiodRV6e4Wx"
            "MWHLcSJvzAIa7Qbc0ypSQ7PrqcBRwyH3mHOltO+Ui6WivQq18lgUwXoHVF7SqJDCedf50HxNr2AyAlYcb0y1Ydcr"
            "jNKTQapO02lsApW3AIIdlkubUO6ToWoOWrUBVNkKCKzaQKpuh/SRzdW2QypvItW3Q3I3kRrbIW3QKdgkuAqZhM6h"
            "0gE1cKtHTkD1JA5fAXBIDR9wwc0UMIu1DIZyOdnCIpi1RAhc/8kIdSdQPhOun6GovUKsZSud/2o5H01vJfPTt+UM"
            "JvItC8s1HXZtRGSyuTrP2SXTu6mw2FNzNBNwjkpGYx4u1SHYFg06xxnI5acO4DIQ2birsJQz1B6Ttm7ihhVgHvNT"
            "3wUisfzTiKViDm8ixHJGHhPur5lZEgCDVwtvdTRrh1vKKT4ZQHkDoOaxnMkiw2ikGI63im7E4TnDKsNJvII4fHWw"
            "pZwa+NCYNQA/fhJE2c3swDecvoYV+cYfPw0u85GDc6mhYxotgyZBHOYUggyxsoaYEEwob6lniMmedmjVJeA0WPNh"
            "OPq8QD3UKg5XaPzz0I5Wkn2FxdMTsNKAXxeh6POMORvTEJQ88FpHI6k0HQiwCMKXQAQS6wF8BSLjm71k17Yd+ZNe"
            "DAVe+DFBSSzsQxE4UP4U30PoqLRCqukRxFCtXC91yx2oHbEVUqjB1nr6A60tKDj9t3uFYrHXKNbb7rKpy4Y0Fmaz"
            "53StyS54qu3bmZkKsLx1SYGSJ0gt8S82KDjYGY39rMcTjGrE8pRQkItobBTeDrmA3r79SSb95C3nQNZgOmnVyXK6"
            "r6SJYACNPA7seKsGKgEat2W0arBz7FIZWP0Af7HDSdGc5SaSl/T63oYORbr8o+syGpl2xOk5UAIcGtCflO6lbdvZ"
            "ddUy++3T9tvz4/YJuZj/ctI7X7w7Xrw76fZOSGfx7s0h/PVOuot3b4975IceOW+/6Z2Si/aPR2dHiGISrGSHm+So"
            "uO1up1vrft3kgOMc/19S4/V3J9+37rn5gZV/4DngIv/7b5tDEiIdpb2YzH/xYTPknAo24DnIVu9UGt1us/2AbMX6"
            "QbVfLL+Q7XGy/cFuPKPTxa0kP3J9yUgb3GMCKhe35IxOuBRM850c3nSrtXatWS/e96br9rqdfq/y4s1n8+aGGy/m"
            "7yUzPOBECXDWhER8Bl5lO4QuBwWL2xlywJ/fzZgUnAzoBFp3iGEBjOYTApg+X9wwTZQ2ADLiPtNIDsklmVD94WZx"
            "K6ik8FQdsIgT6gMeJQMmOAyTTBI6FYvbCaEoGWAILoeWBYBABlyTCZsRHwpaMqZmcavh0TKAdXgEiZ7Pdsl3dlkK"
            "a9jJmcWR0rGM4hkaaI21WB9u2IQLn5FJLARuUCwN4/NfwWIzvwtIlBEcpkGbtQxsQhlLdkZm1B6P9CmZ0hAKOzjP"
            "EZzdJGATw3yud/Nk1V6106lXMX+uC1235pY7zc5LaDxbaFzQ0JKUcOmBh62XgUAy9mL08G8//7tLuWbkRP32838g"
            "BBbA6QBcDpyavxfIGBkH4FTgMTAVRk4ZkMbgxIdZd92pzUq9369bP/1JG2s/X9LHZ55WFFIPypZQs4iB2hT2IdK/"
            "7AmcPmPZAwxAfpy/ISY2VsiAPCgbMgAB+nCTCsnAVkVK+PHiJgZtikKgCmiGwDEGpIO8hlHH8zsx4VMQaLiVsYxR"
            "yA5BMAV5HetYEKNCoCwoYpQshOyT1Gou1SjoZDR/b9V8TVjXZBcs5GMgOohhZiPKoA/ggaX9lEYUlgEtDEk0v9PM"
            "Q/n/cCNQhq1uMg/ULp/WlfvNat1tP9C6WqPYa/Zs64vWPVNRB5nQZ5iwIZvbRIdZDvy/8xhVZTyLgTkxEAIzK9II"
            "uEOH9npGSSyAFEAJVMWEVQboxQ0fUA0XdfKPC2D3P4GOdxKzLGTQiEug3YDKCWR1FtFBkreRc5jikV5YkEygEEGu"
            "2fyKJYutC8C5mnsUQiYpANLni6kV8YAmyXhMBa6xk7LY8nf+6zp/PVhAiTjcwa2l/Wi4rxY3thOAYJcYdvQynkAI"
            "YZUyv/M0M1l9A5u4dxRTKmHxFN6exziO4BeiF8JlBHNmjMAtBuMsT9iU3Ea50ai/PHh/6bDpZgXwehloi8v7xbJV"
            "XSxNsSi2Q+bvZxgOSXXJsACdANkse3cIn3IbF1hwQAawhM5Bi1r9oNg5aPS/blpoPhqb+7z44h+XPMKmv8THOPA8"
            "PrLSmFBN0DzP36VSp93r1kv3qVIsu9Vu3619xVT5cyQk2/ypfrDJbZYdJK8HUV4P22f0T1WsfwUefrzS/+gnCzn4"
            "6R5U681eOU+GO6hWG71V0xo/056Eacwzp0s/rvXatUdnM+i6wrho4vf7cAZwXWu4jWR2ODqmOBmqa2ivlJLlkH2r"
            "Wzgso4LVvWDDtd4xoyDt+IElfpfSGipl1m5HsbG3xeVpo+PS08QxttlX3qHGr3VaUMWwU248sNKt2UlOtkV7mXyZ"
            "46z+tWf/dwAAAP//AwBQSwMEFAAGAAgAAAAhANZks1H0AAAAMQMAABwACAF3b3JkL19yZWxzL2RvY3VtZW50Lnht"
            "bC5yZWxzIKIEASigAAEAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA"
            "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA"
            "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA"
            "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA"
            "AAAAAAAArJLLasMwEEX3hf6DmH0tO31QQuRsSiHb1v0ARR4/qCwJzfThv69ISevQYLrwcq6Yc8+ANtvPwYp3jNR7"
            "p6DIchDojK971yp4qR6v7kEQa1dr6x0qGJFgW15ebJ7Qak5L1PWBRKI4UtAxh7WUZDocNGU+oEsvjY+D5jTGVgZt"
            "XnWLcpXndzJOGVCeMMWuVhB39TWIagz4H7Zvmt7ggzdvAzo+UyE/cP+MzOk4SlgdW2QFkzBLRJDnRVZLitAfi2My"
            "p1AsqsCjxanAYZ6rv12yntMu/rYfxu+wmHO4WdKh8Y4rvbcTj5/oKCFPPnr5BQAA//8DAFBLAwQUAAYACAAAACEA"
            "m4+qHtQGAADRIAAAFQAAAHdvcmQvdGhlbWUvdGhlbWUxLnhtbOxZzW/bNhS/D9j/IOju+kvyR1CnsGW7aZu0QeN2"
            "6JGWaYkxJRokncQoCgztaZcBA7phlwG77TAM67ACK3bZZf9JgBZb90eMpGxLtKmlbdKhKBIDMT9+7/HH9x4fn6Wr"
            "104ibB1ByhCJW3b5Ssm2YOyTEYqDln1v0C80bItxEI8AJjFs2XPI7Gvbn35yFWzxEEbQEvIx2wItO+R8ulUsMl8M"
            "A3aFTGEs5saERoCLLg2KIwqOhd4IFyulUq0YARTbVgwiofbOeIx8aA2k6J+/2NtL9T0s/sWcyQEf0wOpHGoyCjua"
            "lOUXmzMPU+sI4JYtVhqR4wE84baFAeNiomWX1J9d3L5aXAlhniObkeurv4XcQmA0qSg5GgxXgo7jOrX2Sr8CYL6J"
            "69V7tV5tpU8BgO+LnSZcdJ31iucssBlQ0jTo7ta71bKGz+ivbuDbrvxoeAVKms4Gvt/3UhtmQEnT3cC7nWanq+tX"
            "oKRZ28DXS+2uU9fwChRiFE820CW3VvWWu11BxgTvGOFN1+nXKwt4iipmoiuRj3lerEXgkNC+ACjnAo5ii8+ncAx8"
            "gfMARkOKrF0UhCLwpiAmTAyXKqV+qSr+y4+jWsqjYAuCjHQy5LONIcnHYj5FU96ybwqtdgby8sWL08fPTx//dvrk"
            "yenjnxdrb8rtgDjIyr3+4at/vvvc+vvX718//dqMZ1n8q5++ePX7H/+lnmu0vnn26vmzl99++dePTw3wNgXDLHyA"
            "Isis2/DYuksisUHDAnBI305iEAKUlWjHAQMxkDIGdI+HGvr2HGBgwHWgbsf7VKQLE/D67FAjfBDSGUcG4K0w0oB7"
            "hOAOocY93ZJrZa0wiwPz4nSWxd0F4Mi0trfm5d5sKuIemVR6IdRo7mPhchDAGHJLzpEJhAaxBwhpdt1DPiWMjLn1"
            "AFkdgIwmGaChFk2p0A6KhF/mJoLC35pt9u5bHYJN6rvwSEeKswGwSSXEmhmvgxkHkZExiHAWuQt4aCJ5MKe+ZnDG"
            "hacDiInVG0HGTDJ36Fyje0ukGbPb9/A80pGUo4kJuQsIySK7ZOKFIJoaOaM4zGJvsIkIUWDtE24kQfQTIvvCDyDO"
            "dfd9BDV3n32274k0ZA4QOTOjpiMBiX4e53gMoEl5m0Zaim1TZIyOzizQQnsXQgyOwQhC694NE55MNZunpG+GIqvs"
            "QJNtbgI9VmU/hkxWSyemRLyLmBayBzAgOXz25muJZw7iCNA8zbcnesj0xFUXGeMV+xMtlSIqD62ZxB0WafvL1bof"
            "Ai2sZJ+Z43VONf+9yRkTMofvIAPfWkYk9je2zQBgbYE0YAZAVBmmdCtENPenIvI4KbGZUW6sH9rUDcW1oidC8ZkV"
            "0Frt4/4/tY9B4mKqHjPwPPVOXkpZr3LycOu1jUfoCH34pU0XzOJ9KG4TA/SysrmsbD76yibvPF/WM5f1zGU9YxZ5"
            "D/VMWsKoB0HLxz1KS5T77GeMMD7gcwx3mSp+mDj7o74YVB0ltHrUNA1Fc7GchgsoUG2LEv4Z4uFBCKZimbJaIWAL"
            "1QGzpoSJ8kkNG3XLCTyL9sgoGS2Xl083hQDg6bgov5bjoljjyWitnj7GW6lXvUA9bl0SkLJvQyKzmE6iaiBRXw6e"
            "QULt7EJYNA0sGlJ9Lgv1tfCKuJwsIB+Nu07CSISbCOmR9FMiv/TuhXs6z5j6tiuG7TUl14vxtEYiE246iUwYhuLy"
            "WB++YF83U5dq9KQpNmnUG+/D1zKJrOUGHOs961icuaor1Phg2rLH4oeTaEZToY/JTAVwELdsny8M/S6ZZUoZ7wIW"
            "JjA1lew/QhxSC6NIxHrWDThOuZUrdbnHD5Rcs/ThWU59ZZ0Mx2Po85yRtCvmEiXG2XOCZYfMBOmDcHRsDfGM3gXC"
            "UG69LA04QoyvrDlCNBPcqRXX0tXiKGpvXdIjCvA0BIsbJZvME7hqr+hk9qGYru9K7y82Mwykk859654tJCcySTPn"
            "ApG3pjl/vL9LPsMqzfsaqyR1r+e65jLX5d0S578QMtTSxTRqkrGBWjqqU7vAgiCz3Co08+6Ii74N1qNWXhDLulL1"
            "Nl5vk+GhiPyuqFZnmDNFVfxqocBbvphMMoEaXWaXE27NKGrZD0tu2/EqrlcoNdxewak6pULDbVcLbdetlntuudTt"
            "VB4Jo/AwKrvJ2n3xYx/PF+/v1fjGO/xoWWpf8UlUJKoOLiph9Q6/XDG+w5fztoWEZR7WKv1mtdmpFZrVdr/gdDuN"
            "QtOrdQrdmlfv9rue22j2H9nWkQI77arn1HqNQq3seQWnVpL0G81C3alU2k693eg57UcLW4udL7+X5lW8tv8FAAD/"
            "/wMAUEsDBBQABgAIAAAAIQDQbjEUdgQAACINAAARAAAAd29yZC9zZXR0aW5ncy54bWy0V1Fv4zYMfh+w/xDkeTk7"
            "ju3kvEsPTVKvPTS7ockwYG+yJSdCJcuQ5KS5w/77KNmK0zU4NB360sj8yI8URZHqp89PnPV2RCoqyml/+MHv90iZ"
            "C0zLzbT/5zodTPo9pVGJERMlmfYPRPU/X/3806d9oojWoKZ6QFGqhOfT/lbrKvE8lW8JR+qDqEgJYCEkRxo+5cbj"
            "SD7W1SAXvEKaZpRRffAC34/7LY2Y9mtZJi3FgNNcCiUKbUwSURQ0J+2Ps5Cv8duYLERec1Jq69GThEEMolRbWinH"
            "xt/KBuDWkex+tIkdZ05vP/Rfsd29kPho8ZrwjEElRU6UggPizAVIy85x+ILo6PsD+G63aKnAfOjb1Wnk0WUEwQuC"
            "OCdPl3FMWg4PLE95KL6MJz7y0C6xw/htwZwQ4PoiimDk4jA/xvyES2GNt5fRuTPyjC3SaIvUsSIbxoJdxhieMDYF"
            "xkT+eMpJLktadCQ88O4M1cuwzlR1A93TTCLZ9Iy2pHme3G1KIVHGIBwo7R5UZ89GZ/7CIZsfuyRPVm5y2y4KZhaQ"
            "+itoad+E4L19UhGZw72Gfuj7fc8AcJtEsdJIA2OiKsKYbZA5IwgC2CcbiTi0NiexNpgUqGZ6jbKVFhUo7RDsc+xP"
            "Gnh7qLaktA3ob2itDg+DqMHzLZIo10SuKpSDt7kotRTM6WHxu9BzaKMSbnlrYZtqt1o1DRosSsQhM8+a7lJgYiKv"
            "JX39ERoD633ogjzrSMBAkRSTtTmRlT4wkkLwK/qNXJf4S600BUa78/8RwY8CgLyC569QQ+tDRVKCdA1peidn9iRS"
            "RqsllVLIuxJD7bybM1oURIIDCrW4hPKiUuxtnm8JwjDH38lvrchfoAxXeLSGsnycCa0Fv+1q+O1+7RXzTssXXiNY"
            "ucWDEPqo6t9M/PH1qInUoB0yGofxx/ZuPUei2A9H4TlkHEb+LDiHfAzHaTo+h1zfhPFkfg6ZR9Hk5qxNF7V33B1P"
            "zIvhD+lW5or0eGMxRzyTFPWW5k3hGY1MPs5o6fCMQGMkp8iqzhw4GDSA4oixFA7LATbRPMFUVQtS2DVbIrnpeFsN"
            "eVYK/ezLkcv0RyJ/k6KuGnQvUdWUvlMZhmFrSUt9T7mTqzpbOasSWvkJVJf4607aPHXp2ScaSsm2kHtkS9LqajlY"
            "P7Qly+TKlBtZoqpqqjbbDKd9RjdbPTSFpuELw9PTfmSboMUCiwUNZj9QbnYG2u2ikwVOdqI3crJRJwudLOxkkZNF"
            "nSx2stjIYBoQCUPlES6QWxp5IRgTe4JvO/yFqEmC2qKKLJqZA+UlGkE7hFRvl5AnmGgEUw0v+opijp7MgAtiY95q"
            "M3QQtX6mazCjXD1nMO+LtmV4z4xtif8nFjMLcwrluDrwrBthvzSBM6qg3VQw7bSQDvvVYsMwwSK/M/M7bMdnPE79"
            "dNbcsmFkp6S2HQnO/YEUM6QIbjFnGjWm30ejxThOZ/4gTeeLQbiIh4PZOA4G0U0U+YtRnAbzyT/tJXX/3Fz9CwAA"
            "//8DAFBLAwQUAAYACAAAACEACPFJsOAQAADvqQAADwAAAHdvcmQvc3R5bGVzLnhtbOxdyXLcRhK9T8T8A4Inz4Hm"
            "0s1FCtMOkpIshSVZFikrfKwGqtkw0ageLKKon/Fxzr77Zs1/TW3oRjNRALKQbNOKCUWIjSVfLS9fVmVh++a7j/Mk"
            "+MCzPBbpydbe17tbAU9DEcXp1cnWu8tn28dbQV6wNGKJSPnJ1i3Pt7779p//+ObmcV7cJjwPJECaP56HJ1uzolg8"
            "3tnJwxmfs/xrseCpPDgV2ZwVcjO72pmz7LpcbIdivmBFPImTuLjd2d/dPdyyMFkfFDGdxiF/IsJyztNC2+9kPJGI"
            "Is1n8SKv0G76oN2ILFpkIuR5Lhs9TwzenMXpEmZvDIDmcZiJXEyLr2VjbI00lDTf29W/5skK4AAHsA8ADkP+EYdx"
            "bDF2pGUdJ45wOIdLnDiq4fhVpgYQlSiI/VFVD/VHmdew8qiIZji4iqMdZcsKNmP5bB1xmuAQxzVE42CJCK/rmBzX"
            "aQdLwNu54nAePn5xlYqMTRKJJL0ykI4VaGD1v+RH/dE/+Ue9X3WL/TFN1A/Za99K6UYifMKnrEyKXG1mbzK7abf0"
            "n2ciLfLg5jHLwzi+lPWVhc5jWf7z0zSPt+QRzvLiNI9Z48GZ+tF4JMyL2u6zOIq3dlSJ1zxL5eEPTHb8vtmVf1ru"
            "WO45V5Va25ew9KraV2Tbl2/rlTvZ4un2uwu1ayKLOtli2fbFqTbcGz9O4itWlJmMY2pLI5hwl0Xnsv38Y1GyRJ28"
            "YzvG/K111+Lulq7lgoWxrhSbFlxGtb3DXVWDJFZBdP/gUbXxtlRcsrIQthANYP4uYXcAYzLYydB3YSKwPMqnL6Wv"
            "8eiikAdOtnRZcue7F2+yWGQyyp5sPdJlyp0XfB4/j6OIp7UT01kc8fcznr7LebTa/9Mz7ch2RyjKVP4eHR1qL0ry"
            "6OnHkC9U3JVHU6Y4fa0MEnV2Ga8K1+b/rsD2LG1N9jPO1OAT7N2F0NVHQewri7zW2mbM8k7b9VmogkabKmi8qYIO"
            "NlXQ4aYKOtpUQcebKkjD3GdBcRrJcUSfD4sBqF04DjWicRxiQ+M4tITGcUgFjeNQAhrH4ehoHIcfo3EcborAKUTo"
            "8sKas48c3t6O2z1G+OF2Dwl+uN0jgB9ud8D3w+2O73643eHcD7c7evvhdgdrPK6ZagUvpMzSYrDKpkIUqSh4oCa9"
            "g9FYKrF0Rk6DpwY9npE0kgDGRDY7EA9GC5ne7vYQLVL/8bxQiWMgpsE0vlIpz+CK8/QDT8SCByyKJB4hYMZlUubo"
            "ER+fzviUZzwNOaVj04GqTDBIy/mEwDcX7IoMi6cRcfdViCRBYenQMn+eKZHEBE49Z2EmhldNMLL48DLOh/eVAgnO"
            "yiThRFivaVxMYw3PDTTM8NRAwwzPDDTM8MSgxhlVF1k0op6yaEQdZtGI+s34J1W/WTSifrNoRP1m0Yb322VcJDrE"
            "12cde/3X7s4Toa6hDK7HRXyV6lXZwUh2zTR4wzJ2lbHFLFCr2s2w9TZjyzkT0W1wSTGmLZGo5vXaRdRadpyWwzt0"
            "DY1KXEs8Inkt8YgEtsQbLrFXcpqsJmjPafKZi3JSNIpWI/US7QVLSjOhHa42Vgz3sJUAnsVZTiaDZlgCD36tprOK"
            "TorIt6rl8IqtsIbL6m5UIq2ehSSopbrgShOGn98ueCbTsuvBSM9EkogbHtEhXhSZML5Wl/y+pqSX5J/OFzOWxzpX"
            "WoPoP9RXd18Er9hicIPeJCxOaXh7uj1ncRLQzSCeX756GVyKhUozVcfQAJ6JohBzMky7EvjVez75F00FT2USnN4S"
            "tfaUaHlIg53HBIOMQRIREZKcZsZpTDKGarwf+O1EsCyiQXuTcXM/SsGJEC/YfGEmHQTaknHxRsYfgtmQxvuZZbFa"
            "F6IS1SUJWG3ZMC8nv/JweKh7LQKSlaEfy0KvP+qprramgxs+TViDGz5F0GzK4UH5L0Fj1+CGN3YNjqqx5wnL89h5"
            "CdUbj6q5FR51e4cnfxZPJCKblgldB1aAZD1YAZJ1oUjKeZpTtljjETZY41G3l9BlNB7BkpzG+z6LIzIyNBgVExqM"
            "igYNRsWBBiMlYPgdOjWw4bfp1MCG36tjwIimADUwKj8jHf6JrvLUwKj8TINR+ZkGo/IzDUblZ6MnAZ9O5SSYboip"
            "QVL5XA2SbqBJCz5fiIxlt0SQTxN+xQgWSA3am0xM1ZMwIjU3cRNAqjXqhHCybeCoSH7PJ2RVU1iU9SJYEWVJIgTR"
            "2tpqwNGW6/eudZnpJ0EGV+FNwkI+E0nEM0eb3LYyX74wj2Xcrb6uRq9lz5fx1awILmbL1f46zOFup2WVsK+ZdRfY"
            "1OeH1cMvTWaveBSX86qi8GGKw1F/Y+3Ra8bjbuPVTGLN8qCnJSzzsNtyNUteszzqaQnLPO5pqXW6Ztmmhycsu250"
            "hKM2/1nmeA7nO2rzoqVxY7FtjrS0bHLBozYvWpNKcBqG6moBZKefZtz2/cTjtseoyI2CkZMbpbeu3BBtAnvLP8Rq"
            "ZMcETV3e8u4JEPf1JLpX5PypFGbdfu2CU/+Hul7IiVOa86ARZ9T/wtValHH3Y+9w44boHXfcEL0DkBuiVyRymqNC"
            "khuld2xyQ/QOUm4IdLSCIwIuWkF7XLSC9j7RCqL4RKsBswA3RO/pgBsCLVQIgRbqgJmCGwIlVGDuJVSIghYqhEAL"
            "FUKghQonYDihQnucUKG9j1Ahio9QIQpaqBACLVQIgRYqhEALFUKgheo5t3eaewkVoqCFCiHQQoUQaKHq+eIAoUJ7"
            "nFChvY9QIYqPUCEKWqgQAi1UCIEWKoRACxVCoIUKIVBCBeZeQoUoaKFCCLRQIQRaqOZRQ3+hQnucUKG9j1Ahio9Q"
            "IQpaqBACLVQIgRYqhEALFUKghQohUEIF5l5ChShooUIItFAhBFqo+mLhAKFCe5xQob2PUCGKj1AhClqoEAItVAiB"
            "FiqEQAsVQqCFCiFQQgXmXkKFKGihQgi0UCFEm3/aS5Su2+z38Kuezjv2+1+6spV6W3+Uuw416g9V1cqN1f9ZhDMh"
            "roPGBw9HOt/oBxJPkljoJWrHZfU6rr4lAnXh88fz9id86ugDX7pkn4XQ10wB+LivJVhTGbe5fN0SJHnjNk+vW4JZ"
            "57gt+tYtwTA4bgu6WpfVTSlyOALGbWGmZrznMG+L1jVz2MVtMbpmCHu4LTLXDGEHt8XjmuFBoILzXeuDnv10uLy/"
            "FCC0uWMN4ciN0OaWkKsqHENh9CXNjdCXPTdCXxrdCCg+nTB4Yt1QaIbdUH5UQ5lhqfYXqhsBSzVE8KIawPhTDaG8"
            "qYZQflTDwIilGiJgqfYPzm4EL6oBjD/VEMqbagjlRzUcyrBUQwQs1RABS/XAAdkJ4081hPKmGkL5UQ0nd1iqIQKW"
            "aoiApRoieFENYPyphlDeVEMoP6pBloymGiJgqYYIWKohghfVAMafagjlTTWEaqNar6KsUY1iuGaOm4TVDHEDcs0Q"
            "F5xrhh7ZUs3aM1uqIXhmS5CrinNctlQnzY3Qlz03Ql8a3QgoPp0weGLdUGiG3VB+VOOypSaq/YXqRsBSjcuWnFTj"
            "sqVWqnHZUivVuGzJTTUuW2qiGpctNVHtH5zdCF5U47KlVqpx2VIr1bhsyU01LltqohqXLTVRjcuWmqgeOCA7Yfyp"
            "xmVLrVTjsiU31bhsqYlqXLbURDUuW2qiGpctOanGZUutVOOypVaqcdmSm2pcttRENS5baqIaly01UY3LlpxU47Kl"
            "Vqpx2VIr1bhs6ZU0iQleAXUxZ1kR0L0v7jnLZwUb/nLCd2nGc5F84FFA29SXqFbu3Kx9/kph608RyvML2WfqDei1"
            "x5Ui8wZYC6hPfBEtP1OljFVNAvv1MLtbV9herjUlasOOopbgZyy53gPYqy9bafgJk036UXUFKDlVb0Vs2K+8odqv"
            "yzifscwcWjlpdYKV4aoVN4+zPI6qw7u7T493j07txV372bJrzhevZeF6n9qQzPBcb62+aDZRbxOTbR+ZT5rZD5wd"
            "W70K876mlx+SZUmWNFtG6+fl2K8tn5dTB5/afer42hfm1ixXX5hTu8+WX5gLlb6reu0/Oxg/0grXJ2vtn2wxrXzt"
            "LHq3uh1FAp09MwirD9JVl5nrH6Qz+2qfikO7zb7TbWzkIXCb/R5us5KiOW9NiPfsWPZbeZ2OVUWDL8yxRpbpumOZ"
            "fUMca+R0LMsUgWONvhDHqjrb4Vhd7rMJJ9m387S1z2HqfUOcZOx0EnsrD4GTjB+4kxzXfaQK8tBHdAn0PhKb/89N"
            "7YZ6zBBfOHD6gr05i8AXDr4MX9DieHjxYgj75guvTezbJJWA/cMHzv64zr6TfK2GjQaCg0fq311XUF9UWjnCZay+"
            "1HuqyRriB0dOP7BLDQR+cPRF+EHV1fcZBDbJ/LGTeTv1IGD++IEy38W19v3NDv5H6l8f5p8MngU+cjJvKSFg/tHf"
            "lPmqc+91qCfnOpSdzUL7inXHitjPLMvZbcJS+7af6S/s02W8iIEzOL6p5ODSrnZ1cemue6HWZlvqbbxMreAKUNW1"
            "V7i7nK23txWTxNAtf7xIlbPd2C/bm5pGH5mBksfPeZK8YuZssXCfmvCp0os8urer36555/jEfCjCaZ/pKwpOgJ31"
            "ypjNdl8xn460j7o4+lwtmvNfxHVDh+snr4b2dU9PDstcds6FOuFuDVfLo3ereMb++1vy+ffrYC9YxaA7Ec2phcZY"
            "Zl3cGcfcken/a6E4Rs3KpZPRfSpG7Qpc35Hpyyd4yJokhmCzgugkeERFsKWHnuC/KtevkzVkbRBDllnJc5I1piLL"
            "rj0+HLI2vUiHIcUsqTlJOaAixS4CfjkKoqXBrG05aTikosGuxv0ttEG/eoFhxKwyORk5omLEros9UGH8tRyY9R4n"
            "B8dUHNgB8O8xYtxvtt/BiFmHcTLyiIoR2/EPdbjY0OraDyItzxpuOzLvv2jqZezami1hxVrDioxN1lCLZ2CFzFwa"
            "U6tjsvvMCrnaeFsqP2NlIapuTlU3liyx7943vfcAbtlYtUi3ervqlmueLXt/NZOu9hzYAbc+tzb76HRZ5/Cup6hj"
            "gZXn598+/06hz8orO/zlYWa5m6exWdqnSXHLPgG+lh/ophC3KaNV23YtCqXttJybH3ECb66yB+95yRs7LQG879l8"
            "ZLNZcI2Qu8SbQzQCtb7VwfkDn0neM2UuVaYFoMZ8QoBGkGnRJsd9O7vxHGrr963pM34NK0uVyHJdLlBmy1xzvKv+"
            "9eGMOide9hTUyeff04JKJ+7rUBUVnSrZaPc1e62+vFK1FvSYfmJh9UmOLj+GXTGyC2oor4z1BS91uUq9P8/6Y9vs"
            "rqfPLBv9ffJzmV2VCZsz0GTwGjy8nzR4BGqs7PaODd6h9X3SGNjWv5HS5Rh9ApwpqC3EjXyyicVZpP+aS6P6vFx6"
            "j/3O9id1R536IX1KBRKtON3Vnkvjy4uo91yScn3bsq7HJdSWcaWaro4PdW30NV2zpU8ZGvX/0sXQmgfdddfv//zj"
            "z/8kf/4RUI4BVhod3vrgdO6KiWfsQ5mVTqWv3sBJERGrq3yoiDgxJdueymUYSc7ZgqbfwJyxutnSK2qaIPeeT0B3"
            "2vtEvpLH/tUZN4fcQuIz+dvVccNsnpaFsKdYodl4Ys/SW/CkjmWajgUZtT4153nwmt8Eb8Wc6cf+qhyi8aBO4xuP"
            "hDncrdtfz9eXdxMsXcdGwrXEoHq7KNPdtapPkW1fvtXH9saysVesKDNZntrShqlItVz8RQn856LIRHpFosH9oRqE"
            "zap+5d/+DwAA//8DAFBLAwQUAAYACAAAACEAJt76SG8BAAAtBAAAFAAAAHdvcmQvd2ViU2V0dGluZ3MueG1snNPd"
            "bsIgFADg+yV7h4Z7pTo1S2M1WRaX3SxLtj0AwqklAqcBXHVPP6jV1Xhjd1MO0PPl8Ddf7rVKvsE6iSYno2FKEjAc"
            "hTSbnHx9rgaPJHGeGcEUGsjJARxZLu7v5nVWw/oDvA9/uiQoxmWa56T0vsoodbwEzdwQKzBhskCrmQ9du6Ga2e2u"
            "GnDUFfNyLZX0BzpO0xlpGXuLgkUhOTwj32kwvsmnFlQQ0bhSVu6k1bdoNVpRWeTgXFiPVkdPM2nOzGhyBWnJLTos"
            "/DAspq2ooUL6KG0irf6AaT9gfAXMOOz7GY+tQUNm15GinzM7O1J0nP8V0wHErhcxfjjVEZuY3rGc8KLsx53OiMZc"
            "5lnJXHkpFqqfOOmIxwumkG+7JvTbtOkZPOh4hppnrxuDlq1VkMKtTMLFSho4fsP5xKYJYd+Mx21pg0LFIOzaIrxf"
            "rLzU8gdWaJ8s1g4sjcNMKazf315Ch1488sUvAAAA//8DAFBLAwQUAAYACAAAACEAyBhMNjkCAAAICAAAEgAAAHdv"
            "cmQvZm9udFRhYmxlLnhtbOSTTY+bMBCG75X6H5DvGwwJSTbaZLVtN1KlqofV9gc4xgRr/YFsJyT/vmPzUdoo2tBD"
            "L+UA5h3Pw8zr4eHxJEV0ZMZyrdYomWAUMUV1ztV+jX68bu+WKLKOqJwIrdganZlFj5uPHx7qVaGVsxHkK7uSdI1K"
            "56pVHFtaMknsRFdMQbDQRhIHr2YfS2LeDtUd1bIiju+44O4cpxjPUYsxt1B0UXDKvmh6kEy5kB8bJoColS15ZTta"
            "fQut1iavjKbMWuhZioYnCVc9JpldgCSnRltduAk001YUUJCe4LCS4hcgGwdILwBzyk7jGMuWEUPmkMPzcZx5z+H5"
            "gPN3xQwA+WEUIp12dfiHTx+wbO7ychyuO6PY5xJHSmLL34mFGEecDYjNgAlN34ZMNs60rAeepT9DSVdf90obshNA"
            "gqmMYLCiAPZ3OB//CEt2Crq3pV0Uwi/AtU3750b1ShEJoM9E8J3hIVARpS1LIHYk0D7YtMUZ9naleIan/o5iv5GW"
            "xFjmIWHjU9rIBZFcnDvV1tzaJlBxR8tOPxLDfRNNyPI9BA52h9foeYZx+rzdokZJoDr4HdPZ4lOrpFBUc923yrRX"
            "sFdo4ITXpOHQwOn3wDfjxoELJ165ZDb6zuroRUuirjiS4jk4kYEf3pnpFUeaL/3piAncUY7gC0dAWSyzf+JIOxvR"
            "N74v3dUJ8XPxf0zIi95pp4MRN503OcDusceNn4btZSCkyaJT+vbS5P327t9rr13YzU8AAAD//wMAUEsDBBQABgAI"
            "AAAAIQBXayfZgQEAAPkCAAARAAgBZG9jUHJvcHMvY29yZS54bWwgogQBKKAAAQAAAAAAAAAAAAAAAAAAAAAAAAAA"
            "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA"
            "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA"
            "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA"
            "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAACMkktuwjAQhveVeofI++A8KlRFIagP0Q1IVUsf"
            "YufaA7gktmUbAnfpIXqH9mB1EgiNyqK7Gc83/4x/Ox1ui9zbgDZcigEKewHyQFDJuFgM0NN05F8iz1giGMmlgAHa"
            "gUHD7PwspSqhUsO9lgq05WA8pyRMQtUALa1VCcaGLqEgpucI4YpzqQtiXaoXWBG6IgvAURD0cQGWMGIJrgR91Sqi"
            "vSSjraRa67wWYBRDDgUIa3DYC/GRtaALc7KhrvwiC253Ck6ih2JLbw1vwbIse2Vco27/EL9Oxo/1VX0uKq8ooCxl"
            "NLHc5pCl+Bi6yKzf3oHa5rhNXEw1ECt19izzFRHe7OtjPPr+vLt6qMlDtfJ9BbtSamacRidzGANDNVfWvWYzoXPg"
            "6JwYO3HPO+fArnenhv2Fqj4NG179kSyqiTZN94Y3CwLznFFJY+uh8hLf3E5HKIuCqO8HsR+H0yhKLqIkCGbVjp3+"
            "o2CxX+CfinESBl3Fg0BjU/ezZj8AAAD//wMAUEsDBBQABgAIAAAAIQBNWThqdgEAAMgCAAAQAAgBZG9jUHJvcHMv"
            "YXBwLnhtbCCiBAEooAABAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA"
            "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA"
            "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA"
            "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA"
            "AAAAAAAAAJxSy27CMBC8V+o/RLmDAwdE0WJUgaoe+kAihbPlbBKrjm3ZBsHfd0NKmqq35rQ7ux7PTAyrc6OTE/qg"
            "rFmmk3GWJmikLZSplulH/jSap0mIwhRCW4PL9IIhXfH7O9h669BHhSEhChOWaR2jWzAWZI2NCGMaG5qU1jciUusr"
            "ZstSSdxYeWzQRDbNshnDc0RTYDFyPWHaMS5O8b+khZWtvrDPL474OOTYOC0i8rf2pAbWA5DbKHSuGuQzgvsGtqLC"
            "wCfAugIO1hfUzx+AdSWsa+GFjBQen2Rz2hwA8OicVlJEypW/KultsGVM3q9ik5YA2HAFyMAO5dGreOEZsGELL8qQ"
            "Arq4K0iaF5UXrg582urrO9hJoXFN1nkpdEBgPwCsbeOEITrWV8T3GT5cbjdtFN9HfoMDlwcV650Tso1lOpsP/Q5G"
            "sCMUCzLQa+gBeKb/4XV7AZ01FRa3nb+DNsF99zD5ZDbO6LtGdsPIeP9i+BcAAAD//wMAUEsBAi0AFAAGAAgAAAAh"
            "AN+k0mxaAQAAIAUAABMAAAAAAAAAAAAAAAAAAAAAAFtDb250ZW50X1R5cGVzXS54bWxQSwECLQAUAAYACAAAACEA"
            "HpEat+8AAABOAgAACwAAAAAAAAAAAAAAAACTAwAAX3JlbHMvLnJlbHNQSwECLQAUAAYACAAAACEAKCrLOd4HAAAg"
            "JAAAEQAAAAAAAAAAAAAAAACzBgAAd29yZC9kb2N1bWVudC54bWxQSwECLQAUAAYACAAAACEA1mSzUfQAAAAxAwAA"
            "HAAAAAAAAAAAAAAAAADADgAAd29yZC9fcmVscy9kb2N1bWVudC54bWwucmVsc1BLAQItABQABgAIAAAAIQCbj6oe"
            "1AYAANEgAAAVAAAAAAAAAAAAAAAAAPYQAAB3b3JkL3RoZW1lL3RoZW1lMS54bWxQSwECLQAUAAYACAAAACEA0G4x"
            "FHYEAAAiDQAAEQAAAAAAAAAAAAAAAAD9FwAAd29yZC9zZXR0aW5ncy54bWxQSwECLQAUAAYACAAAACEACPFJsOAQ"
            "AADvqQAADwAAAAAAAAAAAAAAAACiHAAAd29yZC9zdHlsZXMueG1sUEsBAi0AFAAGAAgAAAAhACbe+khvAQAALQQA"
            "ABQAAAAAAAAAAAAAAAAAry0AAHdvcmQvd2ViU2V0dGluZ3MueG1sUEsBAi0AFAAGAAgAAAAhAMgYTDY5AgAACAgA"
            "ABIAAAAAAAAAAAAAAAAAUC8AAHdvcmQvZm9udFRhYmxlLnhtbFBLAQItABQABgAIAAAAIQBXayfZgQEAAPkCAAAR"
            "AAAAAAAAAAAAAAAAALkxAABkb2NQcm9wcy9jb3JlLnhtbFBLAQItABQABgAIAAAAIQBNWThqdgEAAMgCAAAQAAAA"
            "AAAAAAAAAAAAAHE0AABkb2NQcm9wcy9hcHAueG1sUEsFBgAAAAALAAsAwQIAAB03AAAAAA=="
        )

    def _uyari_sablon_olustur(self, hedef_yolu):
        """Gömülü şablonu DATA_DIR'e yaz."""
        import base64, shutil
        # Önce uygulama yanındaki dosyayı dene
        for aday in [
            Path(__file__).parent / "APARTMAN_YÖNETİMİNDEN_BİLGİLENDİRME_VE_TALEP_YAZISI.docx",
            Path.home() / "APARTMAN_YÖNETİMİNDEN_BİLGİLENDİRME_VE_TALEP_YAZISI.docx",
        ]:
            if aday.exists():
                shutil.copy(str(aday), str(hedef_yolu))
                return
        # Gömülü şablonu kullan
        hedef_yolu.write_bytes(base64.b64decode(self._uyari_gomulu_b64()))

    def _uyari_secili_yazdir(self):
        sel = self._uw_tree.selection()
        if not sel:
            messagebox.showinfo("Seçim", "Lütfen listeden bir daire seçin.", parent=self)
            return
        no = int(sel[0])
        d  = next((x for x in self.data["daireler"] if x["no"] == no), {})
        b  = self.data["bina"]
        ba = buay()
        tum = ay_listesi(b.get("baslangic", ba), ba)
        _bugun = datetime.date.today()
        _son_gun = b.get("son_gun", 10)
        odenmemis = [
            ay for ay in tum
            if not any(o["daireNo"] == no and o["ay"] == ay
                       for o in self.data["odemeler"])
            # Bu ay ödeme günü henüz geçmemişse hariç tut
            and not (ay == buay() and _bugun.day < _son_gun)
        ]
        anapara = round(sum(ay_aidat(self.data, ay, no) for ay in odenmemis), 2)
        faiz    = round(sum(faiz_hesapla(ay_aidat(self.data, ay, no),
                            b["faiz"], ay, b["son_gun"])
                            for ay in odenmemis), 2)
        toplam  = round(anapara + faiz, 2)
        self._uyari_kaydet([{"no": no, "ana": toplam}])

    def _uyari_tumu_yazdir(self):
        b  = self.data["bina"]
        ba = buay()
        tum_aylar = ay_listesi(b.get("baslangic", ba), ba)
        _bugun   = datetime.date.today()
        _son_gun = b.get("son_gun", 10)
        borclu = []
        for d in self.data["daireler"]:
            no = d["no"]
            odenmemis = [
                ay for ay in tum_aylar
                if not any(o["daireNo"] == no and o["ay"] == ay
                           for o in self.data["odemeler"])
                and not (ay == buay() and _bugun.day < _son_gun)
            ]
            anapara = round(sum(ay_aidat(self.data, ay, no) for ay in odenmemis), 2)
            faiz    = round(sum(faiz_hesapla(ay_aidat(self.data, ay, no),
                                b["faiz"], ay, b["son_gun"])
                                for ay in odenmemis), 2)
            toplam  = round(anapara + faiz, 2)
            if toplam > 0:
                borclu.append({"no": no, "ana": toplam})
        if not borclu:
            messagebox.showinfo("Bilgi", "Borçlu daire bulunamadı.", parent=self)
            return
        self._uyari_kaydet(borclu)

    def _uyari_secili_word(self):
        """Raporlar sekmesi: seçili daireye Word uyarı yazısı kaydet."""
        import tkinter.filedialog as fd
        sel = self._uw_tree.selection()
        if not sel:
            messagebox.showinfo("Seçim", "Lütfen listeden bir daire seçin.", parent=self)
            return
        no = int(sel[0])
        d  = next((x for x in self.data["daireler"] if x["no"] == no), {})
        b  = self.data["bina"]
        ba = buay()
        tum = ay_listesi(b.get("baslangic", ba), ba)
        _bugun   = datetime.date.today()
        _son_gun = b.get("son_gun", 10)
        odenmemis = [
            ay for ay in tum
            if not any(o["daireNo"] == no and o["ay"] == ay
                       for o in self.data["odemeler"])
            and not (ay == buay() and _bugun.day < _son_gun)
        ]
        anapara = round(sum(ay_aidat(self.data, ay, no) for ay in odenmemis), 2)
        faiz    = round(sum(faiz_hesapla(ay_aidat(self.data, ay, no),
                            b["faiz"], ay, b["son_gun"])
                            for ay in odenmemis), 2)
        toplam  = round(anapara + faiz, 2)
        dosya = fd.asksaveasfilename(
            parent=self,
            defaultextension=".docx",
            filetypes=[("Word Belgesi","*.docx"),("Tüm Dosyalar","*.*")],
            initialfile=f"uyari_daire_{no:03d}.docx",
            title="Uyarı Yazısını Word Olarak Kaydet")
        if not dosya: return
        try:
            ana_str = fmt(toplam).replace("₺","").strip()
            doc_bytes = self._uyari_sablon_doc(no, ana_str)
            with open(dosya, "wb") as f:
                f.write(doc_bytes)
            messagebox.showinfo("Kaydedildi",
                f"Word uyarı yazısı kaydedildi:\n{dosya}", parent=self)
        except Exception as e:
            messagebox.showerror("Hata", f"Belge oluşturulamadı:\n{e}", parent=self)

    def _uyari_tumu_pdf(self):
        """Tüm borçlu dairelere toplu PDF uyarı yazısı üret ve klasöre kaydet."""
        import tkinter.filedialog as fd
        b  = self.data["bina"]
        ba = buay()
        tum_aylar = ay_listesi(b.get("baslangic", ba), ba)
        _bugun   = datetime.date.today()
        _son_gun = b.get("son_gun", 10)
        borclu = []
        for d in self.data["daireler"]:
            no = d["no"]
            odenmemis = [
                ay for ay in tum_aylar
                if not any(o["daireNo"] == no and o["ay"] == ay
                           for o in self.data["odemeler"])
                and not (ay == buay() and _bugun.day < _son_gun)
            ]
            anapara = round(sum(ay_aidat(self.data, ay, no) for ay in odenmemis), 2)
            faiz    = round(sum(faiz_hesapla(ay_aidat(self.data, ay, no),
                                b["faiz"], ay, b["son_gun"])
                                for ay in odenmemis), 2)
            toplam  = round(anapara + faiz, 2)
            if toplam > 0:
                borclu.append({"no": no, "ana": toplam})
        if not borclu:
            messagebox.showinfo("Bilgi", "Borçlu daire bulunamadı.", parent=self)
            return
        klasor = fd.askdirectory(parent=self,
            title="PDF Uyarı Yazılarını Kaydedeceğiniz Klasörü Seçin")
        if not klasor: return
        hata = 0
        for item in borclu:
            no  = item["no"]
            ana = item["ana"]
            try:
                ana_str   = fmt(ana).replace("₺","").strip()
                pdf_bytes = self._uyari_pdf_olustur(no, ana_str)
                dosya = Path(klasor) / f"uyari_daire_{no:03d}.pdf"
                with open(dosya, "wb") as f:
                    f.write(pdf_bytes)
            except Exception:
                hata += 1
        basarili = len(borclu) - hata
        messagebox.showinfo("Tamamlandı",
            f"{basarili} PDF uyarı yazısı {klasor} klasörüne kaydedildi."
            + (f"\n{hata} belgede hata oluştu." if hata else ""),
            parent=self)

    def _uyari_kaydet(self, daire_listesi):
        import tkinter.filedialog as fd
        if len(daire_listesi) == 1:
            no  = daire_listesi[0]["no"]
            ana = daire_listesi[0]["ana"]
            dosya = fd.asksaveasfilename(
                parent=self,
                defaultextension=".pdf",
                filetypes=[
                    ("PDF Belgesi","*.pdf"),
                    ("Word Belgesi","*.docx"),
                    ("Tüm Dosyalar","*.*")],
                initialfile=f"uyari_daire_{no:03d}.pdf",
                title="Uyarı Yazısını Kaydet")
            if not dosya: return
            try:
                ana_str = fmt(ana).replace("₺","").strip()
                if dosya.lower().endswith(".docx"):
                    out_bytes = self._uyari_sablon_doc(no, ana_str)
                else:
                    out_bytes = self._uyari_pdf_olustur(no, ana_str)
                with open(dosya, "wb") as f:
                    f.write(out_bytes)
                messagebox.showinfo("Kaydedildi",
                    f"Uyarı yazısı kaydedildi:\n{dosya}", parent=self)
            except Exception as e:
                messagebox.showerror("Hata", f"Belge oluşturulamadı:\n{e}", parent=self)
        else:
            klasor = fd.askdirectory(parent=self,
                title="Uyarı Yazılarını Kaydedeceğiniz Klasörü Seçin")
            if not klasor: return
            hata_sayisi = 0
            for item in daire_listesi:
                no  = item["no"]
                ana = item["ana"]
                ana_str = fmt(ana).replace("₺","").strip()
                try:
                    pdf_bytes = self._uyari_pdf_olustur(no, ana_str)
                    dosya = Path(klasor) / f"uyari_daire_{no:03d}.pdf"
                    with open(dosya, "wb") as f:
                        f.write(pdf_bytes)
                except Exception:
                    hata_sayisi += 1
            basarili = len(daire_listesi) - hata_sayisi
            messagebox.showinfo("Tamamlandı",
                f"{basarili} PDF uyarı yazısı {klasor} klasörüne kaydedildi."
                + (f"\n{hata_sayisi} belgede hata oluştu." if hata_sayisi else ""),
                parent=self)

    # ── Daireler sekmesi Uyarı Yazısı metodları ───────────────────────────

    def _daire_sec_uyari_bilgi(self):
        """Seçili dairenin no ve toplam borcunu (anapara + faiz) anlık hesaplayarak döndür."""
        sel = self._tv.selection()
        if not sel:
            messagebox.showinfo("Seçim", "Lütfen listeden bir daire seçin.", parent=self)
            return None, None
        no = int(sel[0])
        d  = next((x for x in self.data["daireler"] if x["no"] == no), {})
        b  = self.data["bina"]
        ba = buay()
        tum = ay_listesi(b.get("baslangic", ba), ba)
        _bugun = datetime.date.today()
        _son_gun = b.get("son_gun", 10)
        odenmemis = [
            ay for ay in tum
            if not any(o["daireNo"] == no and o["ay"] == ay
                       for o in self.data["odemeler"])
            # Bu ay ödeme günü henüz geçmemişse hariç tut
            and not (ay == buay() and _bugun.day < _son_gun)
        ]
        anapara = round(sum(ay_aidat(self.data, ay, no) for ay in odenmemis), 2)
        faiz    = round(sum(faiz_hesapla(ay_aidat(self.data, ay, no),
                            b["faiz"], ay, b["son_gun"])
                            for ay in odenmemis), 2)
        toplam  = round(anapara + faiz, 2)
        if toplam <= 0:
            messagebox.showinfo("Borç Yok",
                f"Daire {no} için ödenmemiş borç bulunamadı.", parent=self)
            return None, None
        return no, toplam

    def _daire_uyari_yazisi_docx(self):
        """Daireler sekmesinden seçili daire için Word uyarı yazısı kaydet."""
        import tkinter.filedialog as fd
        no, ana = self._daire_sec_uyari_bilgi()
        if no is None: return
        dosya = fd.asksaveasfilename(
            parent=self,
            defaultextension=".docx",
            filetypes=[("Word Belgesi","*.docx"),("Tüm Dosyalar","*.*")],
            initialfile=f"uyari_daire_{no:03d}.docx",
            title="Uyarı Yazısını Kaydet")
        if not dosya: return
        try:
            doc_bytes = self._uyari_sablon_doc(no, fmt(ana).replace("₺","").strip())
            with open(dosya, "wb") as f:
                f.write(doc_bytes)
            messagebox.showinfo("Kaydedildi",
                f"Uyarı yazısı kaydedildi:\n{dosya}", parent=self)
        except Exception as e:
            messagebox.showerror("Hata", f"Belge oluşturulamadı:\n{e}", parent=self)

    def _uyari_pdf_olustur(self, daire_no, ana_para_str):
        """
        reportlab + Windows Arial TTF ile Türkçe destekli PDF üretir.
        Font bulunamazsa Helvetica ile devam eder.
        """
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.units import cm
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer
        from reportlab.lib.styles import ParagraphStyle
        from reportlab.lib.enums import TA_CENTER, TA_JUSTIFY, TA_LEFT
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont
        import io, os

        b_veri   = self.data["bina"]
        bina_adi = b_veri.get("adi", "Apartman")
        sablon   = b_veri.get("uyari_sablon", {})

        def _s(key, varsayilan):
            """Şablondan değer al, yoksa varsayılanı kullan."""
            v = sablon.get(key, varsayilan)
            return (v.replace("{bina_adi}", bina_adi)
                     .replace("{daire_no}", str(daire_no))
                     .replace("{toplam_borc}", str(ana_para_str)))

        baslik_metin  = _s("baslik",
            "APARTMAN YÖNETİMİNDEN BİLGİLENDİRME VE TALEP YAZISI")
        konu_metin    = _s("konu",
            "KONU: Aidat Borcu Ödeme Talebi")
        giris_metin   = _s("giris",
            f"{bina_adi} Yönetimi olarak sizlere, apartmanımızın düzenli bakım, "
            "temizlik ve diğer ortak giderlerinin karşılanabilmesi adına belirlenen "
            "aylık aidatların önemini bir kez daha hatırlatmak isteriz. Ortak "
            "alanlarımızın sorunsuz ve temiz bir şekilde kullanılabilmesi için tüm "
            "sakinlerimizin aidat ödemelerini zamanında yapması gerekmektedir.")
        borc_metin    = _s("borc_bilgi",
            f"Yapılan incelemeler sonucunda Daire {daire_no} bağımsız bölüm numaralı "
            f"daireye ait {ana_para_str} TL tutarında ödenmemiş aidat borcu olduğu "
            "tespit edilmiştir. Kat Mülkiyeti Kanunu ve Genel Kurul toplantısında "
            "alınan kararlara göre, aidatların belirlenen tarihlerde ödenmemesi "
            "durumunda yasal takip süreci başlatılabilecektir.")
        talep_metin   = _s("talep",
            "Siz değerli sakinimizden, ödenmemiş aidat borcunuzun bu yazının "
            "tarafınıza ulaştığı tarihten itibaren 7 (Yedi) gün içerisinde banka "
            "hesabımıza yatırılarak kapatılmasını önemle rica ederiz. Ödeme "
            "yapılmaması halinde, yasal süreç başlatılacak olup, bu süreçten "
            "doğacak masraf ve avukatlık ücretlerinin de tarafınıza yansıtılacağı "
            "hususunu bilginize sunarız.")
        kapanis_metin = _s("kapanis",
            "Düzenli ödemeleriniz ve apartmanımıza gösterdiğiniz özen için teşekkür "
            "eder, iyi günler dileriz.")
        imza_metin    = _s("imza", f"{bina_adi} Yönetimi")

        # TTF font ara - Windows Arial öncelikli, Linux fallback
        _font_candidates = [
            os.path.join(os.environ.get("SystemRoot", "C:/Windows"), "Fonts", "arial.ttf"),
            os.path.join(os.environ.get("SystemRoot", "C:/Windows"), "Fonts", "Arial.ttf"),
            "/usr/share/fonts/truetype/liberation/LiberationSans-Regular.ttf",
            "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
        ]
        _bold_candidates = [
            os.path.join(os.environ.get("SystemRoot", "C:/Windows"), "Fonts", "arialbd.ttf"),
            os.path.join(os.environ.get("SystemRoot", "C:/Windows"), "Fonts", "ArialBD.ttf"),
            "/usr/share/fonts/truetype/liberation/LiberationSans-Bold.ttf",
            "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf",
        ]

        font_normal = 'Helvetica'
        font_bold   = 'Helvetica-Bold'

        for fp in _font_candidates:
            if fp and os.path.exists(fp):
                try:
                    pdfmetrics.registerFont(TTFont('TRNormal', fp))
                    font_normal = 'TRNormal'
                except Exception:
                    pass
                break

        for fp in _bold_candidates:
            if fp and os.path.exists(fp):
                try:
                    pdfmetrics.registerFont(TTFont('TRBold', fp))
                    font_bold = 'TRBold'
                except Exception:
                    pass
                break

        buf = io.BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=A4,
            leftMargin=2.5*cm, rightMargin=2.5*cm,
            topMargin=2.5*cm, bottomMargin=2.5*cm)

        baslik_s = ParagraphStyle("baslik", fontSize=13, fontName=font_bold,
            alignment=TA_CENTER, spaceAfter=6)
        konu_s   = ParagraphStyle("konu",   fontSize=11, fontName=font_bold,
            alignment=TA_CENTER, spaceAfter=18)
        normal_s = ParagraphStyle("normal", fontSize=11, fontName=font_normal,
            alignment=TA_JUSTIFY, spaceAfter=10, leading=16)
        imza_s   = ParagraphStyle("imza",   fontSize=11, fontName=font_normal,
            alignment=TA_LEFT, spaceAfter=6)
        imza_b   = ParagraphStyle("imzab",  fontSize=11, fontName=font_bold,
            alignment=TA_LEFT, spaceAfter=6)

        story = [
            Paragraph(baslik_metin, baslik_s),
            Spacer(1, 0.3*cm),
            Paragraph(konu_metin, konu_s),
            Paragraph(f"Sayın {bina_adi} Sakinleri,", normal_s),
            Paragraph(giris_metin, normal_s),
            Paragraph(borc_metin, normal_s),
            Paragraph(talep_metin, normal_s),
            Paragraph(kapanis_metin, normal_s),
            Spacer(1, 1*cm),
            Paragraph("Saygılarımızla,", imza_s),
            Paragraph(imza_metin, imza_b),
        ]

        doc.build(story)
        return buf.getvalue()


    def _daire_uyari_yazisi_pdf(self):
        """Daireler sekmesinden seçili daire için PDF uyarı yazısı oluştur (reportlab)."""
        import tkinter.filedialog as fd
        no, toplam = self._daire_sec_uyari_bilgi()
        if no is None: return
        dosya = fd.asksaveasfilename(
            parent=self,
            defaultextension=".pdf",
            filetypes=[("PDF Belgesi","*.pdf"),("Tüm Dosyalar","*.*")],
            initialfile=f"uyari_daire_{no:03d}.pdf",
            title="Uyarı Yazısını PDF Olarak Kaydet")
        if not dosya: return
        try:
            ana_str = fmt(toplam).replace("₺","").strip()
            pdf_bytes = self._uyari_pdf_olustur(no, ana_str)
            with open(dosya, "wb") as f:
                f.write(pdf_bytes)
            messagebox.showinfo("Kaydedildi",
                f"PDF uyarı yazısı kaydedildi:\n{dosya}", parent=self)
        except Exception as e:
            messagebox.showerror("Hata", f"PDF oluşturulamadı:\n{e}", parent=self)


    def _tab_ayarlar(self):
        outer=frm(self._content)
        outer.pack(fill="both",expand=True,padx=18,pady=14)
        pw=pane(outer,orient="horizontal")
        pw.pack(fill="both",expand=True)

        sol=card_panel(pw,"Ayar  Bina Ayarları","")
        pw.add(sol,weight=1)
        si=frm(sol,bg=T["bg2"]); si.pack(fill="both",expand=True,padx=20,pady=10)
        b=self.data["bina"]
        self._av={}
        for key,txt,val in [
            ("adi",    "Bina Adı",             b["adi"]),
            ("aidat",  "Aylık Aidat (₺)",       str(b["aidat"])),
            ("faiz",   "Geç Ödeme Faizi (%/ay)",str(b["faiz"])),
            ("son_gun","Son Ödeme Günü",         str(b["son_gun"])),
        ]:
            lbl(si,txt,fg=T["text2"],bg=T["bg2"],font=("Segoe UI",9,"bold")).pack(anchor="w",pady=(10,2))
            var=tk.StringVar(value=str(val))
            ent(si,textvariable=var,width=32).pack(fill="x",ipady=7)
            self._av[key]=var

        # Başlangıç ayı
        lbl(si,"Aidat Başlangıç Ayı",fg=T["text2"],bg=T["bg2"],
            font=("Segoe UI",9,"bold")).pack(anchor="w",pady=(12,4))
        self._bas_wgt=AySecici(si,initial=b.get("baslangic",buay()),bg=T["bg2"])
        self._bas_wgt.pack(anchor="w")

        def kaydet():
            b["adi"] = self._av["adi"].get().strip()
            b["baslangic"] = self._bas_wgt.get()
            eski_aidat = b["aidat"]
            try:
                yeni_aidat = float(self._av["aidat"].get())
            except:
                yeni_aidat = eski_aidat
            try: b["faiz"]    = float(self._av["faiz"].get())
            except: pass
            try: b["son_gun"] = int(self._av["son_gun"].get())
            except: pass

            # Aidat değiştiyse gelecek ayları güncelle
            if yeni_aidat != eski_aidat:
                b["aidat"] = yeni_aidat
                degisen = _guncelle_gelecek_aylar(
                    self.data, yeni_aidat, baslangic_ay=buay()
                )
                save_data(self.data)
                mesaj = f"Ayarlar kaydedildi.\n\nAidat {fmt(eski_aidat)} → {fmt(yeni_aidat)} olarak güncellendi."
                if degisen:
                    mesaj += f"\n{bugun_str()} tarihinden itibaren {len(degisen)} aya yeni aidat uygulandı."
                messagebox.showinfo("Kaydedildi", mesaj)
            else:
                b["aidat"] = yeni_aidat
                save_data(self.data)
                messagebox.showinfo("Kaydedildi", "Ayarlar başarıyla kaydedildi.")

            for w in self.winfo_children(): w.destroy()
            self._build()
        btn(si,"Kaydet  Ayarları Kaydet",kaydet,"gold").pack(anchor="w",pady=(16,4),ipadx=16,ipady=4)

        tk.Frame(si,bg=T["border"],height=1).pack(fill="x",pady=(8,8))

        # ── Aidat Periyodu Ayarı ─────────────────────────────────────
        lbl(si,"Takvim  Aidat Periyodu Ayarı",fg=T["text2"],bg=T["bg2"],
            font=("Segoe UI",9,"bold")).pack(anchor="w",pady=(0,4))
        lbl(si,"Belirli ay aralığına özel aidat tutarı uygula",
            fg=T["text3"],bg=T["bg2"],font=("Segoe UI",8)).pack(anchor="w",pady=(0,8))
        btn(si,"Takvim  Aylık / Yıllık Aidat Planla",
            lambda: AidatPeriyotPencere(self,self.data,callback=self.refresh_all),
            "purple").pack(anchor="w",ipadx=12,ipady=4)

        tk.Frame(si,bg=T["border"],height=1).pack(fill="x",pady=(14,10))

        # ── Şifre Değiştirme ──────────────────────────────────────────────
        lbl(si,"Anahtar  Ayarlar Şifresi",fg=T["text2"],bg=T["bg2"],
            font=("Segoe UI",10,"bold")).pack(anchor="w",pady=(0,3))
        lbl(si,"Ayarlar sekmesine giriş şifresini değiştir",
            fg=T["text3"],bg=T["bg2"],font=("Segoe UI",8)).pack(anchor="w",pady=(0,10))

        sf = frm(si,bg=T["bg2"]); sf.pack(anchor="w")
        for col,txt in enumerate(["Mevcut Şifre","Yeni Şifre","Yeni Şifre (Tekrar)"]):
            lbl(sf,txt,fg=T["text3"],bg=T["bg2"],
                font=("Segoe UI",8)).grid(row=0,column=col,sticky="w",
                padx=(0 if col==0 else 10,0),pady=(0,2))

        ev=tk.StringVar(); nv=tk.StringVar(); tv=tk.StringVar()
        for col,var in enumerate([ev,nv,tv]):
            e2=ent(sf,textvariable=var,width=13)
            e2.config(show="●")
            e2.grid(row=1,column=col,padx=(0 if col==0 else 10,0),ipady=6)

        def sifre_degistir():
            kayitli=self.data["bina"].get("ayar_sifre","1234")
            if ev.get()!=kayitli:
                messagebox.showerror("Hata","Mevcut şifre yanlış!",parent=self); return
            if not nv.get():
                messagebox.showerror("Hata","Yeni şifre boş olamaz!",parent=self); return
            if nv.get()!=tv.get():
                messagebox.showerror("Hata","Yeni şifreler eşleşmiyor!",parent=self); return
            self.data["bina"]["ayar_sifre"]=nv.get()
            save_data(self.data)
            ev.set(""); nv.set(""); tv.set("")
            messagebox.showinfo("Başarılı","Şifre başarıyla değiştirildi.",parent=self)

        btn(sf,"Anahtar Değiştir",sifre_degistir,"gold").grid(
            row=1,column=3,padx=(10,0),ipady=4)

        # ── Uyarı Yazısı Şablonu ──────────────────────────────────────────
        uw=card_panel(pw,"Mektup  Uyarı Yazısı Şablonu","Daire No ve Toplam Borç otomatik doldurulur")
        pw.add(uw,weight=1)
        ui=frm(uw,bg=T["bg2"]); ui.pack(fill="both",expand=True,padx=20,pady=10)

        b_uw = self.data["bina"]
        sablon = b_uw.get("uyari_sablon", {})

        self._uw_av = {}
        uw_alanlar = [
            ("baslik",  "Başlık",
             sablon.get("baslik",
                "APARTMAN YÖNETİMİNDEN BİLGİLENDİRME VE TALEP YAZISI"), 2),
            ("konu",    "Konu Satırı",
             sablon.get("konu",
                "KONU: Aidat Borcu Ödeme Talebi"), 2),
            ("giris",   "Giriş Paragrafı",
             sablon.get("giris",
                "{bina_adi} Yönetimi olarak sizlere, apartmanımızın düzenli bakım, "
                "temizlik ve diğer ortak giderlerinin karşılanabilmesi adına belirlenen "
                "aylık aidatların önemini bir kez daha hatırlatmak isteriz. Ortak "
                "alanlarımızın sorunsuz ve temiz bir şekilde kullanılabilmesi için tüm "
                "sakinlerimizin aidat ödemelerini zamanında yapması gerekmektedir."), 4),
            ("borc_bilgi", "Borç Bilgisi Paragrafı",
             sablon.get("borc_bilgi",
                "Yapılan incelemeler sonucunda Daire {daire_no} bağımsız bölüm numaralı "
                "daireye ait {toplam_borc} TL tutarında ödenmemiş aidat borcu olduğu "
                "tespit edilmiştir. Kat Mülkiyeti Kanunu ve Genel Kurul toplantısında "
                "alınan kararlara göre, aidatların belirlenen tarihlerde ödenmemesi "
                "durumunda yasal takip süreci başlatılabilecektir."), 4),
            ("talep",   "Talep Paragrafı",
             sablon.get("talep",
                "Siz değerli sakinimizden, ödenmemiş aidat borcunuzun bu yazının "
                "tarafınıza ulaştığı tarihten itibaren 7 (Yedi) gün içerisinde banka "
                "hesabımıza yatırılarak kapatılmasını önemle rica ederiz. Ödeme "
                "yapılmaması halinde, yasal süreç başlatılacak olup, bu süreçten "
                "doğacak masraf ve avukatlık ücretlerinin de tarafınıza yansıtılacağı "
                "hususunu bilginize sunarız."), 4),
            ("kapanis",  "Kapanış Cümlesi",
             sablon.get("kapanis",
                "Düzenli ödemeleriniz ve apartmanımıza gösterdiğiniz özen için teşekkür "
                "eder, iyi günler dileriz."), 2),
            ("imza",    "İmza / Gönderen",
             sablon.get("imza", "{bina_adi} Yönetimi"), 2),
        ]

        lbl(ui, "Ipucu  {daire_no} ve {toplam_borc} ve {bina_adi} otomatik değiştirilir.",
            fg=T["text3"], bg=T["bg2"], font=("Segoe UI",8)).pack(anchor="w", pady=(0,8))

        uw_scroll_frm = frm(ui, bg=T["bg2"])
        uw_scroll_frm.pack(fill="both", expand=True)

        for key, txt, val, rows in uw_alanlar:
            lbl(uw_scroll_frm, txt, fg=T["text2"], bg=T["bg2"],
                font=("Segoe UI",9,"bold")).pack(anchor="w", pady=(8,2))
            var = tk.StringVar(value=val)
            t_widget = tk.Text(uw_scroll_frm, height=rows, width=40,
                bg=T["bg4"], fg=T["text"], insertbackground=T["text"],
                relief="flat", font=("Segoe UI",9), wrap="word",
                highlightthickness=1, highlightbackground=T["border"])
            t_widget.insert("1.0", val)
            t_widget.pack(fill="x", ipady=2)
            self._uw_av[key] = t_widget

        def uw_kaydet():
            sablon_yeni = {k: w.get("1.0","end-1c").strip()
                           for k, w in self._uw_av.items()}
            self.data["bina"]["uyari_sablon"] = sablon_yeni
            save_data(self.data)
            messagebox.showinfo("Kaydedildi",
                "Uyarı yazısı şablonu kaydedildi.", parent=self)

        def uw_sifirla():
            if not messagebox.askyesno("Sıfırla",
                "Şablon varsayılan değerlere döndürülsün mü?", parent=self):
                return
            if "uyari_sablon" in self.data["bina"]:
                del self.data["bina"]["uyari_sablon"]
            save_data(self.data)
            for w in self.winfo_children(): w.destroy()
            self._build()

        uw_btn_row = frm(ui, bg=T["bg2"])
        uw_btn_row.pack(anchor="w", pady=(10,0))
        btn(uw_btn_row, "Kaydet Şablonu Kaydet", uw_kaydet, "gold").pack(
            side="left", ipadx=12, ipady=4)
        btn(uw_btn_row, "↺ Varsayılana Döndür", uw_sifirla, "gray").pack(
            side="left", padx=8, ipadx=8, ipady=4)

        sag=card_panel(pw,"Kaydet  Yedekleme & Geri Yükleme","")
        pw.add(sag,weight=1)
        ri=frm(sag,bg=T["bg2"]); ri.pack(fill="both",expand=True,padx=20,pady=10)

        lbl(ri,"Yedek Oluştur",fg=T["text2"],bg=T["bg2"],
            font=("Segoe UI",10,"bold")).pack(anchor="w",pady=(0,4))

        # Buton 1 — ZIP Yedeği
        zip_row=frm(ri,bg=T["bg2"]); zip_row.pack(anchor="w",fill="x",pady=(0,6))
        btn(zip_row,"️  ZIP Yedeği Al (.zip)",self._zip_yedek_al,"blue").pack(side="left",ipadx=14,ipady=5)
        lbl(zip_row,"  Tüm JSON dosyalarını tek ZIP'e sıkıştırır.\nYedekten geri yükleme için bu formatı kullanın.",
            fg=T["text3"],bg=T["bg2"],font=("Segoe UI",9),justify="left").pack(side="left",padx=(10,0))

        # Buton 2 — Excel Yedeği
        xl_row=frm(ri,bg=T["bg2"]); xl_row.pack(anchor="w",fill="x",pady=(0,4))
        btn(xl_row,"Grafik  Excel'e Aktar (.xlsx)",self._yedek_al,"green").pack(side="left",ipadx=14,ipady=5)
        lbl(xl_row,"  Tüm veriler Excel formatına dönüştürülür.\nDoğrudan Excel'de açılabilir (geri yükleme desteklenmez).",
            fg=T["text3"],bg=T["bg2"],font=("Segoe UI",9),justify="left").pack(side="left",padx=(10,0))

        tk.Frame(ri,bg=T["border"],height=1).pack(fill="x",pady=14)

        lbl(ri,"Yedekten Geri Yükle",fg=T["text2"],bg=T["bg2"],
            font=("Segoe UI",10,"bold")).pack(anchor="w",pady=(0,4))
        lbl(ri,"ZIP yedek dosyasını (.zip) seçerek tüm veriler geri yüklenir.\nEski JSON yedekleri de desteklenir.",
            fg=T["text3"],bg=T["bg2"],font=("Segoe UI",9)).pack(anchor="w",pady=(0,10))
        btn(ri,"♻️  Yedekten Geri Yükle (.zip / .json)",self._yedek_yukle,"orange").pack(anchor="w",ipadx=14,ipady=5)

        tk.Frame(ri,bg=T["border"],height=1).pack(fill="x",pady=14)

        lbl(ri,"⛔  Tehlikeli Bölge",fg=T["red"],bg=T["bg2"],
            font=("Segoe UI",10,"bold")).pack(anchor="w",pady=(0,6))
        btn(ri,"Sil  Sistemi Sıfırla",self._sifirla,"red").pack(anchor="w",ipadx=10,ipady=4)
        lbl(ri,f"\nKaydet Veri Klasörü:\n{DATA_DIR}",fg=T["text3"],bg=T["bg2"],
            font=("Segoe UI",9),justify="left").pack(anchor="w",pady=(10,0))

    def _sifirla(self):
        if messagebox.askyesno("Sıfırla",
                "TÜM VERİLER SİLİNECEK!\nBu işlem geri alınamaz.\nEmin misiniz?"):
            import shutil
            if DATA_DIR.exists():
                shutil.rmtree(DATA_DIR)
            import sys
            self.destroy()
            os.execv(sys.executable, [sys.executable] + sys.argv)

    def _zip_yedek_al(self):
        """Tüm JSON veri dosyalarını tek bir ZIP arşivine sıkıştır."""
        from tkinter import filedialog
        import zipfile

        now = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        default_name = f"bina_yedek_{now}.zip"
        hedef = filedialog.asksaveasfilename(
            title="ZIP Yedeği Kaydet",
            initialfile=default_name,
            defaultextension=".zip",
            filetypes=[("ZIP Arşivi","*.zip"),("Tüm Dosyalar","*.*")],
            parent=self,
        )
        if not hedef: return

        try:
            eklenen = []
            with zipfile.ZipFile(hedef, "w", compression=zipfile.ZIP_DEFLATED) as zf:
                # Ayarlar dosyası
                if FILE_AYAR.exists():
                    zf.write(FILE_AYAR, FILE_AYAR.name)
                    eklenen.append(FILE_AYAR.name)
                # Genel veriler dosyası
                if FILE_VER.exists():
                    zf.write(FILE_VER, FILE_VER.name)
                    eklenen.append(FILE_VER.name)
                # Daire dosyaları
                for dosya in sorted(DATA_DIR.glob("daire_*.json")):
                    zf.write(dosya, dosya.name)
                    eklenen.append(dosya.name)

            messagebox.showinfo("ZIP Yedeği Alındı",
                f"ZIP yedeği oluşturuldu:\n{hedef}\n\n"
                f"İçindeki dosyalar ({len(eklenen)} adet):\n" +
                "\n".join(f"  • {ad}" for ad in eklenen),
                parent=self)
        except Exception as e:
            messagebox.showerror("Hata", f"ZIP yedeği oluşturulamadı:\n{e}", parent=self)

    def _yedek_al(self):
        """Tüm veriyi .xlsx dosyasına aktar — Excel'de açılabilir."""
        from tkinter import filedialog
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        except ImportError:
            messagebox.showerror("Eksik Modül",
                "openpyxl kurulu değil.\n"
                "Komut satırında şunu çalıştırın:\n"
                "pip install openpyxl", parent=self)
            return

        now = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        default_name = f"bina_yedek_{now}.xlsx"
        hedef = filedialog.asksaveasfilename(
            title="Yedeği Kaydet (.xlsx)",
            initialfile=default_name,
            defaultextension=".xlsx",
            filetypes=[("Excel Dosyası","*.xlsx"),("Tüm Dosyalar","*.*")],
            parent=self,
        )
        if not hedef: return

        try:
            wb = openpyxl.Workbook()
            wb.remove(wb.active)

            HDR_FILL = PatternFill("solid", fgColor="1E2535")
            HDR_FONT = Font(bold=True, color="E8C97A", name="Segoe UI", size=10)
            TITLE_FONT = Font(bold=True, color="C9A84C", name="Segoe UI", size=12)
            CENTER = Alignment(horizontal="center", vertical="center")
            LEFT   = Alignment(horizontal="left",   vertical="center")

            def baslik_sat(ws, row, cols):
                for col,(txt,w) in enumerate(cols,1):
                    c=ws.cell(row=row,column=col,value=txt)
                    c.font=HDR_FONT; c.fill=HDR_FILL; c.alignment=CENTER
                    ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width=w
                ws.row_dimensions[row].height=22

            def p(v):
                try: return round(float(v),2)
                except: return 0.0

            b = self.data["bina"]

            # 1. Bina Ayarları
            ws1=wb.create_sheet("Bina Ayarlari")
            ws1["A1"]=f"{b['adi']} — Bina Ayarları"
            ws1["A1"].font=TITLE_FONT
            ws1.merge_cells("A1:B1"); ws1.row_dimensions[1].height=28
            baslik_sat(ws1,2,[("Ayar",30),("Değer",25)])
            for i,(k,v) in enumerate([
                ("Bina Adı",b.get("adi","")),("Daire Sayısı",b.get("daire",0)),
                ("Kat Sayısı",b.get("kat",0)),("Aylık Aidat (TL)",p(b.get("aidat",0))),
                ("Faiz Oranı (%/ay)",p(b.get("faiz",0))),
                ("Son Ödeme Günü",b.get("son_gun",10)),("Başlangıç Ayı",b.get("baslangic",""))
            ],3):
                ws1.cell(row=i,column=1,value=k).alignment=LEFT
                ws1.cell(row=i,column=2,value=v).alignment=LEFT
            ozel=b.get("ozel_aidatlar",{})
            if ozel:
                r=12
                ws1.cell(row=r,column=1,value="ÖZEL AİDAT PLANI").font=Font(bold=True)
                r+=1
                baslik_sat(ws1,r,[("Dönem",20),("Özel Tutar (TL)",20)])
                r+=1
                for ay,tut in sorted(ozel.items()):
                    ws1.cell(row=r,column=1,value=ay_label(ay))
                    ws1.cell(row=r,column=2,value=p(tut))
                    r+=1

            # 2. Daireler
            ws2=wb.create_sheet("Daireler")
            ws2["A1"]=f"{b['adi']} — Daire Listesi"; ws2["A1"].font=TITLE_FONT
            ws2.merge_cells("A1:H1"); ws2.row_dimensions[1].height=28
            baslik_sat(ws2,2,[("Daire No",10),("Sakin",20),("Telefon",16),("E-posta",22),
                              ("Anapara Borc (TL)",18),("Faiz Borc (TL)",16),
                              ("Toplam Borc (TL)",16),("Son Odeme",18)])
            for i,d in enumerate(self.data["daireler"],3):
                for j,v in enumerate([
                    f"Daire {d['no']}",d.get("isim",""),d.get("tel",""),d.get("email",""),
                    p(d.get("borc",0)),p(d.get("faiz",0)),
                    p(d.get("borc",0)+d.get("faiz",0)),
                    tarih_str(d.get("son_odeme","")) if d.get("son_odeme") else ""
                ],1):
                    ws2.cell(row=i,column=j,value=v).alignment=CENTER if j in(1,5,6,7,8) else LEFT

            # 3. Ödemeler
            ws3=wb.create_sheet("Odemeler")
            ws3["A1"]=f"{b['adi']} — Tüm Aidat Ödemeleri"; ws3["A1"].font=TITLE_FONT
            ws3.merge_cells("A1:G1"); ws3.row_dimensions[1].height=28
            baslik_sat(ws3,2,[("Daire",10),("Sakin",18),("Donem",18),
                              ("Tutar (TL)",14),("Yontem",14),("Tarih",18),("Makbuz No",16)])
            dis={d["no"]:d.get("isim","") for d in self.data["daireler"]}
            for i,o in enumerate(sorted(self.data["odemeler"],key=lambda x:x.get("tarih","")),3):
                for j,v in enumerate([f"Daire {o['daireNo']}",dis.get(o["daireNo"],""),
                    ay_label(o["ay"]),p(o["tutar"]),o.get("yontem",""),
                    tarih_str(o.get("tarih","")),o.get("makbuzNo","")],1):
                    ws3.cell(row=i,column=j,value=v).alignment=CENTER if j in(1,4,5,6,7) else LEFT

            # 4. Gelirler
            ws4=wb.create_sheet("Gelirler")
            ws4["A1"]=f"{b['adi']} — Gelir Kayıtları"; ws4["A1"].font=TITLE_FONT
            ws4.merge_cells("A1:E1"); ws4.row_dimensions[1].height=28
            baslik_sat(ws4,2,[("Fatura Tarihi",18),("Kayıt Tarihi",18),
                              ("Tur",16),("Aciklama",32),("Tutar (TL)",14)])
            for i,g in enumerate(sorted(self.data["gelirler"],key=lambda x:x.get("tarih","")),3):
                ft=g.get("fatura_tarihi") or g.get("tarih","")
                for j,v in enumerate([tarih_str(ft),tarih_str(g.get("tarih","")),
                    g.get("tur",""),g.get("aciklama",""),p(g.get("tutar",0))],1):
                    ws4.cell(row=i,column=j,value=v).alignment=CENTER if j in(1,2,5) else LEFT

            # 5. Giderler
            ws5=wb.create_sheet("Giderler")
            ws5["A1"]=f"{b['adi']} — Gider Kayıtları"; ws5["A1"].font=TITLE_FONT
            ws5.merge_cells("A1:E1"); ws5.row_dimensions[1].height=28
            baslik_sat(ws5,2,[("Fatura Tarihi",18),("Kayıt Tarihi",18),
                              ("Tur",16),("Aciklama",32),("Tutar (TL)",14)])
            for i,g in enumerate(sorted(self.data["giderler"],key=lambda x:x.get("tarih","")),3):
                ft=g.get("fatura_tarihi") or g.get("tarih","")
                for j,v in enumerate([tarih_str(ft),tarih_str(g.get("tarih","")),
                    g.get("tur",""),g.get("aciklama",""),p(g.get("tutar",0))],1):
                    ws5.cell(row=i,column=j,value=v).alignment=CENTER if j in(1,2,5) else LEFT

            # 6. Alacaklılar
            ws6=wb.create_sheet("Alacaklilar")
            ws6["A1"]=f"{b['adi']} — Alacaklı Kayıtları"; ws6["A1"].font=TITLE_FONT
            ws6.merge_cells("A1:H1"); ws6.row_dimensions[1].height=28
            baslik_sat(ws6,2,[("Fatura Tarihi",16),("Tur",14),("Alacakli",18),
                              ("Aciklama",28),("Toplam (TL)",14),("Odenen (TL)",14),
                              ("Kalan (TL)",14),("Durum",14)])
            r6=3
            for a in sorted(self.data.get("alacaklilar",[]),
                            key=lambda x:x.get("fatura_tarihi") or x.get("kayit_tarihi","")):
                od=sum(pp["tutar"] for pp in a.get("odemeler",[]))
                kalan=round(a["tutar"]-od,2)
                tam=a.get("odendi",False) or kalan<=0
                durum="Tamamlandi" if tam else ("Kismi" if od>0 else "Bekliyor")
                ft=a.get("fatura_tarihi") or a.get("kayit_tarihi","")
                for j,v in enumerate([tarih_str(ft),a.get("tur",""),a.get("kisi",""),
                    a.get("aciklama",""),p(a["tutar"]),p(od),p(kalan),durum],1):
                    ws6.cell(row=r6,column=j,value=v).alignment=CENTER if j in(1,5,6,7,8) else LEFT
                r6+=1
                for pp in a.get("odemeler",[]):
                    ws6.cell(row=r6,column=1,value=tarih_str(pp["tarih"]))
                    ws6.cell(row=r6,column=2,value="  -> Odeme")
                    ws6.cell(row=r6,column=5,value=p(pp["tutar"]))
                    ws6.cell(row=r6,column=8,value=pp.get("not_",""))
                    for j in range(1,9):
                        ws6.cell(row=r6,column=j).font=Font(color="8A9AB5",italic=True,size=9)
                    r6+=1

            # 7. Özet
            ws7=wb.create_sheet("Ozet")
            ws7["A1"]=f"{b['adi']} — Finansal Özet"; ws7["A1"].font=TITLE_FONT
            ws7.merge_cells("A1:B1"); ws7.row_dimensions[1].height=28
            baslik_sat(ws7,2,[("Kalem",32),("Tutar (TL)",20)])
            topG=sum(g["tutar"] for g in self.data["gelirler"])
            topGd=sum(g["tutar"] for g in self.data["giderler"])
            topB=sum(d["borc"] for d in self.data["daireler"])
            topF=sum(d["faiz"] for d in self.data["daireler"])
            topAl=sum(max(0,round(a["tutar"]-sum(pp["tutar"] for pp in a.get("odemeler",[])),2))
                      for a in self.data.get("alacaklilar",[]) if not a.get("odendi",False))
            for i,(k,v) in enumerate([
                ("Toplam Gelir",topG),("Toplam Gider",topGd),("Net Bakiye",topG-topGd),
                ("",""),
                ("Toplam Aidat Borcu",topB),("Toplam Faiz Borcu",topF),
                ("Toplam Alacakli Kalan",topAl),("GENEL TOPLAM BORC",topB+topF+topAl),
            ],3):
                ws7.cell(row=i,column=1,value=k).alignment=LEFT
                if v!="":
                    ws7.cell(row=i,column=2,value=p(v)).alignment=CENTER
                if k.startswith("GENEL") or k=="Net Bakiye":
                    ws7.cell(row=i,column=1).font=Font(bold=True)
                    ws7.cell(row=i,column=2).font=Font(bold=True)

            wb.save(hedef)
            messagebox.showinfo("Yedek Alindi",
                f"Excel yedegi olusturuldu:\n{hedef}\n\n"
                "7 sayfa: Bina Ayarları, Daireler, Ödemeler,\n"
                "Gelirler, Giderler, Alacaklılar, Özet",
                parent=self)
            import subprocess, platform
            if platform.system()=="Windows":
                os.startfile(hedef)
            elif platform.system()=="Darwin":
                subprocess.call(["open",hedef])
            else:
                subprocess.call(["xdg-open",hedef])

        except Exception as e:
            messagebox.showerror("Hata",f"Excel yedegi olusturulamadi:\n{e}",parent=self)

    def _yedek_yukle(self):
        """ZIP veya JSON yedekten geri yükle."""
        from tkinter import filedialog
        import zipfile, shutil

        kaynak = filedialog.askopenfilename(
            title="Yedek Dosyası Seç",
            filetypes=[
                ("ZIP Yedeği","*.zip"),
                ("JSON Yedek","*.json"),
                ("Tüm Dosyalar","*.*"),
            ],
            parent=self,
        )
        if not kaynak: return

        ext = Path(kaynak).suffix.lower()

        # ── Excel seçilirse uyar ─────────────────────────────────────────────
        if ext == ".xlsx":
            messagebox.showinfo("Bilgi",
                "Excel dosyalarından geri yükleme desteklenmiyor.\n"
                "Lütfen ZIP (.zip) veya JSON (.json) yedek dosyası seçin.",
                parent=self)
            return

        # ── ZIP geri yükleme ─────────────────────────────────────────────────
        if ext == ".zip":
            try:
                with zipfile.ZipFile(kaynak, "r") as zf:
                    dosya_listesi = zf.namelist()
            except Exception as e:
                messagebox.showerror("Hata", f"ZIP dosyası okunamadı:\n{e}", parent=self)
                return

            json_dosyalar = [d for d in dosya_listesi if d.endswith(".json")]
            if not json_dosyalar:
                messagebox.showerror("Hata",
                    "ZIP arşivi içinde JSON dosyası bulunamadı.\n"
                    "Geçerli bir bina_yonetim ZIP yedeği seçin.",
                    parent=self)
                return

            if not messagebox.askyesno("Geri Yükle",
                    f"ZIP arşivi içindeki {len(json_dosyalar)} dosya geri yüklenecek:\n"
                    + "\n".join(f"  • {d}" for d in json_dosyalar)
                    + "\n\nMevcut veriler silinecek. Emin misiniz?",
                    parent=self):
                return

            try:
                # Mevcut veri klasörünü temizle
                if DATA_DIR.exists():
                    shutil.rmtree(DATA_DIR)
                DATA_DIR.mkdir(parents=True, exist_ok=True)

                # ZIP içindeki tüm JSON dosyalarını veri klasörüne çıkart
                with zipfile.ZipFile(kaynak, "r") as zf:
                    for dosya in json_dosyalar:
                        hedef_yol = DATA_DIR / Path(dosya).name
                        with zf.open(dosya) as src, open(hedef_yol, "wb") as dst:
                            dst.write(src.read())

                messagebox.showinfo("Tamamlandı",
                    f"{len(json_dosyalar)} dosya başarıyla geri yüklendi.\n"
                    "Program yeniden başlatılıyor.",
                    parent=self)
                import sys
                self.destroy()
                os.execv(sys.executable, [sys.executable] + sys.argv)
            except Exception as e:
                messagebox.showerror("Hata", f"Geri yükleme başarısız:\n{e}", parent=self)
            return

        # ── Eski JSON (tek dosya) geri yükleme ───────────────────────────────
        try:
            with open(kaynak, "r", encoding="utf-8") as f:
                test = json.load(f)
            if "bina" not in test or "daireler" not in test:
                raise ValueError("Geçersiz yedek dosyası.")
        except Exception as e:
            messagebox.showerror("Hata", f"Dosya okunamadı:\n{e}", parent=self)
            return

        if not messagebox.askyesno("Geri Yükle",
                "Mevcut veriler silinip JSON yedekten yüklenecek.\nEmin misiniz?",
                parent=self):
            return
        try:
            if DATA_DIR.exists():
                shutil.rmtree(DATA_DIR)
            shutil.copy2(kaynak, DATA_FILE_LEGACY)
            messagebox.showinfo("Tamamlandı",
                "Veriler geri yüklendi. Yeniden başlatılıyor.", parent=self)
            import sys
            self.destroy()
            os.execv(sys.executable, [sys.executable] + sys.argv)
        except Exception as e:
            messagebox.showerror("Hata", f"Geri yükleme başarısız:\n{e}", parent=self)



# ══════════════════════════════════════════════════════════════════════════════
#  ALACAKLI ÖDEME PENCERESİ
# ══════════════════════════════════════════════════════════════════════════════
class AlacakliOdemePencere(tk.Toplevel):
    """
    Seçili alacaklıya kısmi veya tam ödeme yap.
    Alacaklı adı ve miktar el ile düzenlenebilir.
    Kalan tutar anlık güncellenir.
    """
    def __init__(self, parent, data, alacakli, callback=None):
        super().__init__(parent)
        self.data      = data
        self.alacakli  = alacakli
        self.callback  = callback

        kalan = round(alacakli["tutar"] -
                      sum(p["tutar"] for p in alacakli.get("odemeler",[])), 2)
        self.title(f"Kart  Ödeme Yap — {alacakli.get('kisi') or alacakli['tur']}")
        self.configure(bg=T["bg"])
        self.resizable(False, False)
        self.grab_set()
        self.update_idletasks()
        sw, sh = self.winfo_screenwidth(), self.winfo_screenheight()
        self.geometry(f"500x550+{(sw-500)//2}+{(sh-550)//2}")
        self._kalan_toplam = kalan
        self._build(kalan)

    def _build(self, kalan):
        a = self.alacakli
        tk.Frame(self, bg=T["gold"], height=3).pack(fill="x")

        # Başlık
        lbl(self, "Kart  Ödeme Yap", font=("Segoe UI",13,"bold")).pack(
            pady=(14,4), padx=20, anchor="w")

        # Özet kartı
        ozet = frm(self, bg=T["bg3"],
            highlightthickness=1, highlightbackground=T["border"])
        ozet.pack(fill="x", padx=20, pady=(0,12))
        ozet_i = frm(ozet, bg=T["bg3"]); ozet_i.pack(fill="x", padx=14, pady=10)

        for col,(ltext,val,fg) in enumerate([
            ("Alacaklı / Kişi", a.get("kisi") or a["tur"],          T["text"]),
            ("Fatura Tutarı",   fmt(a["tutar"]),                     T["text"]),
            ("Daha Önce Ödenen",fmt(a["tutar"]-kalan),               T["green"]),
            ("Kalan Borç",      fmt(kalan),                          T["red"]),
        ]):
            cf = frm(ozet_i, bg=T["bg3"]); cf.pack(side="left", expand=True, anchor="w")
            lbl(cf, ltext, fg=T["text3"], bg=T["bg3"], font=("Segoe UI",8)).pack(anchor="w")
            lbl(cf, val,   fg=fg,         bg=T["bg3"], font=("Segoe UI",10,"bold")).pack(anchor="w")

        tk.Frame(self, bg=T["border"], height=1).pack(fill="x", padx=20)

        # Form
        form = frm(self); form.pack(fill="x", padx=20, pady=12)

        # Ödemeyi yapan kişi (düzenlenebilir)
        lbl(form, "Ödemeyi Yapan Kişi", fg=T["text2"],
            font=("Segoe UI",9,"bold")).pack(anchor="w", pady=(0,3))
        self._kisi_var = tk.StringVar(value=a.get("kisi",""))
        ent(form, textvariable=self._kisi_var, width=36).pack(
            fill="x", ipady=7, pady=(0,10))

        # Ödeme tutarı — hızlı butonlar
        lbl(form, "Ödeme Tutarı (₺)", fg=T["text2"],
            font=("Segoe UI",9,"bold")).pack(anchor="w", pady=(0,3))
        hiz = frm(form); hiz.pack(fill="x", pady=(0,4))
        self._tutar_var = tk.StringVar(value=str(kalan))
        ent(form, textvariable=self._tutar_var, width=20).pack(
            anchor="w", ipady=8)

        # Hızlı tutar butonları
        hiz2 = frm(form); hiz2.pack(anchor="w", pady=(4,10))
        for etiket, deger in [
            ("Tamamı",   kalan),
            ("%50",      round(kalan*0.5, 2)),
            ("%25",      round(kalan*0.25, 2)),
            ("Özel",     None),
        ]:
            d = deger
            btn(hiz2, etiket,
                lambda d2=d: self._tutar_var.set("" if d2 is None else str(d2)),
                "gray").pack(side="left", padx=(0,4))

        # Kalan göstergesi — anlık güncellenir
        self._kalan_lbl = lbl(form, f"Bu ödemeden sonra kalan: {fmt(kalan)}",
            fg=T["gold2"], font=("Segoe UI",10,"bold"))
        self._kalan_lbl.pack(anchor="w", pady=(0,4))
        self._tutar_var.trace_add("write", self._guncelle_kalan)

        # Ödeme tarihi
        lbl(form, "Ödeme Tarihi (GG.AA.YYYY)", fg=T["text2"],
            font=("Segoe UI",9,"bold")).pack(anchor="w", pady=(4,3))
        self._tarih_var = tk.StringVar(
            value=datetime.date.today().strftime("%d.%m.%Y"))
        ent(form, textvariable=self._tarih_var, width=16).pack(
            anchor="w", ipady=6)

        # Not
        lbl(form, "Not (isteğe bağlı)", fg=T["text2"],
            font=("Segoe UI",9,"bold")).pack(anchor="w", pady=(10,3))
        self._not_var = tk.StringVar()
        ent(form, textvariable=self._not_var, width=36).pack(
            fill="x", ipady=6)

        # Butonlar
        tk.Frame(self, bg=T["border"], height=1).pack(fill="x", padx=20, pady=(10,0))
        btn_f = frm(self); btn_f.pack(fill="x", padx=20, pady=12)
        btn(btn_f, "X İptal",    self.destroy,  "gray").pack(side="right", padx=(6,0))
        btn(btn_f, "OK Ödemeyi Kaydet", self._kaydet, "gold").pack(side="right")

    def _guncelle_kalan(self, *_):
        try:
            odenen = float(self._tutar_var.get())
            sonraki_kalan = round(self._kalan_toplam - odenen, 2)
            if sonraki_kalan < 0:
                self._kalan_lbl.config(
                    text=f"Uyari Fazla ödeme: {fmt(abs(sonraki_kalan))}",
                    fg=T["orange"])
            elif sonraki_kalan == 0:
                self._kalan_lbl.config(
                    text="OK Borç tamamen kapanacak", fg=T["green"])
            else:
                self._kalan_lbl.config(
                    text=f"Bu ödemeden sonra kalan: {fmt(sonraki_kalan)}",
                    fg=T["gold2"])
        except:
            self._kalan_lbl.config(
                text="Geçersiz tutar", fg=T["red"])

    def _kaydet(self):
        try:
            tutar = float(self._tutar_var.get())
            if tutar <= 0: raise ValueError
        except:
            messagebox.showerror("Hata", "Geçerli ve pozitif tutar girin.", parent=self)
            return
        try:
            t = datetime.datetime.strptime(self._tarih_var.get(), "%d.%m.%Y")
            tarih_iso = t.isoformat(timespec="seconds")
        except:
            messagebox.showerror("Hata", "Tarih GG.AA.YYYY formatında olmalı.", parent=self)
            return

        kisi = self._kisi_var.get().strip()
        if kisi:
            self.alacakli["kisi"] = kisi  # kişi adını güncelle

        odeme_kaydi = {
            "id"   : yeni_id(),
            "tutar": round(tutar, 2),
            "tarih": tarih_iso,
            "not_" : self._not_var.get().strip(),
        }
        self.alacakli.setdefault("odemeler", []).append(odeme_kaydi)

        # Kalan sıfırsa tamamen ödendi işaretle
        kalan_sonra = round(self._kalan_toplam - tutar, 2)
        if kalan_sonra <= 0:
            self.alacakli["odendi"]       = True
            self.alacakli["odeme_tarihi"] = tarih_iso

        # Ortak Gider ise ayrı makbuz oluştur
        if self.alacakli.get("tur") == "Ortak Gider":
            og_makbuz_yazdir(self.data, self.alacakli, odeme_kaydi)

        # Nakit Harcama HARİÇ: ödeme yapılınca muhasebe kaydı oluştur
        a_tur = self.alacakli.get("tur", "")
        aciklama_parcalari = []
        if self.alacakli.get("kisi"):
            aciklama_parcalari.append(self.alacakli["kisi"])
        if self.alacakli.get("aciklama"):
            aciklama_parcalari.append(self.alacakli["aciklama"])
        aciklama = " — ".join(aciklama_parcalari) if aciklama_parcalari else a_tur

        if a_tur == "Ortak Gider":
            # Ortak gider tahsilatı → gelir hesabına
            # Açıklama: Daire No — Sakin İsmi — İş Adı
            dno = self.alacakli.get("daire_no")
            d_obj = next((x for x in self.data.get("daireler", [])
                          if x["no"] == dno), {}) if dno else {}
            d_isim = d_obj.get("isim", "").strip()
            is_adi = self.alacakli.get("kisi", "").strip()  # kisi = gider adı
            og_aciklama_parts = [f"Daire {dno}" if dno else ""]
            if d_isim:
                og_aciklama_parts.append(d_isim)
            if is_adi:
                og_aciklama_parts.append(is_adi)
            og_aciklama = " — ".join(p for p in og_aciklama_parts if p)
            self.data.setdefault("gelirler", []).append({
                "id"         : odeme_kaydi["id"] + 1,
                "tur"        : "Ortak Gider Tahsilatı",
                "aciklama"   : og_aciklama,
                "tarih"      : tarih_iso,
                "tutar"      : round(tutar, 2),
                "alacakli_id": self.alacakli["id"],
            })
        elif a_tur != "Nakit Harcama":
            # Diğer alacaklı ödemeleri → gider hesabına
            self.data.setdefault("giderler", []).append({
                "id"            : odeme_kaydi["id"] + 1,
                "tur"           : a_tur,
                "aciklama"      : aciklama,
                "tarih"         : tarih_iso,
                "fatura_tarihi" : tarih_iso,
                "tutar"         : round(tutar, 2),
                "alacakli_id"   : self.alacakli["id"],
            })

        save_data(self.data)
        self.destroy()
        if self.callback: self.callback()


# ══════════════════════════════════════════════════════════════════════════════
#  ALACAKLI ÖDEME GEÇMİŞİ PENCERESİ
# ══════════════════════════════════════════════════════════════════════════════
class AlacakliGecmisPencere(tk.Toplevel):
    """
    Seçili alacaklının (veya ortak giderin) ödeme geçmişini göster.
    Ortak Gider türünde: düzenle ve sil işlemleri de desteklenir.
    """
    def __init__(self, parent, alacakli, data=None, callback=None):
        super().__init__(parent)
        self.alacakli = alacakli
        self.data     = data        # None ise salt-görüntüleme modu
        self.callback = callback
        a = alacakli
        is_og = (a.get("tur") == "Ortak Gider")
        baslik = a.get("kisi") or a.get("aciklama","") or a["tur"]
        self.title(f"  Ödeme Geçmişi — {baslik}")
        self.configure(bg=T["bg"])
        w, h = (720, 520) if is_og else (620, 440)
        self.geometry(f"{w}x{h}"); self.minsize(560, 380)
        self.grab_set()
        self.update_idletasks()
        sw, sh = self.winfo_screenwidth(), self.winfo_screenheight()
        self.geometry(f"{w}x{h}+{(sw-w)//2}+{(sh-h)//2}")
        self._build(is_og)

    def _build(self, is_og):
        a = self.alacakli
        tk.Frame(self, bg=T["gold"], height=3).pack(fill="x")

        # Başlık
        hdr_frm = frm(self, bg=T["bg2"]); hdr_frm.pack(fill="x")
        hf = frm(hdr_frm, bg=T["bg2"]); hf.pack(fill="x", padx=14, pady=8)
        baslik_txt = a.get("kisi") or a.get("aciklama","") or a["tur"]
        lbl(hf, f"  {baslik_txt}",
            font=("Segoe UI",13,"bold"), bg=T["bg2"]).pack(side="left")
        if is_og and a.get("daire_no"):
            lbl(hf, f"  •  Daire {a['daire_no']}",
                fg=T["text2"], bg=T["bg2"], font=("Segoe UI",10)).pack(side="left")
        tk.Frame(hdr_frm, bg=T["border"], height=1).pack(fill="x")

        # Özet şerit
        self._ozet_frm = frm(self, bg=T["bg3"],
            highlightthickness=1, highlightbackground=T["border"])
        self._ozet_frm.pack(fill="x", padx=16, pady=(10,0))
        self._ozet_guncelle()

        # Liste
        wrap = frm(self); wrap.pack(fill="both", expand=True, padx=16, pady=(8,0))

        cols = [("Tarih",130),("Tutar",110),("Not",220)]
        if is_og:
            cols.append(("Makbuz No", 160))
        tf, self._tree = scrolled(wrap, cols, height=9)
        tf.pack(fill="both", expand=True)
        self._listele()

        if is_og:
            self._tree.bind("<<TreeviewSelect>>", self._on_sec)

        # Butonlar
        btn_row = frm(self); btn_row.pack(fill="x", padx=16, pady=(8,12))
        if is_og and self.data is not None:
            btn(btn_row, "Duzenle  Düzenle",   self._duzenle, "blue"  ).pack(side="left", ipadx=8, ipady=4)
            btn(btn_row, "Sil  Sil",       self._sil,     "red"   ).pack(side="left", padx=6, ipadx=8, ipady=4)
            btn(btn_row, "Yazdir  Makbuz",    self._makbuz,  "gray"  ).pack(side="left", ipadx=8, ipady=4)
        btn(btn_row, "OK Kapat", self.destroy, "gold").pack(side="right", ipadx=16, ipady=4)

    # ── Yardımcılar ───────────────────────────────────────────────────────────
    def _ozet_guncelle(self):
        for w in self._ozet_frm.winfo_children():
            w.destroy()
        a = self.alacakli
        oi = frm(self._ozet_frm, bg=T["bg3"]); oi.pack(fill="x", padx=14, pady=8)
        toplam_odenen = sum(p["tutar"] for p in a.get("odemeler",[]))
        kalan = round(a["tutar"] - toplam_odenen, 2)
        for ltext, val, fg in [
            ("Fatura Tutarı",  fmt(a["tutar"]),                               T["text"]),
            ("Toplam Ödenen",  fmt(toplam_odenen),                            T["green"]),
            ("Kalan",          fmt(kalan) if kalan > 0 else "v Tamam",
                               T["red"] if kalan > 0 else T["green"]),
        ]:
            cf = frm(oi, bg=T["bg3"]); cf.pack(side="left", expand=True, anchor="w")
            lbl(cf, ltext, fg=T["text3"], bg=T["bg3"], font=("Segoe UI",8)).pack(anchor="w")
            lbl(cf, val,   fg=fg,         bg=T["bg3"], font=("Segoe UI",11,"bold")).pack(anchor="w")

    def _listele(self):
        self._tree.delete(*self._tree.get_children())
        a = self.alacakli
        is_og = (a.get("tur") == "Ortak Gider")
        odemeler = a.get("odemeler", [])
        if not odemeler:
            self._tree.insert("","end", tags=("normal",),
                values=("Henüz ödeme yapılmamış","","") + (("—",) if is_og else ()))
        else:
            for p in reversed(odemeler):
                mkb_no = f"OG-{datetime.date.today().year}-{p['id'] % 10000:04d}"
                row = (tarih_str(p["tarih"]), fmt(p["tutar"]), p.get("not_","") or "—")
                if is_og:
                    row += (mkb_no,)
                self._tree.insert("","end", iid=str(p["id"]), tags=("odendi",), values=row)

    def _secili_odeme(self):
        sel = self._tree.selection()
        if not sel: return None
        pid = int(sel[0])
        return next((p for p in self.alacakli.get("odemeler",[]) if p["id"]==pid), None)

    def _on_sec(self, event=None):
        pass  # seçim değişiminde özel işlem gerekmez

    # ── Düzenle ───────────────────────────────────────────────────────────────
    def _duzenle(self):
        p = self._secili_odeme()
        if not p:
            messagebox.showinfo("Seçim","Lütfen düzenlenecek ödemeyi seçin.",parent=self)
            return
        OGOdemeDuzenle(self, self.data, self.alacakli, p,
                       callback=self._yenile)

    # ── Sil ───────────────────────────────────────────────────────────────────
    def _sil(self):
        p = self._secili_odeme()
        if not p:
            messagebox.showinfo("Seçim","Lütfen silinecek ödemeyi seçin.",parent=self)
            return
        a = self.alacakli
        gider_adi = a.get("kisi") or a.get("aciklama","") or "Ortak Gider"
        if not messagebox.askyesno("Odemeyi Sil",
            "Gider: " + gider_adi + "\nTutar: " + fmt(p["tutar"]) + "\n"
            "Tarih: " + tarih_str(p["tarih"]) + "\n\nBu islem geri alinamaz.",
            parent=self):
            return

        # Eşleşen muhasebe kaydını sil (gelir veya gider olabilir)
        if self.data:
            def _eslesmiyor(kayit):
                return not (
                    (kayit.get("odeme_id") == p["id"] or kayit.get("id") == p["id"] + 1)
                    or (kayit.get("alacakli_id") == a["id"]
                        and abs(kayit["tutar"] - p["tutar"]) < 0.01
                        and kayit["tarih"][:10] == p["tarih"][:10])
                )
            self.data["gelirler"] = [
                g for g in self.data.get("gelirler", []) if _eslesmiyor(g)
            ]
            self.data["giderler"] = [
                g for g in self.data.get("giderler", []) if _eslesmiyor(g)
            ]

        # Ödemeyi listeden çıkar
        a["odemeler"] = [x for x in a.get("odemeler",[]) if x["id"] != p["id"]]

        # Durum güncelle
        toplam_odenen = sum(x["tutar"] for x in a["odemeler"])
        kalan = round(a["tutar"] - toplam_odenen, 2)
        if kalan > 0.01:
            a["odendi"] = False
            a["odeme_tarihi"] = ""
        else:
            a["odendi"] = True

        if self.data:
            save_data(self.data)
        self._yenile()

    # ── Makbuz ────────────────────────────────────────────────────────────────
    def _makbuz(self):
        p = self._secili_odeme()
        if not p:
            messagebox.showinfo("Seçim","Lütfen makbuz almak istediğiniz ödemeyi seçin.",parent=self)
            return
        if self.data:
            og_makbuz_yazdir(self.data, self.alacakli, p)
        else:
            messagebox.showinfo("Bilgi","Makbuz için veri bağlamı gerekli.",parent=self)

    # ── Yenile ────────────────────────────────────────────────────────────────
    def _yenile(self):
        self._ozet_guncelle()
        self._listele()
        if self.callback:
            self.callback()


# ══════════════════════════════════════════════════════════════════════════════
#  ORTAK GİDER ÖDEME DÜZENLEME DİYALOĞU
# ══════════════════════════════════════════════════════════════════════════════
class OGOdemeDuzenle(tk.Toplevel):
    """Ortak gider tekil ödemesini düzenle (tutar, tarih, not)."""
    def __init__(self, parent, data, alacakli, odeme, callback=None):
        super().__init__(parent)
        self.data      = data
        self.alacakli  = alacakli
        self.odeme     = odeme
        self.callback  = callback
        gider_adi = alacakli.get("kisi") or alacakli.get("aciklama","") or "Ortak Gider"
        self.title(f"Duzenle  Ödeme Düzenle — {gider_adi}")
        self.configure(bg=T["bg"])
        self.resizable(False, False)
        self.grab_set()
        self.update_idletasks()
        sw, sh = self.winfo_screenwidth(), self.winfo_screenheight()
        self.geometry(f"440x380+{(sw-440)//2}+{(sh-380)//2}")
        self._build(gider_adi)

    def _build(self, gider_adi):
        p  = self.odeme
        a  = self.alacakli
        tk.Frame(self, bg=T["gold"], height=3).pack(fill="x")
        lbl(self, "Duzenle  Ortak Gider Ödemesi Düzenle",
            font=("Segoe UI",13,"bold")).pack(pady=(14,2), padx=20, anchor="w")
        lbl(self, f"{gider_adi}  •  Daire {a.get('daire_no','?')}  •  Toplam: {fmt(a['tutar'])}",
            fg=T["text2"], font=("Segoe UI",9)).pack(padx=20, anchor="w", pady=(0,8))
        tk.Frame(self, bg=T["border"], height=1).pack(fill="x")

        form = frm(self); form.pack(fill="x", padx=20, pady=12)

        # Tutar
        lbl(form,"Tutar (₺)", fg=T["text2"],
            font=("Segoe UI",9,"bold")).pack(anchor="w", pady=(0,3))
        self._tutar_var = tk.StringVar(value=str(p["tutar"]))
        ent(form, textvariable=self._tutar_var, width=30).pack(fill="x", ipady=7)

        # Tarih
        lbl(form,"Ödeme Tarihi (GG.AA.YYYY)", fg=T["text2"],
            font=("Segoe UI",9,"bold")).pack(anchor="w", pady=(12,3))
        try:
            mevcut_tarih = datetime.datetime.fromisoformat(p["tarih"]).strftime("%d.%m.%Y")
        except:
            mevcut_tarih = datetime.date.today().strftime("%d.%m.%Y")
        self._tarih_var = tk.StringVar(value=mevcut_tarih)
        ent(form, textvariable=self._tarih_var, width=20).pack(anchor="w", ipady=7)

        # Not
        lbl(form,"Not", fg=T["text2"],
            font=("Segoe UI",9,"bold")).pack(anchor="w", pady=(12,3))
        self._not_var = tk.StringVar(value=p.get("not_","") or "")
        ent(form, textvariable=self._not_var, width=38).pack(fill="x", ipady=7)

        tk.Frame(self, bg=T["border"], height=1).pack(fill="x")
        bf = frm(self); bf.pack(fill="x", padx=20, pady=12)
        btn(bf,"X İptal",  self.destroy,  "gray").pack(side="right", padx=(6,0))
        btn(bf,"OK Kaydet", self._kaydet,  "gold").pack(side="right")
        btn(bf,"Yazdir Makbuz",self._makbuz,  "gray").pack(side="left")

    def _kaydet(self):
        try:
            tutar = float(self._tutar_var.get())
            if tutar <= 0: raise ValueError
        except:
            messagebox.showerror("Hata","Geçerli ve pozitif tutar girin.",parent=self)
            return
        try:
            t = datetime.datetime.strptime(self._tarih_var.get().strip(), "%d.%m.%Y")
            tarih_iso = t.isoformat(timespec="seconds")
        except:
            messagebox.showerror("Hata","Tarih GG.AA.YYYY formatında olmalı.",parent=self)
            return

        eski_tutar = self.odeme["tutar"]
        self.odeme["tutar"]  = round(tutar, 2)
        self.odeme["tarih"]  = tarih_iso
        self.odeme["not_"]   = self._not_var.get().strip()

        # Eşleşen gider kaydını güncelle
        for g in self.data.get("giderler",[]):
            if (g.get("alacakli_id") == self.alacakli["id"]
                    and abs(g["tutar"] - eski_tutar) < 0.01):
                g["tutar"] = round(tutar, 2)
                g["tarih"] = tarih_iso
                break

        # Alacaklı durum güncelle
        a = self.alacakli
        toplam_odenen = sum(x["tutar"] for x in a.get("odemeler",[]))
        kalan = round(a["tutar"] - toplam_odenen, 2)
        a["odendi"] = kalan <= 0.01
        if a["odendi"] and not a.get("odeme_tarihi"):
            a["odeme_tarihi"] = tarih_iso

        save_data(self.data)
        self.destroy()
        if self.callback: self.callback()

    def _makbuz(self):
        og_makbuz_yazdir(self.data, self.alacakli, self.odeme)


# ══════════════════════════════════════════════════════════════════════════════
#  ORTAK GİDER / TADİLAT YÖNETİMİ
# ══════════════════════════════════════════════════════════════════════════════

def _sonraki_ay(ay_str, n=1):
    """ay_str'den n ay sonrasını döndür (YYYY-MM)."""
    y, m = int(ay_str[:4]), int(ay_str[5:7])
    m += n
    while m > 12:
        m -= 12; y += 1
    return f"{y}-{m:02d}"

def _ortak_gider_aidata_uygula(data, gider):
    """
    Ortak giderin her daire payını, her daire için ayrı bir 'Ortak Gider'
    alacaklı kaydı olarak alacaklilar listesine ekler.
    Taksit bilgisi kayıtta referans olarak tutulur.
    Zaten işlenmiş daireleri tekrar eklemez.
    """
    daire_nos   = gider.get("daire_nos", [])
    taksitler   = gider.get("taksitler", [])
    toplam_taksit = round(sum(t["tutar"] for t in taksitler), 2)

    if len(taksitler) == 1:
        taksit_not = f"Tek seferlik • {ay_label(taksitler[0]['ay'])}"
    elif taksitler:
        taksit_not = (f"{len(taksitler)} taksit • "
                      f"{ay_label(taksitler[0]['ay'])} – {ay_label(taksitler[-1]['ay'])}")
    else:
        taksit_not = ""

    eklenenler = list(gider.get("eklenenler", []))
    zaten_daireler = {k["daire_no"] for k in eklenenler}

    for dno in daire_nos:
        if dno in zaten_daireler:
            continue
        aciklama_parts = []
        if gider.get("aciklama"):
            aciklama_parts.append(gider["aciklama"])
        if taksit_not:
            aciklama_parts.append(f"[{taksit_not}]")
        kayit = {
            "id"            : yeni_id(),
            "tur"           : "Ortak Gider",
            "kisi"          : gider.get("ad", ""),
            "aciklama"      : "  ".join(aciklama_parts),
            "tutar"         : toplam_taksit,
            "fatura_tarihi" : gider.get("tarih", simdi()),
            "kayit_tarihi"  : simdi(),
            "odendi"        : False,
            "odeme_tarihi"  : "",
            "odemeler"      : [],
            "ortak_gider_id": gider["id"],
            "daire_no"      : dno,
            "taksitler"     : taksitler,
        }
        data.setdefault("alacaklilar", []).append(kayit)
        eklenenler.append({"daire_no": dno, "alacakli_id": kayit["id"]})

    gider["eklenenler"] = eklenenler
    gider["uygulandi"]  = True


def _ortak_gider_aidat_geri_al(data, gider):
    """Daha önce oluşturulmuş alacaklı kayıtlarını sil."""
    silinen_idler = {k["alacakli_id"] for k in gider.get("eklenenler", [])
                     if "alacakli_id" in k}
    data["alacaklilar"] = [
        a for a in data.get("alacaklilar", [])
        if a["id"] not in silinen_idler
    ]
    gider["eklenenler"] = []
    gider["uygulandi"]  = False


class OrtakGiderPencere(tk.Toplevel):
    """Yeni ortak gider / tadilat tanımlama ve taksitlendirme penceresi."""

    def __init__(self, parent, data, callback=None, duzenle=None):
        super().__init__(parent)
        self.data     = data
        self.callback = callback
        self.duzenle  = duzenle   # None = yeni kayıt, dict = mevcut kaydı düzenle
        baslık = "Duzenle  Ortak Gider Düzenle" if duzenle else "+  Yeni Ortak Gider / Tadilat"
        self.title(baslık)
        self.configure(bg=T["bg"])
        self.geometry("700x660")
        self.minsize(600, 580)
        self.grab_set()
        self.update_idletasks()
        sw, sh = self.winfo_screenwidth(), self.winfo_screenheight()
        self.geometry(f"700x660+{(sw-700)//2}+{(sh-660)//2}")
        self._build()
        if duzenle:
            self._doldur(duzenle)

    # ─── FORM OLUŞTUR ────────────────────────────────────────────────────────
    def _build(self):
        b = self.data["bina"]
        daire_sayisi = b["daire"]

        tk.Frame(self, bg=T["gold"], height=3).pack(fill="x")

        # Başlık
        hdr = frm(self, bg=T["bg2"]); hdr.pack(fill="x")
        lf  = frm(hdr, bg=T["bg2"]); lf.pack(fill="x", padx=14, pady=10)
        lbl(lf, "Arac  Ortak Gider / Tadilat Tanımla",
            font=("Segoe UI",13,"bold"), bg=T["bg2"]).pack(side="left")
        tk.Frame(hdr, bg=T["border"], height=1).pack(fill="x")

        # ── Kaydırılabilir form alanı
        canvas = tk.Canvas(self, bg=T["bg"], highlightthickness=0)
        vsb    = ttk.Scrollbar(self, orient="vertical", command=canvas.yview,
                               style="Dark.Vertical.TScrollbar")
        canvas.configure(yscrollcommand=vsb.set)
        vsb.pack(side="right", fill="y")
        canvas.pack(side="left", fill="both", expand=True)
        inner = frm(canvas, bg=T["bg"])
        inner.bind("<Configure>",
                   lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.create_window((0,0), window=inner, anchor="nw")

        form = frm(inner, bg=T["bg3"],
                   highlightthickness=1, highlightbackground=T["border"])
        form.pack(fill="x", padx=14, pady=10)
        fi = frm(form, bg=T["bg3"]); fi.pack(fill="x", padx=14, pady=12)
        fi.columnconfigure(1, weight=1)

        # İş Adı
        lbl(fi, "İş / Gider Adı:", fg=T["text2"], bg=T["bg3"],
            font=("Segoe UI",9,"bold")).grid(row=0, column=0, sticky="w", pady=4, padx=(0,10))
        self._ad_var = tk.StringVar()
        _IS_TURLERI = [
            "Tadilat / Tamirat", "Boya / Badana", "Bakım",
            "Asansör Bakımı", "Asansör Tamiri",
            "Su Tesisatı", "Elektrik Tesisatı",
            "Çatı Onarımı", "Isı Yalıtımı",
            "Temizlik", "Peyzaj / Bahçe",
            "Güvenlik Sistemi", "Kapı / Kilit",
            "Cam / Pencere", "Zemin / Döşeme",
            "Diğer",
        ]
        ad_cb = ttk.Combobox(fi, textvariable=self._ad_var,
            values=_IS_TURLERI, style="Dark.TCombobox",
            font=("Segoe UI", 10), width=36)
        ad_cb.grid(row=0, column=1, sticky="ew", ipady=6)
        ad_cb.bind("<FocusOut>", lambda e: None)   # serbest yazıma izin ver

        # Açıklama
        lbl(fi, "Açıklama:", fg=T["text2"], bg=T["bg3"],
            font=("Segoe UI",9,"bold")).grid(row=1, column=0, sticky="nw", pady=8, padx=(0,10))
        self._acik_var = tk.StringVar()
        ent(fi, textvariable=self._acik_var, width=38).grid(
            row=1, column=1, sticky="ew", ipady=6)

        # Toplam Tutar
        lbl(fi, "Toplam Tutar (₺):", fg=T["text2"], bg=T["bg3"],
            font=("Segoe UI",9,"bold")).grid(row=2, column=0, sticky="w", pady=4, padx=(0,10))
        tf = frm(fi, bg=T["bg3"]); tf.grid(row=2, column=1, sticky="w")
        self._toplam_var = tk.StringVar()
        self._toplam_var.trace_add("write", lambda *_: self._hesapla())
        ent(tf, textvariable=self._toplam_var, width=16).pack(side="left", ipady=6)

        # Daire Başı Tutar
        lbl(fi, "Daire Başı Tutar (₺):", fg=T["text2"], bg=T["bg3"],
            font=("Segoe UI",9,"bold")).grid(row=3, column=0, sticky="w", pady=4, padx=(0,10))
        df = frm(fi, bg=T["bg3"]); df.grid(row=3, column=1, sticky="w")
        self._dbasI_var = tk.StringVar()
        self._dbasI_var.trace_add("write", lambda *_: self._hesapla_ters())
        ent(df, textvariable=self._dbasI_var, width=16).pack(side="left", ipady=6)
        self._hesap_lbl = lbl(df, "", fg=T["gold2"], bg=T["bg3"],
                              font=("Segoe UI",9,"bold"))
        self._hesap_lbl.pack(side="left", padx=8)
        self._hesap_lock = False   # döngüsel güncelleme koruması

        # ── Daire Seçimi
        sep_frm = frm(fi, bg=T["border"], height=1)
        sep_frm.grid(row=4, column=0, columnspan=2, sticky="ew", pady=8)

        lbl(fi, "Kapsam:", fg=T["text2"], bg=T["bg3"],
            font=("Segoe UI",9,"bold")).grid(row=5, column=0, sticky="nw", pady=4, padx=(0,10))
        kp = frm(fi, bg=T["bg3"]); kp.grid(row=5, column=1, sticky="w")
        self._kapsam_var = tk.StringVar(value="tum")
        for val, txt in [("tum", "Tüm Daireler"), ("secili", "Seçili Daireler")]:
            tk.Radiobutton(kp, text=txt, variable=self._kapsam_var, value=val,
                bg=T["bg3"], fg=T["text"], selectcolor=T["bg4"],
                activebackground=T["bg3"], font=("Segoe UI",9),
                cursor="hand2",
                command=self._kapsam_degis).pack(side="left", padx=8)

        # Daire seçim listesi (multi-select)
        self._daire_frame = frm(fi, bg=T["bg3"])
        self._daire_frame.grid(row=6, column=0, columnspan=2, sticky="ew", pady=(4,8))
        lbl(self._daire_frame, "Daireleri Seçin:",
            fg=T["text3"], bg=T["bg3"], font=("Segoe UI",8)).pack(anchor="w")
        lb_f = frm(self._daire_frame, bg=T["bg3"],
                   highlightthickness=1, highlightbackground=T["border"])
        lb_f.pack(fill="x")
        vsb2 = ttk.Scrollbar(lb_f, orient="vertical", style="Dark.Vertical.TScrollbar")
        self._daire_lb = tk.Listbox(lb_f, selectmode="multiple", height=5,
            bg=T["entry_bg"], fg=T["text"], font=("Segoe UI",10),
            selectbackground=T["sel"], selectforeground=T["gold2"],
            relief="flat", bd=0, activestyle="none",
            yscrollcommand=vsb2.set)
        vsb2.config(command=self._daire_lb.yview)
        self._daire_lb.pack(side="left", fill="both", expand=True)
        vsb2.pack(side="right", fill="y")
        for i in range(1, daire_sayisi+1):
            d = next((x for x in self.data["daireler"] if x["no"] == i), {})
            isim = d.get("isim", "") or ""
            self._daire_lb.insert("end",
                f"Daire {i}" + (f"  ({isim})" if isim else ""))
        self._daire_frame.grid_remove()

        # ── Taksit
        sep_frm2 = frm(fi, bg=T["border"], height=1)
        sep_frm2.grid(row=7, column=0, columnspan=2, sticky="ew", pady=8)

        lbl(fi, "Ödeme Şekli:", fg=T["text2"], bg=T["bg3"],
            font=("Segoe UI",9,"bold")).grid(row=8, column=0, sticky="nw", pady=4, padx=(0,10))
        op = frm(fi, bg=T["bg3"]); op.grid(row=8, column=1, sticky="w")
        self._odeme_var = tk.StringVar(value="tek")
        for val, txt in [("tek", "Tek Seferlik"), ("taksit", "Taksitli")]:
            tk.Radiobutton(op, text=txt, variable=self._odeme_var, value=val,
                bg=T["bg3"], fg=T["text"], selectcolor=T["bg4"],
                activebackground=T["bg3"], font=("Segoe UI",9),
                cursor="hand2",
                command=self._odeme_degis).pack(side="left", padx=8)

        # Başlangıç ayı
        lbl(fi, "Başlangıç Ayı:", fg=T["text2"], bg=T["bg3"],
            font=("Segoe UI",9,"bold")).grid(row=9, column=0, sticky="w", pady=4, padx=(0,10))
        self._bas_ay = AySecici(fi, bg=T["bg3"],
                                callback=lambda _: self._onizle())
        self._bas_ay.grid(row=9, column=1, sticky="w")

        # Taksit sayısı
        self._taksit_lbl = lbl(fi, "Taksit Sayısı:", fg=T["text2"], bg=T["bg3"],
                                font=("Segoe UI",9,"bold"))
        self._taksit_lbl.grid(row=10, column=0, sticky="w", pady=4, padx=(0,10))
        self._taksit_frame = frm(fi, bg=T["bg3"])
        self._taksit_frame.grid(row=10, column=1, sticky="w")
        self._taksit_var = tk.IntVar(value=3)
        tk.Spinbox(self._taksit_frame, from_=2, to=24,
            textvariable=self._taksit_var, width=6,
            bg=T["entry_bg"], fg=T["text"], insertbackground=T["gold"],
            font=("Segoe UI",10), relief="flat",
            buttonbackground=T["bg4"],
            command=self._onizle).pack(side="left", ipady=4)
        lbl(self._taksit_frame, "  ay", fg=T["text3"], bg=T["bg3"],
            font=("Segoe UI",9)).pack(side="left")

        # Önizleme
        self._onizle_lbl = lbl(fi, "", fg=T["gold2"], bg=T["bg3"],
                                font=("Segoe UI",9,"bold"))
        self._onizle_lbl.grid(row=11, column=0, columnspan=2, sticky="w", pady=(10,0))

        # ── Butonlar
        bf = frm(inner, bg=T["bg"])
        bf.pack(fill="x", padx=14, pady=12)
        btn(bf, "OK  Kaydet ve Daireye Ata", self._kaydet_uygula, "gold"
            ).pack(side="left", ipadx=14, ipady=6)
        btn(bf, "X  İptal",                self.destroy,        "gray"
            ).pack(side="left", padx=8, ipadx=10, ipady=6)

        self._odeme_degis()
        self._onizle()

    # ─── YARDIMCI ────────────────────────────────────────────────────────────
    def _kapsam_degis(self):
        if self._kapsam_var.get() == "secili":
            self._daire_frame.grid()
        else:
            self._daire_frame.grid_remove()
        self._hesapla()

    def _odeme_degis(self):
        if self._odeme_var.get() == "taksit":
            self._taksit_lbl.grid()
            self._taksit_frame.grid()
        else:
            self._taksit_lbl.grid_remove()
            self._taksit_frame.grid_remove()
        self._onizle()

    def _secili_daireler(self):
        """Seçili daire numaralarını döndür."""
        if self._kapsam_var.get() == "tum":
            return [d["no"] for d in self.data["daireler"]]
        sel = self._daire_lb.curselection()
        return [i + 1 for i in sel]

    def _hesapla(self, *_):
        """Toplam → daire başı hesapla."""
        if self._hesap_lock:
            return
        self._hesap_lock = True
        try:
            toplam = float(self._toplam_var.get())
        except:
            self._hesap_lock = False
            self._onizle()
            return
        daireler = self._secili_daireler()
        n = len(daireler) or 1
        db = round(toplam / n, 2)
        self._dbasI_var.set(f"{db:.2f}")
        self._hesap_lock = False
        self._onizle()

    def _hesapla_ters(self, *_):
        """Daire başı → toplam hesapla."""
        if self._hesap_lock:
            return
        self._hesap_lock = True
        try:
            db = float(self._dbasI_var.get())
        except:
            self._hesap_lock = False
            self._onizle()
            return
        daireler = self._secili_daireler()
        n = len(daireler) or 1
        toplam = round(db * n, 2)
        self._toplam_var.set(f"{toplam:.2f}")
        self._hesap_lock = False
        self._onizle()

    def _onizle(self, *_):
        taksitler = self._taksit_listesi()
        if not taksitler:
            self._onizle_lbl.config(text="")
            return
        daireler = self._secili_daireler()
        n = len(daireler)
        ilk = taksitler[0]
        son = taksitler[-1]
        if len(taksitler) == 1:
            msg = (f"→  {n} daire  •  Tek seferlik: {fmt(ilk['tutar'])}/daire  "
                   f"•  Ay: {ay_label(ilk['ay'])}")
        else:
            msg = (f"→  {n} daire  •  {len(taksitler)} ay  •  "
                   f"{fmt(ilk['tutar'])}/ay/daire  "
                   f"•  {ay_label(ilk['ay'])} – {ay_label(son['ay'])}")
        self._onizle_lbl.config(text=msg)

    def _taksit_listesi(self):
        """[{"ay":"YYYY-MM","tutar":X}, ...] listesini oluştur."""
        try:
            db = float(self._dbasI_var.get())
        except:
            return []
        if db <= 0:
            return []
        bas = self._bas_ay.get()
        if self._odeme_var.get() == "tek":
            return [{"ay": bas, "tutar": round(db, 2)}]
        else:
            try:
                n = int(self._taksit_var.get())
            except:
                n = 2
            n = max(2, min(n, 24))
            taksit_tut = round(db / n, 2)
            sonuclar = []
            for i in range(n):
                sonuclar.append({
                    "ay"   : _sonraki_ay(bas, i),
                    "tutar": taksit_tut,
                })
            toplam_yuv = round(taksit_tut * n, 2)
            fark = round(db - toplam_yuv, 2)
            if sonuclar and fark != 0:
                sonuclar[-1]["tutar"] = round(sonuclar[-1]["tutar"] + fark, 2)
            return sonuclar

    def _doldur(self, g):
        """Düzenleme modunda alanları doldur."""
        self._ad_var.set(g.get("ad", ""))
        self._acik_var.set(g.get("aciklama", ""))
        self._toplam_var.set(str(g.get("toplam", "")))
        self._dbasI_var.set(str(g.get("daire_basi", "")))
        kapsam = g.get("kapsam", "tum")
        self._kapsam_var.set(kapsam)
        self._kapsam_degis()
        if kapsam == "secili":
            dno_list = g.get("daire_nos", [])
            for i in range(self._daire_lb.size()):
                if i + 1 in dno_list:
                    self._daire_lb.selection_set(i)
        odeme = g.get("odeme_sekli", "tek")
        self._odeme_var.set(odeme)
        self._odeme_degis()
        if g.get("taksitler"):
            self._bas_ay.set(g["taksitler"][0]["ay"])
            self._taksit_var.set(len(g["taksitler"]))
        self._onizle()

    def _kaydet_kayit(self, uygula):
        """Ortak gider kaydını oluştur / güncelle ve isteğe bağlı olarak aidatlara uygula."""
        ad = self._ad_var.get().strip()
        if not ad:
            messagebox.showerror("Hata", "İş adı zorunludur.", parent=self)
            return
        taksitler = self._taksit_listesi()
        if not taksitler:
            messagebox.showerror("Hata", "Geçerli tutar ve başlangıç ayı girin.", parent=self)
            return
        daireler = self._secili_daireler()
        if not daireler:
            messagebox.showerror("Hata", "En az bir daire seçin.", parent=self)
            return
        try:
            toplam  = float(self._toplam_var.get())
            daire_b = float(self._dbasI_var.get())
        except:
            messagebox.showerror("Hata", "Geçerli tutar girin.", parent=self)
            return

        if self.duzenle:
            # Mevcut kaydı güncelle: önce eski aidatları geri al
            if self.duzenle.get("uygulandi"):
                _ortak_gider_aidat_geri_al(self.data, self.duzenle)
            g = self.duzenle
        else:
            g = {"id": yeni_id()}
            self.data["ortak_giderler"].append(g)

        g.update({
            "ad"          : ad,
            "aciklama"    : self._acik_var.get().strip(),
            "toplam"      : round(toplam, 2),
            "daire_basi"  : round(daire_b, 2),
            "kapsam"      : self._kapsam_var.get(),
            "daire_nos"   : daireler,
            "odeme_sekli" : self._odeme_var.get(),
            "taksitler"   : taksitler,
            "tarih"       : g.get("tarih", simdi()),
            "uygulandi"   : False,
            "eklenenler"  : [],
        })

        if uygula:
            _ortak_gider_aidata_uygula(self.data, g)

        # Ortak gider kaydını gider hesabına yansıt
        # Daha önce aynı id ile kayıt varsa güncelle, yoksa ekle
        mevcut_gider = next(
            (x for x in self.data["giderler"] if x.get("ortak_gider_id") == g["id"]),
            None
        )
        if mevcut_gider:
            mevcut_gider["tutar"]    = round(toplam, 2)
            mevcut_gider["aciklama"] = f"{ad}" + (f" — {self._acik_var.get().strip()}" if self._acik_var.get().strip() else "")
            mevcut_gider["tarih"]    = g["tarih"]
        else:
            self.data["giderler"].append({
                "id"             : yeni_id(),
                "tur"            : "Ortak Gider",
                "tarih"          : g["tarih"],
                "tutar"          : round(toplam, 2),
                "aciklama"       : f"{ad}" + (f" — {self._acik_var.get().strip()}" if self._acik_var.get().strip() else ""),
                "ortak_gider_id" : g["id"],
            })

        save_data(self.data)
        if self.callback:
            self.callback()
        self.destroy()
        mesaj = ("Gider kaydedildi ve dairelere atandı.\n"
                 "Tahsilat sırasında daire bazında ödenebilir.") if uygula else \
                "Gider kaydedildi (henüz dairelere atanmadı)."
        messagebox.showinfo("Tamam", mesaj)

    def _kaydet_uygula(self):
        self._kaydet_kayit(uygula=True)

    def _kaydet_sadece(self):
        self._kaydet_kayit(uygula=False)


# ══════════════════════════════════════════════════════════════════════════════
#  SAKİN LİSTESİ PENCERESİ
# ══════════════════════════════════════════════════════════════════════════════
class SakinListePencere(tk.Toplevel):
    """Tüm daire sakinlerini bir tabloda listeler; Excel/CSV dışa aktarımı sunar."""

    def __init__(self, parent, data):
        super().__init__(parent)
        self.data = data
        self.title("  Bina Sakin Listesi")
        self.configure(bg=T["bg"])
        self.geometry("820x560")
        self.minsize(640, 400)
        self.grab_set()
        self.update_idletasks()
        sw, sh = self.winfo_screenwidth(), self.winfo_screenheight()
        self.geometry(f"820x560+{(sw-820)//2}+{(sh-560)//2}")
        self._build()
        self._doldur()

    def _build(self):
        # Üst başlık çubuğu
        tk.Frame(self, bg=T["gold"], height=3).pack(fill="x")
        hdr = frm(self, bg=T["bg2"]); hdr.pack(fill="x")
        hf  = frm(hdr, bg=T["bg2"]); hf.pack(fill="x", padx=14, pady=10)

        lbl(hf, "  Bina Sakin Listesi",
            font=("Segoe UI", 13, "bold"), bg=T["bg2"]).pack(side="left")

        # Dışa aktarma butonu sağda
        bf = frm(hf, bg=T["bg2"]); bf.pack(side="right")
        btn(bf, "Grafik  Excel Kaydet", self._excel_kaydet, "green").pack(side="left", padx=4)
        btn(bf, "X  Kapat",      self.destroy,     "gray").pack(side="left", padx=4)

        tk.Frame(hdr, bg=T["border"], height=1).pack(fill="x")

        # Arama çubuğu
        ara_frm = frm(self, bg=T["bg3"],
            highlightthickness=1, highlightbackground=T["border"])
        ara_frm.pack(fill="x", padx=14, pady=(10,0))
        ara_ic = frm(ara_frm, bg=T["bg3"]); ara_ic.pack(fill="x", padx=12, pady=8)

        lbl(ara_ic, "Ara Ara:", fg=T["text2"], bg=T["bg3"],
            font=("Segoe UI", 9, "bold")).pack(side="left", padx=(0,6))
        self._ara_var = tk.StringVar()
        self._ara_var.trace_add("write", lambda *_: self._doldur())
        ent(ara_ic, textvariable=self._ara_var, width=30).pack(side="left", ipady=4)

        # Filtre: tüm / sakinli / boş
        self._filtre = tk.StringVar(value="tumu")
        lbl(ara_ic, "  Göster:", fg=T["text2"], bg=T["bg3"],
            font=("Segoe UI", 9, "bold")).pack(side="left", padx=(16,4))
        for val, txt in [("tumu","Tümü"), ("sakinli","Sakinli"), ("bos","Boş")]:
            tk.Radiobutton(ara_ic, text=txt, variable=self._filtre, value=val,
                bg=T["bg3"], fg=T["text"], selectcolor=T["bg4"],
                activebackground=T["bg3"], font=("Segoe UI", 9),
                cursor="hand2", command=self._doldur).pack(side="left", padx=4)

        # Tablo
        tbl_frm = frm(self); tbl_frm.pack(fill="both", expand=True, padx=14, pady=10)

        tbl_hdr = frm(tbl_frm, bg=T["bg2"]); tbl_hdr.pack(fill="x")
        tk.Frame(tbl_hdr, bg=T["gold"], height=2).pack(fill="x")
        lhf = frm(tbl_hdr, bg=T["bg2"]); lhf.pack(fill="x", padx=10, pady=6)
        lbl(lhf, "Liste  Sakin Tablosu",
            font=("Segoe UI", 10, "bold"), bg=T["bg2"]).pack(side="left")
        self._sayac_lbl = lbl(lhf, "",
            fg=T["text3"], bg=T["bg2"], font=("Segoe UI", 9))
        self._sayac_lbl.pack(side="left", padx=10)

        tf, self._tree = scrolled(tbl_frm, [
            ("Daire No",   80),
            ("Ad Soyad",  180),
            ("Telefon",   140),
            ("E-posta",   220),
            ("Durum",     110),
        ], height=16)
        tf.pack(fill="both", expand=True)

        self._tree.tag_configure("sakinli", foreground=T["green"])
        self._tree.tag_configure("bos",     foreground=T["text3"])

    def _doldur(self):
        self._tree.delete(*self._tree.get_children())
        ara   = self._ara_var.get().strip().lower()
        filtre = self._filtre.get()

        toplam = 0
        for d in sorted(self.data["daireler"], key=lambda x: x["no"]):
            isim  = d.get("isim",  "") or ""
            tel   = d.get("tel",   "") or ""
            email = d.get("email", "") or ""
            dolu  = bool(isim.strip())

            # Filtre uygula
            if filtre == "sakinli" and not dolu: continue
            if filtre == "bos"     and dolu:     continue

            # Arama uygula
            if ara and ara not in isim.lower() and ara not in tel.lower() \
                    and ara not in email.lower() \
                    and ara not in str(d["no"]):
                continue

            durum = "v Dolu" if dolu else "— Boş"
            tag   = "sakinli" if dolu else "bos"
            self._tree.insert("", "end",
                iid=str(d["no"]),
                tags=(tag,),
                values=(f"Daire {d['no']}", isim or "—", tel or "—",
                        email or "—", durum))
            toplam += 1

        dolu_say  = sum(1 for d in self.data["daireler"] if d.get("isim","").strip())
        bos_say   = len(self.data["daireler"]) - dolu_say
        self._sayac_lbl.config(
            text=f"  {toplam} kayıt gösteriliyor  •  "
                 f"v {dolu_say} dolu  •  — {bos_say} boş")

    def _excel_kaydet(self):
        import tkinter.filedialog as fd
        try:
            from openpyxl import Workbook
            from openpyxl.styles import (Font, PatternFill, Alignment,
                                         Border, Side)
        except ImportError:
            messagebox.showerror(
                "Eksik Kütüphane",
                "openpyxl kurulu değil.\n\nKurmak için:\n  pip install openpyxl",
                parent=self)
            return

        dosya = fd.asksaveasfilename(
            parent=self,
            defaultextension=".xlsx",
            filetypes=[("Excel Dosyası","*.xlsx"),("Tüm Dosyalar","*.*")],
            initialfile="sakin_listesi.xlsx",
            title="Sakin Listesini Excel Olarak Kaydet")
        if not dosya: return

        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Sakin Listesi"

            bina_adi = self.data["bina"].get("ad","Bina") or "Bina"

            # ── Renk paleti ───────────────────────────────────────────────
            ALTIN   = "C9A84C"
            KOYU_BG = "1E2535"
            YESIL   = "2ECC71"
            YESIL_BG= "1A3A2A"
            GRI_BG  = "252D3D"
            BEYAZ   = "E8EAF0"
            GRI_YAZ = "8A9AB5"

            ince = Side(style="thin",  color="2A3550")
            kalin= Side(style="medium",color=ALTIN)

            def dolu_dolgu(renk):
                return PatternFill("solid", start_color=renk, end_color=renk)

            def kenar(ust=None,alt=None,sol=None,sag=None):
                return Border(top=ust,bottom=alt,left=sol,right=sag)

            # ── Başlık satırı (1. satır) ──────────────────────────────────
            ws.merge_cells("A1:F1")
            bas = ws["A1"]
            bas.value       = f"Bina  {bina_adi}  —  Sakin Listesi"
            bas.font        = Font(name="Arial", bold=True, size=14, color=ALTIN)
            bas.fill        = dolu_dolgu(KOYU_BG)
            bas.alignment   = Alignment(horizontal="center", vertical="center")
            bas.border      = kenar(alt=Side(style="medium",color=ALTIN))
            ws.row_dimensions[1].height = 32

            # ── Tarih satırı (2. satır) ────────────────────────────────────
            ws.merge_cells("A2:F2")
            tarih_huc = ws["A2"]
            tarih_huc.value     = f"Oluşturma Tarihi: {bugun_str()}"
            tarih_huc.font      = Font(name="Arial", size=9, color=GRI_YAZ)
            tarih_huc.fill      = dolu_dolgu(KOYU_BG)
            tarih_huc.alignment = Alignment(horizontal="center")
            ws.row_dimensions[2].height = 18

            # ── Boş ayırıcı satır ─────────────────────────────────────────
            ws.row_dimensions[3].height = 6

            # ── Sütun başlıkları (4. satır) ───────────────────────────────
            basliklar = ["#","Daire No","Ad Soyad","Telefon","E-posta","Durum"]
            genislikler = [5, 12, 28, 18, 32, 12]
            for col, (bas_txt, gen) in enumerate(zip(basliklar, genislikler), 1):
                from openpyxl.utils import get_column_letter
                harf = get_column_letter(col)
                huc = ws.cell(row=4, column=col, value=bas_txt)
                huc.font      = Font(name="Arial", bold=True, size=10, color=ALTIN)
                huc.fill      = dolu_dolgu(GRI_BG)
                huc.alignment = Alignment(horizontal="center", vertical="center")
                huc.border    = kenar(
                    ust=Side(style="medium",color=ALTIN),
                    alt=Side(style="thin",  color=ALTIN),
                    sol=ince, sag=ince)
                ws.column_dimensions[harf].width = gen
            ws.row_dimensions[4].height = 24

            # ── Veri satırları ────────────────────────────────────────────
            daireler = sorted(self.data["daireler"], key=lambda x: x["no"])
            satir = 5
            for sira, d in enumerate(daireler, 1):
                isim  = d.get("isim","")  or ""
                tel   = d.get("tel","")   or ""
                email = d.get("email","") or ""
                dolu  = bool(isim.strip())
                durum = "Dolu" if dolu else "Boş"

                zebra = "161B27" if sira % 2 == 0 else "1A2030"
                dolgu = dolu_dolgu(YESIL_BG if dolu else zebra)
                yazi_renk = YESIL if dolu else "4A5A7A"

                degerler = [sira, f"Daire {d['no']}", isim, tel, email, durum]
                for col, val in enumerate(degerler, 1):
                    huc = ws.cell(row=satir, column=col, value=val)
                    huc.fill      = dolgu
                    huc.font      = Font(name="Arial", size=10,
                                         color=(yazi_renk if col > 1 else GRI_YAZ))
                    huc.alignment = Alignment(
                        horizontal="center" if col in (1,2,6) else "left",
                        vertical="center")
                    huc.border    = kenar(
                        alt=Side(style="thin", color="2A3550"),
                        sol=ince, sag=ince)
                ws.row_dimensions[satir].height = 22
                satir += 1

            # ── Özet satırı ───────────────────────────────────────────────
            satir += 1
            dolu_say = sum(1 for d in daireler if d.get("isim","").strip())
            bos_say  = len(daireler) - dolu_say

            ws.merge_cells(f"A{satir}:C{satir}")
            ozet_huc = ws[f"A{satir}"]
            ozet_huc.value     = (f"Toplam: {len(daireler)} daire  •  "
                                   f"Dolu: {dolu_say}  •  Boş: {bos_say}")
            ozet_huc.font      = Font(name="Arial", bold=True, size=9, color=ALTIN)
            ozet_huc.fill      = dolu_dolgu(KOYU_BG)
            ozet_huc.alignment = Alignment(horizontal="left", vertical="center")
            ozet_huc.border    = kenar(ust=Side(style="medium",color=ALTIN))
            for col in range(4, 7):
                huc = ws.cell(row=satir, column=col)
                huc.fill   = dolu_dolgu(KOYU_BG)
                huc.border = kenar(ust=Side(style="medium",color=ALTIN))
            ws.row_dimensions[satir].height = 20

            # ── Sayfa ayarları ────────────────────────────────────────────
            ws.sheet_view.showGridLines = False
            ws.freeze_panes = "A5"

            wb.save(dosya)
            messagebox.showinfo("Kaydedildi",
                f"Excel dosyası başarıyla kaydedildi:\n{dosya}", parent=self)
        except Exception as e:
            messagebox.showerror("Hata", f"Dosya kaydedilemedi:\n{e}", parent=self)


# ══════════════════════════════════════════════════════════════════════════════
#  AİDAT PERİYOT PENCERESİ — aylık/yıllık plan
# ══════════════════════════════════════════════════════════════════════════════
class AidatPeriyotPencere(tk.Toplevel):
    """
    Aylık veya yıllık periyotta aidat tutarı belirle.
    Kullanıcı başlangıç-bitiş aralığını ve tutarı girerek seçili aylara uygular.
    """
    def __init__(self, parent, data, callback=None):
        super().__init__(parent)
        self.data = data; self.callback = callback
        self.title("Takvim  Aylık / Yıllık Aidat Planı")
        self.configure(bg=T["bg"])
        self.geometry("760x600"); self.minsize(660,500)
        self.grab_set()
        self.update_idletasks()
        sw,sh=self.winfo_screenwidth(),self.winfo_screenheight()
        self.geometry(f"760x600+{(sw-760)//2}+{(sh-600)//2}")
        self._build()
        self._onizle()

    def _build(self):
        b=self.data["bina"]
        tk.Frame(self,bg=T["gold"],height=3).pack(fill="x")

        hdr=frm(self,bg=T["bg2"]); hdr.pack(fill="x")
        lf=frm(hdr,bg=T["bg2"]); lf.pack(fill="x",padx=14,pady=10)
        lbl(lf,"Takvim  Aidat Periyot Planı",
            font=("Segoe UI",13,"bold"),bg=T["bg2"]).pack(side="left")
        lbl(lf,f"  Mevcut varsayılan: {fmt(b['aidat'])}",
            fg=T["text3"],bg=T["bg2"],font=("Segoe UI",9)).pack(side="left",padx=8)
        tk.Frame(hdr,bg=T["border"],height=1).pack(fill="x")

        # Form
        form_wrap=frm(self,bg=T["bg3"],
            highlightthickness=1,highlightbackground=T["border"])
        form_wrap.pack(fill="x",padx=14,pady=10)
        form=frm(form_wrap,bg=T["bg3"]); form.pack(fill="x",padx=14,pady=12)
        form.columnconfigure(1,weight=1)

        # Periyot tipi
        lbl(form,"Periyot Tipi:",fg=T["text2"],bg=T["bg3"],
            font=("Segoe UI",9,"bold")).grid(row=0,column=0,sticky="w",pady=4,padx=(0,12))
        self._periyot=tk.StringVar(value="aylik")
        pf=frm(form,bg=T["bg3"]); pf.grid(row=0,column=1,sticky="w")
        for val,txt in [("aylik","Aylık  (tek ay seç)"),
                        ("yillik","Yıllık  (tüm yılı seç)"),
                        ("aralik","Aralık  (başlangıç–bitiş)")]:
            tk.Radiobutton(pf,text=txt,variable=self._periyot,value=val,
                bg=T["bg3"],fg=T["text"],selectcolor=T["bg4"],
                activebackground=T["bg3"],font=("Segoe UI",9),
                cursor="hand2",command=self._periyot_degis).pack(side="left",padx=6)

        # Başlangıç ayı
        lbl(form,"Başlangıç:",fg=T["text2"],bg=T["bg3"],
            font=("Segoe UI",9,"bold")).grid(row=1,column=0,sticky="w",pady=8,padx=(0,12))
        self._bas=AySecici(form,bg=T["bg3"],callback=lambda _:self._onizle())
        self._bas.grid(row=1,column=1,sticky="w")

        # Bitiş ayı (aralık modunda görünür)
        self._bitis_lbl=lbl(form,"Bitiş:",fg=T["text2"],bg=T["bg3"],
            font=("Segoe UI",9,"bold"))
        self._bitis_lbl.grid(row=2,column=0,sticky="w",pady=4,padx=(0,12))
        self._son=AySecici(form,bg=T["bg3"],callback=lambda _:self._onizle())
        self._son.grid(row=2,column=1,sticky="w")

        # Tutar
        lbl(form,"Aidat Tutarı (₺):",fg=T["text2"],bg=T["bg3"],
            font=("Segoe UI",9,"bold")).grid(row=3,column=0,sticky="w",pady=8,padx=(0,12))
        tf2=frm(form,bg=T["bg3"]); tf2.grid(row=3,column=1,sticky="w")
        self._tutar=tk.StringVar(value=str(b["aidat"]))
        ent(tf2,textvariable=self._tutar,width=14).pack(side="left",ipady=6)
        self._tutar.trace_add("write",lambda *_:self._onizle())

        # Hızlı tutar
        for etiket,deger in [("Mevcut",b["aidat"]),("+%5",round(b["aidat"]*1.05)),
                              ("+%10",round(b["aidat"]*1.10)),("+%20",round(b["aidat"]*1.20))]:
            d=deger
            btn(tf2,etiket,lambda d2=d:self._tutar.set(str(d2)),"gray").pack(side="left",padx=3)

        # Üzerine yazma kuralı
        lbl(form,"Kural:",fg=T["text2"],bg=T["bg3"],
            font=("Segoe UI",9,"bold")).grid(row=4,column=0,sticky="w",pady=4,padx=(0,12))
        self._kapsam=tk.StringVar(value="bos")
        kf=frm(form,bg=T["bg3"]); kf.grid(row=4,column=1,sticky="w")
        for val,txt in [("bos","Sadece boş ayları güncelle"),
                        ("hepsi","Tüm seçili ayları güncelle")]:
            tk.Radiobutton(kf,text=txt,variable=self._kapsam,value=val,
                bg=T["bg3"],fg=T["text"],selectcolor=T["bg4"],
                activebackground=T["bg3"],font=("Segoe UI",9),
                cursor="hand2",command=self._onizle).pack(side="left",padx=6)

        # Önizleme
        self._onizle_lbl=lbl(form,"",fg=T["gold2"],bg=T["bg3"],
            font=("Segoe UI",9,"bold"))
        self._onizle_lbl.grid(row=5,column=0,columnspan=2,sticky="w",pady=(10,0))

        # Butonlar
        bf=frm(form,bg=T["bg3"]); bf.grid(row=6,column=0,columnspan=2,sticky="w",pady=(12,0))
        btn(bf,"OK  Uygula",self._uygula,"gold").pack(side="left",ipadx=14,ipady=5)
        btn(bf,"X  Kapat",self.destroy,"gray").pack(side="left",padx=8,ipadx=10,ipady=5)

        # Tablo — mevcut özel aidatlar
        tbl_wrap=frm(self); tbl_wrap.pack(fill="both",expand=True,padx=14,pady=(0,12))
        t_hdr=frm(tbl_wrap,bg=T["bg2"]); t_hdr.pack(fill="x")
        tk.Frame(t_hdr,bg=T["gold"],height=2).pack(fill="x")
        lbl(frm(t_hdr,bg=T["bg2"]).pack(fill="x",padx=10,pady=6) or t_hdr,
            "Liste  Mevcut Özel Aidat Planı",
            font=("Segoe UI",10,"bold"),bg=T["bg2"]).pack(anchor="w",padx=10,pady=6)
        tf3,self._tree=scrolled(tbl_wrap,[
            ("Dönem",150),("Tür",120),("Tutar",120),("Varsayılan",130),("Fark",100)],height=8)
        tf3.pack(fill="both",expand=True)
        self._tablo_doldur()

        self._periyot_degis()

    def _periyot_degis(self):
        mod=self._periyot.get()
        if mod=="aralik":
            self._bitis_lbl.grid()
            self._son.grid()
        else:
            self._bitis_lbl.grid_remove()
            self._son.grid_remove()
        self._onizle()

    def _hedef_aylar(self):
        """Periyot moduna göre uygulanacak ayları döndür."""
        mod=self._periyot.get()
        bas=self._bas.get()
        if mod=="aylik":
            return [bas]
        elif mod=="yillik":
            y=int(bas[:4])
            return [f"{y}-{m:02d}" for m in range(1,13)]
        else:  # aralik
            son=self._son.get()
            if son<bas: son=bas
            return ay_listesi(bas,son)

    def _onizle(self):
        try:
            tutar=float(self._tutar.get())
        except:
            self._onizle_lbl.config(text="Uyari Geçersiz tutar"); return
        aylar=self._hedef_aylar()
        ozel=self.data["bina"].get("ozel_aidatlar",{})
        kapsam=self._kapsam.get()
        degisecek=[a for a in aylar if kapsam=="hepsi" or a not in ozel]
        self._onizle_lbl.config(
            text=f"→  {len(degisecek)} ay etkilenecek  •  Toplam: {fmt(len(degisecek)*tutar)}")

    def _uygula(self):
        try:
            tutar=float(self._tutar.get())
        except:
            messagebox.showerror("Hata","Geçerli tutar girin.",parent=self); return
        aylar=self._hedef_aylar()
        if not aylar:
            messagebox.showwarning("Uyarı","Uygulanacak ay bulunamadı.",parent=self); return
        ozel=self.data["bina"].setdefault("ozel_aidatlar",{})
        kapsam=self._kapsam.get()
        degisen=[]
        for ay in aylar:
            if kapsam=="hepsi" or ay not in ozel:
                ozel[ay]=tutar; degisen.append(ay)
        save_data(self.data)
        self._tablo_doldur()
        if self.callback: self.callback()
        messagebox.showinfo("Uygulandı",
            f"{len(degisen)} aya {fmt(tutar)} uygulandı.",parent=self)

    def _tablo_doldur(self):
        self._tree.delete(*self._tree.get_children())
        b=self.data["bina"]
        ozel=b.get("ozel_aidatlar",{})
        bas=b.get("baslangic",buay())
        today=buay()
        ty,tm=int(today[:4]),int(today[5:])
        tm+=24; ty+=tm//12; tm=tm%12 or 12
        bitis=f"{ty}-{tm:02d}"
        tum=ay_listesi(bas,bitis)
        for ay in reversed(tum):
            ov=ozel.get(ay)
            if ov is None: continue
            durum="o İleri" if ay>today else "* Özel"
            fark=round(ov-b["aidat"],2)
            fs=(f"+{fmt(fark)}" if fark>0 else fmt(fark)) if fark!=0 else "—"
            self._tree.insert("","end",iid=ay,tags=("ozel",),
                values=(ay_label(ay),durum,fmt(ov),fmt(b["aidat"]),fs))

# ─── BAŞLAT ───────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    app = App()
    app.mainloop()
